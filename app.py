from __future__ import annotations

import io
import json
import re
import zipfile
from dataclasses import dataclass, asdict
from datetime import date, datetime
from difflib import SequenceMatcher
from typing import Any, Dict, List, Optional, Tuple
import xml.etree.ElementTree as ET

import pandas as pd
import pymupdf as fitz
import requests
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from dateutil import parser as dateparser
from dateutil.relativedelta import relativedelta

APP_TITLE = "HVPQ / PIQ / Q88 / Class Status Checker v19"
APP_SUBTITLE = "Extraction-first verifier with built-in machine-readable observation priority library and validation rules. No external rules/observation uploads required."

# ----------------------------- Data models -----------------------------

@dataclass
class FieldRecord:
    source: str
    field_id: str
    label: str
    value: str = ""
    date_value: str = ""
    confidence: str = "deterministic"
    raw: str = ""

@dataclass
class Finding:
    area: str
    check: str
    status: str
    risk: str
    hvpq_value: str = ""
    piq_value: str = ""
    class_value: str = ""
    q88_value: str = ""
    xml_value: str = ""
    reason: str = ""
    action: str = ""

# ----------------------------- General helpers -----------------------------

MONTHS_7_DAYS = 7 * 30.4375
MONTHS_12_DAYS = 12 * 30.4375
DATE_PATTERNS = [
    r"\b\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4}\b",
    r"\b[A-Za-z]{3,9}\s+\d{1,2},\s*\d{4}\b",
    r"\b\d{4}-\d{2}-\d{2}\b",
    r"\b\d{1,2}[./-]\d{1,2}[./-]\d{2,4}\b",
]
DATE_RE = re.compile("|".join(DATE_PATTERNS), re.I)

CERT_ALIASES = {
    "safety equipment certificate": "safety_equipment",
    "sec": "safety_equipment",
    "safety radio certificate": "safety_radio",
    "src": "safety_radio",
    "safety construction certificate": "safety_construction",
    "scc": "safety_construction",
    "loadline certificate": "loadline",
    "load line": "loadline",
    "international loadline certificate": "loadline",
    "international oil pollution prevention": "iopp",
    "oil pollution prevention": "iopp",
    "ioppc": "iopp",
    "opp": "iopp",
    "international ballast water management": "bwm",
    "ballast water management": "bwm",
    "ibwmc": "bwm",
    "safety management certificate": "smc",
    "ism safety management certificate": "smc",
    "smc": "smc",
    "document of compliance": "doc",
    "doc": "doc",
    "international ship security": "issc",
    "issc": "issc",
    "uscg certificate of compliance": "uscg_coc",
    "us certificate of compliance": "uscg_coc",
    "uscgcoc": "uscg_coc",
    "civil liability convention certificate": "clc_oil",
    "civil liability certificates": "clc_oil",
    "civil liability convention 1992": "clc_oil",
    "bunker oil pollution": "clc_bunker",
    "bunker": "clc_bunker",
    "wreck removal": "wreck_removal",
    "wrc": "wreck_removal",
    "financial responsibility": "cofr",
    "cofr": "cofr",
    "vessel general permit": "vgp",
    "certificate of fitness": "cof_chemical",
    "cof": "cof_chemical",
    "certificate of class": "class_certificate",
    "class certificate": "class_certificate",
    "international tonnage": "tonnage",
    "international air pollution prevention": "iapp",
    "iapp": "iapp",
    "air pollution prevention": "iapp",
    "international sewage pollution prevention": "ispp",
    "ispp": "ispp",
    "maritime labour certificate": "mlc",
    "mlc": "mlc",
    "ship sanitation": "ship_sanitation",
}

FIELD_LABELS = {
    "vessel.name": "Vessel name",
    "vessel.imo": "IMO number",
    "vessel.flag": "Flag",
    "vessel.port_registry": "Port of registry",
    "vessel.type": "Vessel type",
    "vessel.call_sign": "Call sign",
    "vessel.mmsi": "MMSI",
    "owner.registered_owner": "Registered owner",
    "owner.technical_operator": "Technical operator",
    "owner.commercial_operator": "Commercial operator",
    "insurance.pni_club": "P&I club",
    "classification.class_society": "Class society",
    "classification.class_notation": "Class notation",
    "classification.conditions_of_class": "Conditions of class",
    "classification.memo_of_class": "Memoranda of class",
    "classification.flag_dispensation": "Flag/Class dispensation",
    "surveys.last_drydock": "Last dry dock",
    "surveys.next_drydock_due": "Next dry dock due",
    "surveys.last_iws": "Last IWS",
    "surveys.next_iws_due": "Next IWS due",
    "surveys.last_special": "Last special survey",
    "surveys.next_special_due": "Next special survey due",
    "surveys.last_annual": "Last annual survey",
    "surveys.next_annual_due": "Next annual survey due",
    "surveys.last_intermediate": "Last intermediate survey",
    "environment.cii_rating": "CII rating",
    "environment.cii_verified_by": "CII verification basis",
    "environment.eexi_rating": "EEXI rating",
    "environment.eexi_verified_by": "EEXI verification basis",
    "incidents.pollution_grounding_collision_allision": "Pollution/grounding/collision/allision incident",
    "incidents.other_incidents": "Other incidents",
    "psc.last_date": "Last PSC date",
    "psc.last_port": "Last PSC port",
    "psc.detained_36m": "Detained last 36 months",
    "moc.retrofit": "Equipment retrofitted",
    "moc.structural_change": "Structural change",
    "moc.equipment_replaced": "Equipment replaced non-like-for-like",
}


def clean_text(s: Any) -> str:
    if s is None:
        return ""
    s = str(s).replace("\uFFFE", " ").replace("\u0000", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_key(s: str) -> str:
    s = clean_text(s).lower()
    s = s.replace("&", "and")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, normalize_key(a), normalize_key(b)).ratio()


def normalize_value(s: str) -> str:
    s = clean_text(s).lower()
    s = s.replace("limited", "ltd").replace("limitied", "ltd").replace("co.,", "co")
    s = s.replace("n/a", "na").replace("not applicable", "na")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def normalize_bool(s: str) -> str:
    x = normalize_value(s)
    if x in ["yes", "y", "true", "1"]:
        return "yes"
    if x in ["no", "n", "false", "0", "nil", "none"]:
        return "no"
    if x in ["na", "not applicable", "blank", ""]:
        return ""
    return x


def parse_date_any(s: str) -> Optional[date]:
    if not s:
        return None
    # Avoid question numbers / decimals being parsed as dates
    s = clean_text(s)
    m = DATE_RE.search(s)
    target = m.group(0) if m else s
    try:
        dt = dateparser.parse(target, dayfirst=False, fuzzy=True)
        if dt and 1900 <= dt.year <= 2100:
            return dt.date()
    except Exception:
        pass
    return None


def iso_date(s: str) -> str:
    d = parse_date_any(s)
    return d.isoformat() if d else ""


def extract_dates(line: str) -> List[str]:
    out = []
    for m in DATE_RE.finditer(line or ""):
        d = parse_date_any(m.group(0))
        if d:
            out.append(d.isoformat())
    return out


def add_field(fields: List[FieldRecord], source: str, field_id: str, value: Any, label: str = "", raw: str = "", confidence: str = "deterministic"):
    val = clean_text(value)
    if val == "":
        return
    fields.append(FieldRecord(source=source, field_id=field_id, label=label or FIELD_LABELS.get(field_id, field_id), value=val, date_value=iso_date(val), raw=clean_text(raw)[:800], confidence=confidence))




def add_structured_date_list(fields: List[FieldRecord], source: str, field_id: str, dates: List[date], label: str, raw: str = "", confidence: str = "table-aware"):
    """Store structured list-style date extraction: oldest/newest/count plus the readable list."""
    if not dates:
        return
    unique = sorted(set(dates))
    value = "; ".join(d.isoformat() for d in unique)
    add_field(fields, source, field_id, value, label=label, raw=raw, confidence=confidence)
    add_field(fields, source, field_id + ".oldest", unique[0].isoformat(), label=label + " - oldest", raw=raw, confidence=confidence)
    add_field(fields, source, field_id + ".newest", unique[-1].isoformat(), label=label + " - newest", raw=raw, confidence=confidence)
    add_field(fields, source, field_id + ".count", str(len(unique)), label=label + " - count", raw=raw, confidence=confidence)


def parse_table_dates_by_last_inspection(section: str) -> List[date]:
    """Return inspection dates from HVPQ/Q88 tank coating tables.

    The table normally has two dates per row: original coating date and last
    inspection date. The rule must use the SECOND date before Annual/frequency,
    not the original coating date.
    """
    if not section:
        return []
    txt = clean_text(section)
    date_pat = r"(\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4}|[A-Za-z]{3,9}\s+\d{1,2},\s*\d{4})"
    pairs = re.findall(date_pat + r"\s+" + date_pat + r"\s+(?:Annual|12\s*months?|[0-9]+\s*months?)", txt, flags=re.I)
    out=[]
    for _, insp in pairs:
        d=parse_date_any(insp)
        if d:
            out.append(d)
    return out


def extract_hvpq_tank_coating_fields(fields: List[FieldRecord], source: str, text: str):
    """Extract HVPQ tank coating inspection dates table-aware."""
    cargo_sec = section_text_by_qid(text, "7.1.1")
    ballast_sec = section_text_by_qid(text, "7.1.3")
    if cargo_sec:
        dates = parse_table_dates_by_last_inspection(cargo_sec)
        if dates:
            add_structured_date_list(fields, source, "tank.cargo_hvpq.last_inspection_dates", dates, "Cargo/slop tank last coating inspection dates", raw=cargo_sec)
            add_field(fields, source, "tank.cargo_hvpq.freq_months", "12", label="Cargo/slop tank coating inspection frequency", raw=cargo_sec, confidence="table-aware")
        else:
            add_field(fields, source, "tank.cargo_hvpq.extraction_issue", "Cargo tank table found but last inspection dates were not reliably parsed", label="Cargo tank coating extraction issue", raw=cargo_sec, confidence="manual-needed")
    if ballast_sec:
        dates = parse_table_dates_by_last_inspection(ballast_sec)
        if dates:
            add_structured_date_list(fields, source, "tank.ballast_hvpq.last_inspection_dates", dates, "Ballast tank last coating inspection dates", raw=ballast_sec)
            add_field(fields, source, "tank.ballast_hvpq.freq_months", "12", label="Ballast tank coating inspection frequency", raw=ballast_sec, confidence="table-aware")
        else:
            add_field(fields, source, "tank.ballast_hvpq.extraction_issue", "Ballast tank table found but last inspection dates were not reliably parsed", label="Ballast tank coating extraction issue", raw=ballast_sec, confidence="manual-needed")


def extract_q88_coating_fields(fields: List[FieldRecord], source: str, text: str):
    """Extract Q88 coating inspection dates as a value-add cross-check."""
    cargo_m = re.search(r"Cargo tanks:(.*?)(?:Ballast tanks:|Tank anodes|7\.|8\.|Inert Gas|Cargo Pumps|$)", text, flags=re.I|re.S)
    ballast_m = re.search(r"Ballast tanks:(.*?)(?:Tank anodes|7\.|8\.|Inert Gas|Cargo Pumps|$)", text, flags=re.I|re.S)
    for name, fid, m in [("Cargo/slop tank", "tank.cargo_q88.last_inspection_dates", cargo_m), ("Ballast tank", "tank.ballast_q88.last_inspection_dates", ballast_m)]:
        if not m:
            continue
        sec=m.group(1)
        dates=parse_table_dates_by_last_inspection(sec)
        if dates:
            add_structured_date_list(fields, source, fid, dates, f"Q88 {name} coating inspection dates", raw=sec, confidence="table-aware")

def first_field(fields: List[FieldRecord], source: str, field_id: str) -> str:
    for f in fields:
        if f.source == source and f.field_id == field_id and clean_text(f.value):
            return f.value
    return ""


def values_by_field(fields: List[FieldRecord], source: str) -> Dict[str, str]:
    d = {}
    for f in fields:
        if f.source == source and f.field_id not in d and clean_text(f.value):
            d[f.field_id] = f.value
    return d


def sources_value(fields: List[FieldRecord], field_id: str) -> Dict[str, str]:
    out = {}
    for f in fields:
        if f.field_id == field_id and clean_text(f.value):
            out.setdefault(f.source, f.value)
    return out


def cert_key_from_label(label: str) -> Optional[str]:
    nk = normalize_key(label)
    if "garbage pollution" in nk:
        return None
    # longer aliases first to avoid bunker matching everything
    for alias, key in sorted(CERT_ALIASES.items(), key=lambda kv: len(kv[0]), reverse=True):
        if normalize_key(alias) in nk:
            return key
    return None


def semantically_equivalent(a: str, b: str, field_id: str = "") -> bool:
    if not clean_text(a) or not clean_text(b):
        return False
    da, db = parse_date_any(a), parse_date_any(b)
    if da and db:
        return da == db
    na, nb = normalize_bool(a), normalize_bool(b)
    if na and nb and na == nb:
        return True
    va, vb = normalize_value(a), normalize_value(b)
    if not va or not vb:
        return False
    if va == vb:
        return True
    # Class aliases
    if field_id.endswith("class_society"):
        class_aliases = [
            ("korean register", "kr"), ("nippon kaiji kyokai", "class nk"), ("nippon kaiji kyokai", "nk"),
            ("dnv gl", "dnv"), ("american bureau of shipping", "abs"), ("bureau veritas", "bv"),
        ]
        for x, y in class_aliases:
            if (x in va and y in vb) or (x in vb and y in va):
                return True
    # Vessel type should be broad equivalence, not hard mismatch
    if field_id.endswith("vessel.type"):
        oil_words_a = any(w in va for w in ["oil", "crude", "product", "chemical", "tanker", "carrier"])
        oil_words_b = any(w in vb for w in ["oil", "crude", "product", "chemical", "tanker", "carrier"])
        if oil_words_a and oil_words_b:
            return True
    # P&I / owner fuzzy equivalence
    if similarity(va, vb) >= 0.88:
        return True
    return False

# ----------------------------- PDF/Text extraction -----------------------------

def extract_pdf_pages(file_obj) -> List[Tuple[int, str]]:
    if file_obj is None:
        return []
    data = file_obj.getvalue() if hasattr(file_obj, "getvalue") else file_obj.read()
    pages = []
    try:
        with fitz.open(stream=data, filetype="pdf") as doc:
            for i, page in enumerate(doc, 1):
                txt = page.get_text("text") or ""
                pages.append((i, txt))
    except Exception as e:
        st.warning(f"PDF extraction failed: {e}")
    return pages


def join_pages(pages: List[Tuple[int, str]]) -> str:
    return "\n".join([f"\n--- PAGE {p} ---\n{t}" for p, t in pages])


def lines_from_text(text: str) -> List[str]:
    return [clean_text(x) for x in text.splitlines() if clean_text(x)]


def window_after(lines: List[str], pattern: str, n: int = 8) -> str:
    rx = re.compile(pattern, re.I)
    for i, line in enumerate(lines):
        if rx.search(line):
            return " ".join(lines[i:i+n])
    return ""


def find_value_after_label(lines: List[str], label_regex: str, value_regex: str = r"(.+)", max_lines: int = 3) -> str:
    rx = re.compile(label_regex, re.I)
    for i, line in enumerate(lines):
        if rx.search(line):
            # Same line after label
            m = rx.search(line)
            rest = clean_text(line[m.end():])
            if rest:
                vm = re.search(value_regex, rest, re.I)
                if vm:
                    return clean_text(vm.group(1) if vm.groups() else vm.group(0))
            for j in range(1, max_lines + 1):
                if i + j < len(lines):
                    cand = clean_text(lines[i + j])
                    if cand and not re.match(r"^\d+(\.\d+)*\.?$", cand):
                        vm = re.search(value_regex, cand, re.I)
                        if vm:
                            return clean_text(vm.group(1) if vm.groups() else vm.group(0))
    return ""


def next_value_after(lines: List[str], label_regex: str, max_lines: int = 6) -> str:
    """Return first plausible value after a label line. Designed for HVPQ/Q88 line-per-cell PDFs."""
    rx = re.compile(label_regex, re.I)
    skip = re.compile(r"^(\d+|\d+(\.\d+)+\.?|yes/no|date issued|date expires|last annual|last intermediate|date of|endorsement)$", re.I)
    label_like = re.compile(r"^(general information|certificates|classification|ownership and operation|environmental information|publications|special survey|dry dock|in water survey|condition assessment|date of|name$|flag$|type of|if other|what is|does |has |is )", re.I)
    for i, line in enumerate(lines):
        if rx.search(line):
            # If value exists after colon on same line, take it.
            if ':' in line:
                tail = clean_text(line.split(':', 1)[1])
                if tail and not rx.fullmatch(tail):
                    return tail
            for j in range(1, max_lines + 1):
                if i + j >= len(lines):
                    break
                cand = clean_text(lines[i + j])
                if not cand:
                    continue
                if re.fullmatch(r"\d+(\.\d+)+\.?", cand):
                    break
                if skip.match(cand):
                    continue
                # allow No/Yes/NA as values
                if cand.lower() in ['yes', 'no', 'na', 'n/a', 'not applicable', 'nil', 'none']:
                    return cand
                # skip obvious labels/question text unless it has a date/number/value after same line
                if label_like.match(cand) and not DATE_RE.search(cand):
                    continue
                return cand
    return ''

def q88_blocks(lines: List[str]) -> Dict[str, List[str]]:
    blocks: Dict[str, List[str]] = {}
    current = None
    for line in lines:
        txt = line.strip()
        m = re.match(r"^(\d+\.\d+[a-z]?)(?:\s+(.*))?$", txt, re.I)
        if m:
            current = m.group(1).lower()
            blocks[current] = []
            rest = clean_text(m.group(2) or "")
            if rest:
                blocks[current].append(rest)
            continue
        if current:
            blocks[current].append(line)
    return blocks

def block_answer(block: List[str], skip_label_lines: int = 1, max_answer_lines: int = 8) -> str:
    vals = []
    # Usually first line is question label, answer begins after it.
    for cand in block[skip_label_lines:]:
        c = clean_text(cand)
        if not c:
            continue
        if re.match(r"^(Tel:|Fax:|Telex:|Email:|Web:|IMO:|Company IMO#)", c, re.I) and vals:
            break
        vals.append(c)
        if len(vals) >= max_answer_lines:
            break
    return clean_text(" ".join(vals))


# Class Status date-label safeguards.
# DNV Class Status normally uses "Issued date" for certificate issue and "Valid until" for expiry.
# Never treat a lone "Issued date" as an expiry. Expiry is accepted only when an expiry/validity label is present
# or when a clearly labelled table provides both issue and valid-until dates.
EXPIRY_LABEL_RE = re.compile(r"\b(valid\s*(?:until|to)|validity|expiry|expires?|expiration|date\s*expires|valid\s*until)\b", re.I)
ISSUE_LABEL_RE = re.compile(r"\b(issued?|date\s*issued|issue\s*date|issued\s*date)\b", re.I)

def has_expiry_label(s: str) -> bool:
    return bool(EXPIRY_LABEL_RE.search(clean_text(s)))

def has_issue_label(s: str) -> bool:
    return bool(ISSUE_LABEL_RE.search(clean_text(s)))

def first_date_after_label(text: str, label_re: re.Pattern, max_chars: int = 140) -> str:
    txt = clean_text(text)
    for m in label_re.finditer(txt):
        seg = txt[m.end():m.end()+max_chars]
        dates = extract_dates(seg)
        if dates:
            return dates[0]
    return ""

def labelled_cert_dates(block: str) -> Tuple[str, str]:
    """Return (issue_date, expiry_date) only when labels support the classification.

    This is intentionally conservative for Class Status PDFs. A date near 'Issued date' is issue.
    A date near 'Valid until'/'Expiry' is expiry. A single unlabelled date is not expiry.
    """
    txt = clean_text(block)
    issue = first_date_after_label(txt, ISSUE_LABEL_RE)
    expiry = first_date_after_label(txt, EXPIRY_LABEL_RE)
    if issue or expiry:
        return issue, expiry

    # Fallback for one-line/table rows where labels and dates are both present but OCR order is imperfect.
    dates = extract_dates(txt)
    if has_issue_label(txt) and len(dates) == 1:
        return dates[0], ""
    if has_expiry_label(txt) and len(dates) == 1:
        return "", dates[0]
    if has_issue_label(txt) and has_expiry_label(txt) and len(dates) >= 2:
        return dates[0], dates[-1]
    return "", ""


def add_q88_block_field(fields: List[FieldRecord], blocks: Dict[str, List[str]], qno: str, source: str, field_id: str, label: str = ''):
    b = blocks.get(qno.lower(), [])
    if not b:
        return
    ans = block_answer(b)
    add_field(fields, source, field_id, ans, label=label, raw=" | ".join(b))

def parse_cert_rows_from_sequence(lines: List[str], source: str, cert_order: str = 'hvpq') -> List[FieldRecord]:
    """Parse certificate rows by partitioning between certificate labels, avoiding bleed into next row."""
    fields: List[FieldRecord] = []
    positions: List[Tuple[int, str]] = []
    for i, line in enumerate(lines):
        key = cert_key_from_label(line)
        if not key and re.search(r"certificate|u\.s\.|international|civil|safety|document|vessel general|wreck|financial", line, re.I):
            key = cert_key_from_label(line + " " + (lines[i+1] if i+1 < len(lines) else ""))
        if not key:
            continue
        # avoid generic headers where possible
        if normalize_key(line) in ["certificates", "certificate dates", "certificate description"]:
            continue
        positions.append((i, key))
    for idx, (pos, key) in enumerate(positions):
        next_pos = positions[idx + 1][0] if idx + 1 < len(positions) else min(len(lines), pos + 8)
        block_lines = lines[pos:next_pos]
        block = " ".join(block_lines)
        dates = extract_dates(block)
        if not dates:
            continue
        # Special rows with issue date only
        if key in ["tonnage"]:
            add_field(fields, source, f"cert.{key}.issue", dates[0], raw=block)
            continue
        # VGP in HVPQ often has issue date and an old second date, but Class Status usually has issue only. Keep both for review.
        if cert_order == 'q88':
            if len(dates) >= 1: add_field(fields, source, f"cert.{key}.issue", dates[0], raw=block)
            if len(dates) >= 4: add_field(fields, source, f"cert.{key}.expiry", dates[3], raw=block)
            elif len(dates) >= 2: add_field(fields, source, f"cert.{key}.expiry", dates[-1], raw=block)
            if len(dates) >= 2: add_field(fields, source, f"cert.{key}.last_annual", dates[1], raw=block)
            if len(dates) >= 3: add_field(fields, source, f"cert.{key}.last_intermediate", dates[2], raw=block)
        elif cert_order == 'class':
            issue_dt, expiry_dt = labelled_cert_dates(block)
            if issue_dt:
                add_field(fields, source, f"cert.{key}.issue", issue_dt, raw=block)
            elif len(dates) >= 1 and not has_expiry_label(block):
                # Conservative fallback: an unlabelled/lone Class Status date is issue/review data, not expiry.
                add_field(fields, source, f"cert.{key}.issue", dates[0], raw=block)
            if expiry_dt:
                add_field(fields, source, f"cert.{key}.expiry", expiry_dt, raw=block)
            elif has_expiry_label(block) and len(dates) >= 2:
                add_field(fields, source, f"cert.{key}.expiry", dates[-1], raw=block)
        else:
            if len(dates) >= 1: add_field(fields, source, f"cert.{key}.issue", dates[0], raw=block)
            if len(dates) >= 2: add_field(fields, source, f"cert.{key}.expiry", dates[1], raw=block)
            if len(dates) >= 3: add_field(fields, source, f"cert.{key}.last_annual", dates[2], raw=block)
            if len(dates) >= 4: add_field(fields, source, f"cert.{key}.last_intermediate", dates[3], raw=block)
    return fields

# ----------------------------- Document-specific extractors -----------------------------

def extract_hvpq(pages: List[Tuple[int, str]]) -> List[FieldRecord]:
    source = "HVPQ"
    text = join_pages(pages)
    lines = lines_from_text(text)
    fields: List[FieldRecord] = []

    # Section-aware direct regex fixes for HVPQ line-per-cell layout
    m_owner = re.search(r"1\.3\.1\s+Registered Owner.*?\bName\s+([^\n]+)", text, re.I | re.S)
    if m_owner: add_field(fields, source, "owner.registered_owner", m_owner.group(1), raw=m_owner.group(0)[:500])
    m_pni = re.search(r"1\.1\.13\s+P and I Club.*?If other, then specify\s+([^\n]+)", text, re.I | re.S)
    if m_pni: add_field(fields, source, "insurance.pni_club", m_pni.group(1), raw=m_pni.group(0)[:500])
    m_mmsi = re.search(r"MMSI\) number\?\s+(\d{6,12})", text, re.I)
    if m_mmsi: add_field(fields, source, "vessel.mmsi", m_mmsi.group(1), raw=m_mmsi.group(0))

    # Header basics
    m = re.search(r"IMO/LR Number\s+(\d{7})\s+([A-Z0-9 '\-]+)", text, re.I)
    if m:
        add_field(fields, source, "vessel.imo", m.group(1), raw=m.group(0))
        add_field(fields, source, "vessel.name", clean_text(m.group(2)), raw=m.group(0))
    add_field(fields, source, "hvpq.date_completed", find_value_after_label(lines, r"Date this HVPQ document completed"), label="HVPQ date completed")
    add_field(fields, source, "vessel.flag", next_value_after(lines, r"^Flag$"))
    add_field(fields, source, "vessel.port_registry", next_value_after(lines, r"^Port of Registry$"))
    add_field(fields, source, "vessel.call_sign", next_value_after(lines, r"^Call sign$"))
    add_field(fields, source, "vessel.mmsi", next_value_after(lines, r"MMSI"))
    add_field(fields, source, "vessel.type", next_value_after(lines, r"type of ship as described"))
    other_type = next_value_after(lines, r"^If other, then specify$")
    if other_type and normalize_value(other_type) not in ["na", "not applicable"]:
        add_field(fields, source, "vessel.type_detail", other_type, label="Vessel type detail")
    add_field(fields, source, "owner.registered_owner", next_value_after(lines, r"^Name$"), raw="Registered owner name from 1.3.1")
    add_field(fields, source, "owner.technical_operator", next_value_after(lines, r"Technical operator"))
    add_field(fields, source, "owner.commercial_operator", next_value_after(lines, r"Commercial operator"))
    add_field(fields, source, "insurance.pni_club", find_value_after_label(lines, r"If other, then specify"), raw="P&I if-other extraction")
    # More specific P&I window
    pni_win = window_after(lines, r"P and I Club", 8)
    m = re.search(r"If other, then specify\s+(.+?)(?:\s+3\s+Amount|\s+Amount|$)", pni_win, re.I)
    if m:
        add_field(fields, source, "insurance.pni_club", m.group(1), raw=pni_win)

    # Environment
    cii_win = window_after(lines, r"Carbon Intensity Indicator|CII", 12)
    m = re.search(r"provide CII rating\s+([A-Z])\b", cii_win, re.I)
    if m: add_field(fields, source, "environment.cii_rating", m.group(1), raw=cii_win)
    m = re.search(r"CII rating verified by Class, 3rd Party or Owner\??\s+([A-Za-z0-9 /]+)", cii_win, re.I)
    if m: add_field(fields, source, "environment.cii_verified_by", m.group(1), raw=cii_win)
    else: add_field(fields, source, "environment.cii_verified_by", next_value_after(lines, r"CII rating verified by Class"), raw=cii_win)
    eexi_win = window_after(lines, r"EEXI|Energy Efficiency Existing", 12)
    m = re.search(r"provide EEXI rating\s+([0-9.]+)", eexi_win, re.I)
    if m: add_field(fields, source, "environment.eexi_rating", m.group(1), raw=eexi_win)
    m = re.search(r"EEXI rating verified by Class, 3rd Party or Owner\??\s+([A-Za-z0-9 /]+)", eexi_win, re.I)
    if m: add_field(fields, source, "environment.eexi_verified_by", m.group(1), raw=eexi_win)
    else: add_field(fields, source, "environment.eexi_verified_by", next_value_after(lines, r"EEXI rating verified by Class"), raw=eexi_win)

    # Class / survey
    add_field(fields, source, "classification.class_society", next_value_after(lines, r"^Classification Society$"))
    notation_win = window_after(lines, r"Class notation", 8)
    m = re.search(r"List class notations\s+(.+?)(?:Provide details|1\.5\.3|Change of Classification|$)", notation_win, re.I)
    if m: add_field(fields, source, "classification.class_notation", m.group(1), raw=notation_win)
    for fid, patt in [
        ("surveys.last_drydock", r"Date of last dry dock"),
        ("surveys.next_drydock_due", r"Date next dry dock due"),
        ("surveys.last_iws", r"Date of last IWS"),
        ("surveys.next_iws_due", r"Date next IWS due"),
        ("surveys.last_special", r"Date of last special survey"),
        ("surveys.next_special_due", r"Date next special survey due"),
        ("surveys.last_annual", r"Date of last annual survey"),
        ("surveys.last_intermediate", r"Date of Last Intermediate survey"),
    ]:
        add_field(fields, source, fid, next_value_after(lines, patt), raw=patt)
    add_field(fields, source, "classification.conditions_of_class", next_value_after(lines, r"Does Vessel have any open Conditions of Class"))
    add_field(fields, source, "classification.memo_of_class", next_value_after(lines, r"Does Vessel have any Memoranda of Class"))
    add_field(fields, source, "classification.flag_dispensation", next_value_after(lines, r"Does vessel have any flag state dispensations"))

    # Incidents / PSC
    add_field(fields, source, "incidents.pollution_grounding_collision_allision", next_value_after(lines, r"pollution, grounding, collision or allision"))
    add_field(fields, source, "incidents.other_incidents", next_value_after(lines, r"other incidents during the past 12 months"))
    add_field(fields, source, "psc.last_date", next_value_after(lines, r"Date of last Port State Control Inspection"))
    add_field(fields, source, "psc.last_port", next_value_after(lines, r"Port of last Port State Control Inspection"))
    add_field(fields, source, "psc.detained_36m", next_value_after(lines, r"detained during the last 36 months"))

    # Certificates: restrict extraction to the HVPQ Certificate dates table only.
    cert_start = next((i for i,l in enumerate(lines) if re.search(r"Certificate dates", l, re.I)), -1)
    cert_end = next((i for i,l in enumerate(lines[cert_start+1:], cert_start+1) if re.search(r"Publications", l, re.I)), len(lines)) if cert_start >= 0 else -1
    cert_lines = lines[cert_start:cert_end] if cert_start >= 0 else []
    fields += parse_cert_rows_from_sequence(cert_lines, source, cert_order="hvpq")
    add_section_and_operational_fields(fields, source, text)
    extract_hvpq_tank_coating_fields(fields, source, text)
    return dedupe_fields(fields)


def extract_q88(pages: List[Tuple[int, str]]) -> List[FieldRecord]:
    source = "Q88"
    text = join_pages(pages)
    lines = lines_from_text(text)
    fields: List[FieldRecord] = []

    blocks = q88_blocks(lines)
    # Robust block-based extraction for Q88 table layout
    def ans(q): return block_answer(blocks.get(q, []))
    v = ans("1.1"); add_field(fields, source, "q88.date_updated", v, label="Q88 date updated", raw=" | ".join(blocks.get("1.1", [])))
    v = ans("1.2")
    m = re.search(r"(.+?)\s*\((\d{7})\)", v)
    if m:
        add_field(fields, source, "vessel.name", m.group(1), raw=v); add_field(fields, source, "vessel.imo", m.group(2), raw=v)
    v = ans("1.5")
    if "/" in v:
        a,b = v.split("/",1); add_field(fields, source, "vessel.flag", a, raw=v); add_field(fields, source, "vessel.port_registry", b, raw=v)
    v = ans("1.6")
    if "/" in v:
        a,b = v.split("/",1); add_field(fields, source, "vessel.call_sign", a, raw=v); add_field(fields, source, "vessel.mmsi", b, raw=v)
    for q, fid in [("1.8", "vessel.type"), ("1.8a", "vessel.type_detail"), ("1.10", "owner.registered_owner"), ("1.11", "owner.technical_operator"), ("1.12", "owner.commercial_operator"), ("1.14", "insurance.pni_club"), ("1.18", "classification.class_society"), ("1.19", "classification.class_notation")]:
        v = ans(q); add_field(fields, source, fid, v, raw=" | ".join(blocks.get(q, [])))
    for q, fid in [("1.20", "classification.conditions_of_class"), ("1.20a", "classification.memo_of_class")]:
        blocktxt = " | ".join(blocks.get(q, []))
        yns = re.findall(r"\b(Yes|No)\b", blocktxt, re.I)
        v = yns[-1] if yns else ans(q)
        add_field(fields, source, fid, v, raw=blocktxt)
    # survey dates
    for q in ["1.23", "1.24", "1.25", "1.25a"]:
        v = ans(q); dates = extract_dates(v)
        if q == "1.23" and dates: add_field(fields, source, "surveys.last_drydock", dates[0], raw=v)
        if q == "1.24" and dates:
            add_field(fields, source, "surveys.next_drydock_due", dates[0], raw=v)
            if len(dates)>1: add_field(fields, source, "surveys.next_annual_due", dates[1], raw=v)
        if q == "1.25" and dates:
            add_field(fields, source, "surveys.last_special", dates[0], raw=v)
            if len(dates)>1: add_field(fields, source, "surveys.next_special_due", dates[1], raw=v)
        if q == "1.25a" and dates:
            add_field(fields, source, "surveys.last_iws", dates[0], raw=v)
            if len(dates)>1: add_field(fields, source, "surveys.next_iws_due", dates[1], raw=v)
    # certificates in Q88 section 2
    for q, block in blocks.items():
        if re.match(r"2\.\d+", q):
            blocktxt = " ".join(block)
            key = cert_key_from_label(blocktxt)
            if key:
                dates = extract_dates(blocktxt)
                if len(dates)>=1: add_field(fields, source, f"cert.{key}.issue", dates[0], raw=blocktxt)
                if len(dates)>=4: add_field(fields, source, f"cert.{key}.expiry", dates[3], raw=blocktxt)
                elif len(dates)>=2: add_field(fields, source, f"cert.{key}.expiry", dates[-1], raw=blocktxt)
                if len(dates)>=2: add_field(fields, source, f"cert.{key}.last_annual", dates[1], raw=blocktxt)
                if len(dates)>=3: add_field(fields, source, f"cert.{key}.last_intermediate", dates[2], raw=blocktxt)
    # Q88 line-based extraction fallback
    for line in lines:
        if re.search(r"1\.1\s+Date updated", line, re.I):
            add_field(fields, source, "q88.date_updated", line.split(":")[-1], label="Q88 date updated", raw=line)
        m = re.search(r"1\.2\s+Vessel.*?:\s*(.+?)\s*\((\d{7})\)", line, re.I)
        if m:
            add_field(fields, source, "vessel.name", m.group(1), raw=line)
            add_field(fields, source, "vessel.imo", m.group(2), raw=line)
        m = re.search(r"1\.5\s+Flag/Port of Registry:\s*(.+?)/(.*)$", line, re.I)
        if m:
            add_field(fields, source, "vessel.flag", m.group(1), raw=line)
            add_field(fields, source, "vessel.port_registry", m.group(2), raw=line)
        m = re.search(r"1\.6\s+Call sign/MMSI:\s*([^/]+)/(.+)$", line, re.I)
        if m:
            add_field(fields, source, "vessel.call_sign", m.group(1), raw=line)
            add_field(fields, source, "vessel.mmsi", m.group(2), raw=line)
        m = re.search(r"1\.8\s+Type of vessel.*?:\s*(.+)$", line, re.I)
        if m: add_field(fields, source, "vessel.type", m.group(1), raw=line)
        m = re.search(r"1\.8a.*?specify:\s*(.+)$", line, re.I)
        if m: add_field(fields, source, "vessel.type_detail", m.group(1), label="Vessel type detail", raw=line)
        m = re.search(r"1\.10\s+Registered owner.*?:\s*(.+)$", line, re.I)
        if m: add_field(fields, source, "owner.registered_owner", m.group(1), raw=line)
        m = re.search(r"1\.11\s+Technical operator.*?:\s*(.+)$", line, re.I)
        if m: add_field(fields, source, "owner.technical_operator", m.group(1), raw=line)
        m = re.search(r"1\.12\s+Commercial operator.*?:\s*(.+)$", line, re.I)
        if m: add_field(fields, source, "owner.commercial_operator", m.group(1), raw=line)
        m = re.search(r"1\.14\s+P\s*&\s*I Club.*?:\s*(.+)$", line, re.I)
        if m: add_field(fields, source, "insurance.pni_club", m.group(1), raw=line)
        m = re.search(r"1\.18\s+Classification society:\s*(.+)$", line, re.I)
        if m: add_field(fields, source, "classification.class_society", m.group(1), raw=line)
        # conditions/memo handled by block parser above
        for fid, qno in [("surveys.last_drydock", "1.23"), ("surveys.next_drydock_due", "1.24"), ("surveys.last_special", "1.25"), ("surveys.last_iws", "1.25a")]:
            if line.startswith(qno):
                dates = extract_dates(line)
                if qno == "1.23" and dates:
                    add_field(fields, source, "surveys.last_drydock", dates[0], raw=line)
                elif qno == "1.24" and dates:
                    add_field(fields, source, "surveys.next_drydock_due", dates[0], raw=line)
                    if len(dates) > 1: add_field(fields, source, "surveys.next_annual_due", dates[1], raw=line)
                elif qno == "1.25" and dates:
                    add_field(fields, source, "surveys.last_special", dates[0], raw=line)
                    if len(dates) > 1: add_field(fields, source, "surveys.next_special_due", dates[1], raw=line)
                elif qno == "1.25a" and dates:
                    add_field(fields, source, "surveys.last_iws", dates[0], raw=line)
                    if len(dates) > 1: add_field(fields, source, "surveys.next_iws_due", dates[1], raw=line)
        # Certificate rows Q88 order: issued, last annual, last intermediate, expires
        if re.match(r"^2\.\d+\s+", line):
            key = cert_key_from_label(line)
            if key:
                dates = extract_dates(line)
                if len(dates) >= 1: add_field(fields, source, f"cert.{key}.issue", dates[0], raw=line)
                if len(dates) >= 4: add_field(fields, source, f"cert.{key}.expiry", dates[-1], raw=line)
                elif len(dates) >= 2: add_field(fields, source, f"cert.{key}.expiry", dates[-1], raw=line)
                if len(dates) >= 2: add_field(fields, source, f"cert.{key}.last_annual", dates[1], raw=line)
                if len(dates) >= 3: add_field(fields, source, f"cert.{key}.last_intermediate", dates[2], raw=line)
        m = re.search(r"CII rating.*?\b([A-E])\b", line, re.I)
        if m:
            add_field(fields, source, "environment.cii_rating", m.group(1), raw=line)
        if re.search(r"CII rating verified", line, re.I):
            tail = line.split("?")[-1].strip()
            if tail and tail.lower() != line.lower():
                add_field(fields, source, "environment.cii_verified_by", tail, raw=line)

    add_section_and_operational_fields(fields, source, text)
    extract_q88_coating_fields(fields, source, text)

    # Certificate special: P&I coverage/expiration contains expiry only
    for line in lines:
        m = re.search(r"1\.15\s+P\s*&\s*I Club pollution liability coverage/expiration date:\s*(.+)$", line, re.I)
        if m:
            dates = extract_dates(line)
            if dates:
                add_field(fields, source, "cert.pni_cover.expiry", dates[-1], raw=line)
    # Class notation window
    win = window_after(lines, r"1\.19\s+Class notation", 5)
    if win:
        add_field(fields, source, "classification.class_notation", re.sub(r"^1\.19\s+Class notation:\s*", "", win, flags=re.I), raw=win)

    return dedupe_fields(fields)


def extract_piq(pages: List[Tuple[int, str]]) -> List[FieldRecord]:
    source = "PIQ"
    text = join_pages(pages)
    lines = lines_from_text(text)
    fields: List[FieldRecord] = []
    m = re.search(r"Vessel Name\s+([A-Z0-9 '\-]+)\s+Date\s+(\d{1,2}\s+[A-Za-z]+\s+\d{4})", text, re.I)
    if m:
        add_field(fields, source, "vessel.name", m.group(1), raw=m.group(0))
        add_field(fields, source, "piq.date", m.group(2), label="PIQ date", raw=m.group(0))
    else:
        # Many PIQs extract header as: PIQ Report / VESSEL / Vessel Name / Date / date
        m2 = re.search(r"PIQ Report\s+([A-Z0-9 '\-]{3,80})\s+Vessel Name\s+Date\s+(\d{1,2}\s+[A-Za-z]+\s+\d{4})", text, re.I)
        if m2:
            add_field(fields, source, "vessel.name", clean_text(m2.group(1)), raw=m2.group(0))
            add_field(fields, source, "piq.date", m2.group(2), label="PIQ date", raw=m2.group(0))
    add_field(fields, source, "vessel.type", find_value_after_label(lines, r"^Vessel Type\b"))
    ann = find_value_after_label(lines, r"Annex II cargo")
    add_field(fields, source, "piq.annex_ii_carried_or_intended", ann, label="Annex II carried/intended")
    add_field(fields, source, "classification.last_class_visit", find_value_after_label(lines, r"Date of last visit"))
    add_field(fields, source, "classification.last_class_visit_purpose", find_value_after_label(lines, r"Purpose of visit"))

    # Superintendent visit blocks
    tech_win = window_after(lines, r"Technical Superintendent inspection completed", 25)
    mar_win = window_after(lines, r"Marine Superintendent inspection completed", 25)
    add_visit_fields(fields, source, "technical", tech_win)
    add_visit_fields(fields, source, "marine", mar_win)

    # Tank inspections
    tank_patterns = [
        ("tank.cargo_slop.freq_months", "tank.cargo_slop.oldest", r"cargo tanks", r"cargo and slop"),
        ("tank.ballast.freq_months", "tank.ballast.oldest", r"ballast tanks", r"ballast tanks"),
        ("tank.void.freq_months", "tank.void.oldest", r"void spaces", r"void spaces"),
    ]
    for fid_freq, fid_old, p1, p2 in tank_patterns:
        win = window_after(lines, p1, 8)
        m = re.search(r"Required frequency.*?(\d+)\s+months", win, re.I)
        if m: add_field(fields, source, fid_freq, m.group(1), label=fid_freq, raw=win)
        dates = extract_dates(win)
        if dates: add_field(fields, source, fid_old, dates[-1], label=fid_old, raw=win)

    # CAP / MOC
    add_field(fields, source, "cap.enrolled", find_value_after_label(lines, r"Enrolled in a condition assessment programme"))
    add_field(fields, source, "cap.overall_rating", find_value_after_label(lines, r"Overall rating"))
    add_field(fields, source, "cap.completed", find_value_after_label(lines, r"CAP survey completed"))
    add_field(fields, source, "cap.class_society", find_value_after_label(lines, r"Which classification society issued"))
    add_field(fields, source, "moc.structural_change", find_value_after_label(lines, r"Have any structural changes been made"))
    add_field(fields, source, "moc.retrofit", find_value_after_label(lines, r"Has any new equipment been retrofitted"))
    retrofit_win = window_after(lines, r"Has any new equipment been retrofitted", 10)
    add_field(fields, source, "moc.retrofit_details", retrofit_win, label="Retrofit details", raw=retrofit_win)
    add_field(fields, source, "moc.equipment_replaced", find_value_after_label(lines, r"Equipment replaced"))
    add_field(fields, source, "moc.equipment_replaced_details", window_after(lines, r"Equipment replaced", 8), label="Equipment replaced details")
    add_field(fields, source, "moc.equipment_decommissioned", find_value_after_label(lines, r"Equipment decommissioned"))

    # More PIQ extraction: audits, assessments, training and key yes/no declarations.
    piq_extra_patterns = [
        ("piq.static_nav_assessment", r"Static navigational assessment conducted"),
        ("piq.dynamic_nav_assessment_shore", r"Dynamic navigational assessment conducted by a member"),
        ("piq.dynamic_nav_assessment_third_party", r"Dynamic navigational assessment conducted by a third party"),
        ("piq.remote_nav_assessment", r"Unannounced remote navigational assessment"),
        ("piq.cargo_audit", r"Comprehensive cargo audit"),
        ("piq.engineering_audit", r"Comprehensive engineering audit"),
        ("piq.mooring_anchoring_audit", r"Comprehensive mooring and anchoring audit"),
        ("piq.behavioural_competency", r"Behavioural Competency Assessment programme"),
        ("piq.brm_training", r"BTM/BRM training course attendance"),
        ("piq.cargo_simulator", r"shore based cargo simulation course attended"),
        ("piq.sms_language", r"Primary Language"),
        ("piq.working_language", r"Common working language"),
    ]
    for fid, patt in piq_extra_patterns:
        val = find_value_after_label(lines, patt)
        if not val:
            win = window_after(lines, patt, 5)
            m = re.search(r"\b(Yes|No|English|Not applicable|NA)\b", win, re.I)
            val = m.group(1) if m else ""
        add_field(fields, source, fid, val, label=fid, raw=window_after(lines, patt, 5))

    # PSC block
    psc_win = window_after(lines, r"last three Port State Control", 45)
    add_field(fields, source, "psc.block", psc_win, label="PSC block", raw=psc_win)
    m = re.search(r"Last\s+" + DATE_RE.pattern + r"\s+([^|\n]+?)(?:\s+(?:US Coastguard|Tokyo MoU|Indian Ocean MoU|Paris MoU|Vina Del Mar|Black Sea MoU|Mediterranean MoU)|\s+\d+\.\d+|$)", psc_win, re.I)
    if m:
        dates = extract_dates(m.group(0))
        if dates: add_field(fields, source, "psc.last_date", dates[0], raw=psc_win)
        # Port extraction from full PSC table is handled by the robust fallback below.
    # robust fallback for extracted PSC tables
    m2 = re.search(r"Last\s+(\d{1,2}\s+[A-Za-z]+\s+\d{4})\s+(.+?)\s+(US Coastguard|Tokyo MoU|Indian Ocean MoU|Paris MoU|Vina Del Mar|Black Sea MoU|Mediterranean MoU)", psc_win, re.I)
    if m2:
        if not first_field(fields, source, "psc.last_date"):
            add_field(fields, source, "psc.last_date", m2.group(1), raw=psc_win)
        add_field(fields, source, "psc.last_port", clean_text(m2.group(2)), raw=psc_win)

    # Incidents: PIQ incident section often later, extract yes/no if visible
    for kw, fid in [
        ("pollution", "incidents.pollution"), ("grounding", "incidents.grounding"),
        ("collision", "incidents.collision"), ("allision", "incidents.allision"),
        ("mooring", "incidents.mooring"), ("injury", "incidents.injury"),
        ("machinery", "incidents.machinery"),
    ]:
        win = window_after(lines, kw, 4)
        m = re.search(r"\b(Yes|No)\b", win, re.I)
        if m: add_field(fields, source, fid, m.group(1), label=f"PIQ incident {kw}", raw=win)
    # collapsed incident yes/no for no incidents declared logic
    no_count = sum(1 for f in fields if f.field_id.startswith("incidents.") and normalize_bool(f.value) == "no")
    yes_count = sum(1 for f in fields if f.field_id.startswith("incidents.") and normalize_bool(f.value) == "yes")
    if yes_count == 0 and no_count > 0:
        add_field(fields, source, "incidents.other_incidents", "No", raw="Derived from PIQ incident section")

    return dedupe_fields(fields)


def add_visit_fields(fields: List[FieldRecord], source: str, kind: str, win: str):
    if not win:
        return
    add_field(fields, source, f"superintendent.{kind}.raw", win, label=f"{kind.title()} superintendent raw", raw=win)
    # Rows look like Last 01 October 2025 28 November 2025 59.00 Physical
    rows = re.findall(r"(Second Last|Third Last|Last)\s+(\d{1,2}\s+[A-Za-z]+\s+\d{4})\s+(\d{1,2}\s+[A-Za-z]+\s+\d{4})\s+([0-9.]+)?\s*([A-Za-z]*)", win, re.I)
    for idx, row in enumerate(rows, 1):
        label, frm, to, days, typ = row
        add_field(fields, source, f"superintendent.{kind}.{idx}.from", frm, label=f"{kind.title()} superintendent {label} from", raw=win)
        add_field(fields, source, f"superintendent.{kind}.{idx}.to", to, label=f"{kind.title()} superintendent {label} to", raw=win)
        if days: add_field(fields, source, f"superintendent.{kind}.{idx}.days", days, label=f"{kind.title()} superintendent {label} days", raw=win)
        if typ: add_field(fields, source, f"superintendent.{kind}.{idx}.type", typ, label=f"{kind.title()} superintendent {label} type", raw=win)


def extract_class_status(pages: List[Tuple[int, str]]) -> List[FieldRecord]:
    source = "CLASS"
    text = join_pages(pages)
    lines = lines_from_text(text)
    fields: List[FieldRecord] = []
    if re.search(r"KOREAN REGISTER|\bKR\b", text, re.I):
        class_soc = "Korean Register"
    elif re.search(r"NIPPON KAIJI KYOKAI|ClassNK|NK-SHIPS", text, re.I):
        class_soc = "Nippon Kaiji Kyokai"
    else:
        class_soc = find_value_after_label(lines, r"Class Society|Classification Society")
    add_field(fields, source, "classification.class_society", class_soc)
    add_field(fields, source, "vessel.name", next_value_after(lines, r"^Ship Name$|^Name of Ship$"))
    m_imo = re.search(r"IMO No\s*:?\s*(\d{7})|IMO Number\s*:?\s*(\d{7})", text, re.I)
    add_field(fields, source, "vessel.imo", (m_imo.group(1) or m_imo.group(2)) if m_imo else next_value_after(lines, r"^IMO No\.?$|^IMO Number$"))
    add_field(fields, source, "vessel.flag", next_value_after(lines, r"^Flag$"))
    add_field(fields, source, "vessel.port_registry", next_value_after(lines, r"^Port of Registry$"))
    add_field(fields, source, "classification.class_notation", window_after(lines, r"Class Notation|Classification Character", 6))
    add_field(fields, source, "owner.registered_owner", next_value_after(lines, r"^Owner$|^Registered Owner$"))
    add_field(fields, source, "owner.technical_operator", next_value_after(lines, r"^Technical Manager$|^Management Company$"))

    # KR/ClassNK certificate tables: restrict to certificate description section
    cstart = next((i for i,l in enumerate(lines) if re.search(r"Certificate description", l, re.I)), -1)
    cend = next((i for i,l in enumerate(lines[cstart+1:], cstart+1) if re.search(r"Survey Description", l, re.I)), len(lines)) if cstart >= 0 else -1
    class_cert_lines = lines[cstart:cend] if cstart >= 0 else []
    fields += parse_cert_rows_from_sequence(class_cert_lines, source, cert_order="class")

    # Certificate rows in Class Status. Be strict with DNV-style labels:
    # 'Issued date' is issue only; expiry must be labelled 'Valid until' / 'Expiry' / equivalent.
    for i, line in enumerate(lines):
        key = cert_key_from_label(line)
        if key:
            win = " ".join(lines[max(0, i-2):min(len(lines), i+3)])
            dates = extract_dates(win)
            issue_dt, expiry_dt = labelled_cert_dates(win)
            if issue_dt and not sources_value(fields, f"cert.{key}.issue").get(source):
                add_field(fields, source, f"cert.{key}.issue", issue_dt, raw=win)
            if expiry_dt and not sources_value(fields, f"cert.{key}.expiry").get(source):
                add_field(fields, source, f"cert.{key}.expiry", expiry_dt, raw=win)
            # Safe fallback: multiple dates can be issue+validity only if an expiry label exists in the same window.
            if has_expiry_label(win) and len(dates) >= 2 and not sources_value(fields, f"cert.{key}.expiry").get(source):
                add_field(fields, source, f"cert.{key}.expiry", dates[-1], raw=win)
            elif has_issue_label(win) and len(dates) >= 1 and not sources_value(fields, f"cert.{key}.issue").get(source):
                add_field(fields, source, f"cert.{key}.issue", dates[0], raw=win)
    # Last chance expiry extraction: only if an expiry/validity label is present nearby.
    for i, line in enumerate(lines):
        key = cert_key_from_label(line)
        if key and not sources_value(fields, f"cert.{key}.expiry").get(source):
            win = " ".join(lines[max(0, i-2):min(len(lines), i+4)])
            dates = extract_dates(win)
            if dates and has_expiry_label(win):
                issue_dt, expiry_dt = labelled_cert_dates(win)
                add_field(fields, source, f"cert.{key}.expiry", expiry_dt or dates[-1], raw=win)

    # Survey status rows KR/NK
    for line in lines:
        norm = normalize_key(line)
        dates = extract_dates(line)
        if not dates:
            continue
        if "special survey" in norm:
            if re.search(r"next|due", line, re.I): add_field(fields, source, "surveys.next_special_due", dates[0], raw=line)
            add_field(fields, source, "surveys.last_special", dates[-1], raw=line)
        if "annual survey" in norm:
            if re.search(r"due|range", line, re.I): add_field(fields, source, "surveys.next_annual_due", dates[0], raw=line)
            add_field(fields, source, "surveys.last_annual", dates[-1], raw=line)
        if "intermediate survey" in norm:
            add_field(fields, source, "surveys.last_intermediate", dates[-1], raw=line)
        if "docking survey" in norm or "dock" in norm:
            if re.search(r"next|due", line, re.I): add_field(fields, source, "surveys.next_drydock_due", dates[0], raw=line)
            add_field(fields, source, "surveys.last_drydock", dates[-1], raw=line)
        if "iws" in norm or "in water" in norm:
            if re.search(r"next|due", line, re.I): add_field(fields, source, "surveys.next_iws_due", dates[0], raw=line)
            add_field(fields, source, "surveys.last_iws", dates[-1], raw=line)

    # Conditions / memo / recommendations: be conservative, don't confuse empty sections with mismatch.
    cc_text = " ".join([l for l in lines if re.search(r"condition(s)? of class|recommendation|memoranda|memorandum|memo", l, re.I)])
    if cc_text:
        if re.search(r"no\s+(open\s+)?condition|none|nil", cc_text, re.I):
            add_field(fields, source, "classification.conditions_of_class", "No", raw=cc_text)
        elif re.search(r"condition", cc_text, re.I):
            add_field(fields, source, "classification.conditions_of_class", "Review listed items", raw=cc_text)
        if re.search(r"no\s+(memoranda|memorandum|memo)|none|nil", cc_text, re.I):
            add_field(fields, source, "classification.memo_of_class", "No", raw=cc_text)

    return dedupe_fields(fields)


def extract_xml(file_obj) -> List[FieldRecord]:
    source = "XML"
    fields: List[FieldRecord] = []
    if file_obj is None:
        return fields
    data = file_obj.getvalue() if hasattr(file_obj, "getvalue") else file_obj.read()
    try:
        root = ET.fromstring(data)
        ns = {"x": root.tag.split("}")[0].strip("{")} if "}" in root.tag else {}
        vessel = root.find(".//x:Vessel", ns) if ns else root.find(".//Vessel")
        if vessel is not None:
            add_field(fields, source, "vessel.name", vessel.attrib.get("name", ""), raw=ET.tostring(vessel, encoding="unicode"))
            add_field(fields, source, "vessel.imo", vessel.attrib.get("id", ""), raw=ET.tostring(vessel, encoding="unicode"))
        doc = root.find(".//x:Document", ns) if ns else root.find(".//Document")
        if doc is not None:
            add_field(fields, source, "xml.exported", doc.attrib.get("exported", ""), label="XML exported date")
        templ = root.find(".//x:Template", ns) if ns else root.find(".//Template")
        if templ is not None:
            add_field(fields, source, "xml.template", templ.attrib.get("variant", ""), label="XML template")
    except Exception as e:
        add_field(fields, source, "xml.error", str(e), label="XML parse error")
    return fields


def dedupe_fields(fields: List[FieldRecord]) -> List[FieldRecord]:
    seen = set()
    out = []
    for f in fields:
        k = (f.source, f.field_id, normalize_value(f.value), f.date_value)
        if k in seen:
            continue
        seen.add(k)
        out.append(f)
    return out

# ----------------------------- Optional local LLM extraction assist -----------------------------

def ollama_generate(base_url: str, model: str, prompt: str, timeout: int = 120) -> str:
    url = base_url.rstrip("/") + "/api/generate"
    payload = {"model": model, "prompt": prompt, "stream": False, "options": {"temperature": 0.0}}
    r = requests.post(url, json=payload, timeout=timeout)
    r.raise_for_status()
    return r.json().get("response", "")


def extract_json_from_llm_response(resp: str) -> Dict[str, Any]:
    resp = resp.strip()
    m = re.search(r"\{.*\}", resp, re.S)
    if not m:
        return {}
    try:
        return json.loads(m.group(0))
    except Exception:
        return {}


def llm_assist_extract(doc_type: str, pages: List[Tuple[int, str]], base_url: str, model: str) -> List[FieldRecord]:
    source = doc_type.upper() + "_LLM"
    text = join_pages(pages)
    # Use only relevant first 18k chars to avoid local context overflow; class/Q88 certs in early pages generally.
    snippet = text[:18000]
    prompt = f"""
You are a marine vetting document extraction engine. Extract only values visible in the text. Do not guess.
Document type: {doc_type}
Return ONLY valid JSON. If value is not visible, use null.

Required JSON keys:
{{
  "vessel": {{"name": null, "imo": null, "flag": null, "port_registry": null, "type": null, "call_sign": null, "mmsi": null}},
  "owner": {{"registered_owner": null, "technical_operator": null, "commercial_operator": null}},
  "insurance": {{"pni_club": null}},
  "classification": {{"class_society": null, "class_notation": null, "conditions_of_class": null, "memo_of_class": null, "flag_dispensation": null}},
  "environment": {{"cii_rating": null, "cii_verified_by": null, "eexi_rating": null, "eexi_verified_by": null}},
  "surveys": {{"last_drydock": null, "next_drydock_due": null, "last_iws": null, "next_iws_due": null, "last_special": null, "next_special_due": null, "last_annual": null, "next_annual_due": null, "last_intermediate": null}},
  "certificates": {{
    "cofr": {{"issue": null, "expiry": null}}, "iopp": {{"issue": null, "expiry": null}}, "vgp": {{"issue": null, "expiry": null}},
    "doc": {{"issue": null, "expiry": null, "last_annual": null}}, "smc": {{"issue": null, "expiry": null, "last_annual": null}},
    "issc": {{"issue": null, "expiry": null, "last_annual": null}}, "class_certificate": {{"issue": null, "expiry": null, "last_annual": null, "last_intermediate": null}},
    "safety_equipment": {{"issue": null, "expiry": null, "last_annual": null, "last_intermediate": null}},
    "safety_radio": {{"issue": null, "expiry": null, "last_annual": null, "last_intermediate": null}},
    "safety_construction": {{"issue": null, "expiry": null, "last_annual": null, "last_intermediate": null}},
    "loadline": {{"issue": null, "expiry": null, "last_annual": null, "last_intermediate": null}},
    "cof_chemical": {{"issue": null, "expiry": null, "last_annual": null, "last_intermediate": null}}
  }},
  "incidents": {{"pollution_grounding_collision_allision": null, "other_incidents": null}}
}}

TEXT:
{snippet}
"""
    try:
        resp = ollama_generate(base_url, model, prompt)
        data = extract_json_from_llm_response(resp)
    except Exception as e:
        return [FieldRecord(source=source, field_id="llm.error", label="LLM extraction error", value=str(e), confidence="llm-error")]
    fields: List[FieldRecord] = []
    flatten_llm_json(data, source, fields)
    return fields


def flatten_llm_json(data: Dict[str, Any], source: str, fields: List[FieldRecord], prefix: str = ""):
    for k, v in (data or {}).items():
        fid = f"{prefix}.{k}" if prefix else k
        if isinstance(v, dict):
            if prefix == "certificates":
                cert_key = k
                for sub, val in v.items():
                    if val not in [None, "", []]:
                        add_field(fields, source, f"cert.{cert_key}.{sub}", val, confidence="local-llm")
            else:
                flatten_llm_json(v, source, fields, fid)
        elif v not in [None, "", []]:
            add_field(fields, source, fid, v, confidence="local-llm")

# ----------------------------- Rule engine -----------------------------

def add_finding(findings: List[Finding], **kwargs):
    findings.append(Finding(**kwargs))


def compare_field(findings: List[Finding], fields: List[FieldRecord], field_id: str, sources: List[str], area: str, check_name: str, risk: str = "MEDIUM", hard: bool = False):
    vals = {s: first_field(fields, s, field_id) for s in sources}
    vals = {s: v for s, v in vals.items() if clean_text(v)}
    if len(vals) < 2:
        return
    pairs = list(vals.items())
    base_s, base_v = pairs[0]
    mismatch = False
    for s, v in pairs[1:]:
        if not semantically_equivalent(base_v, v, field_id):
            mismatch = True
    if mismatch:
        status = "MISMATCH" if hard else "MANUAL CHECK"
        add_finding(findings, area=area, check=check_name, status=status, risk=risk,
                    hvpq_value=vals.get("HVPQ", ""), piq_value=vals.get("PIQ", ""), class_value=vals.get("CLASS", ""), q88_value=vals.get("Q88", ""), xml_value=vals.get("XML", ""),
                    reason="Mapped values differ after normalization. Treat as manual check unless the source document confirms exact official value.",
                    action="Verify source documents and update the stale/incorrect declaration.")


def run_rules(fields: List[FieldRecord], ref_date: date, settings: Dict[str, Any], obs_df: pd.DataFrame) -> List[Finding]:
    findings: List[Finding] = []

    # 1. Identity only big mismatches
    compare_field(findings, fields, "vessel.imo", ["HVPQ", "PIQ", "Q88", "CLASS", "XML"], "Identity", "IMO consistency across documents", "CRITICAL", hard=True)
    compare_field(findings, fields, "vessel.name", ["HVPQ", "PIQ", "Q88", "CLASS", "XML"], "Identity", "Vessel name consistency across documents", "HIGH", hard=False)

    # 2. Vessel type as manual consistency, not hard mismatch
    vals_type = sources_value(fields, "vessel.type")
    if len(vals_type) >= 2:
        # Only flag as manual if textual difference visible; not high mismatch.
        normset = set(normalize_value(v) for v in vals_type.values() if v)
        if len(normset) > 1:
            add_finding(findings, area="General", check="Vessel type wording consistency", status="MANUAL CHECK", risk="MEDIUM",
                        hvpq_value=vals_type.get("HVPQ", ""), piq_value=vals_type.get("PIQ", ""), class_value=vals_type.get("CLASS", ""), q88_value=vals_type.get("Q88", ""),
                        reason="Vessel type wording differs across documents. This may be acceptable for oil/product/chemical wording but should be confirmed against IOPP/COF.",
                        action="Confirm official type wording and Annex II/chemical carriage applicability.")

    # 3. Owner/P&I spelling/wording
    compare_field(findings, fields, "owner.registered_owner", ["HVPQ", "Q88", "CLASS"], "Ownership", "Registered owner consistency", "MEDIUM", hard=False)
    vals_pni = sources_value(fields, "insurance.pni_club")
    if len(vals_pni) >= 2:
        hv, qv = vals_pni.get("HVPQ", ""), vals_pni.get("Q88", "")
        if hv and qv and not semantically_equivalent(hv, qv, "insurance.pni_club"):
            risk = "LOW" if similarity(hv, qv) > 0.72 else "MEDIUM"
            add_finding(findings, area="Insurance", check="P&I club spelling / consistency", status="MANUAL CHECK", risk=risk,
                        hvpq_value=hv, q88_value=qv, reason="P&I club name differs or appears misspelled.", action="Correct typo/style in HVPQ/Q88 if required.")

    # 4. Class society broad, official class status only useful here
    compare_field(findings, fields, "classification.class_society", ["HVPQ", "Q88", "CLASS"], "Class", "Class society consistency", "HIGH", hard=False)
    for fid, label in [("classification.conditions_of_class", "Conditions of Class"), ("classification.memo_of_class", "Memoranda of Class"), ("classification.flag_dispensation", "Flag/Class dispensation")]:
        vals = sources_value(fields, fid)
        if len(vals) >= 2:
            bools = {s: normalize_bool(v) for s, v in vals.items()}
            if "yes" in bools.values() and "no" in bools.values():
                add_finding(findings, area="Class", check=f"{label} declaration mismatch", status="MISMATCH", risk="CRITICAL",
                            hvpq_value=vals.get("HVPQ", ""), class_value=vals.get("CLASS", ""), q88_value=vals.get("Q88", ""),
                            reason=f"{label} appears differently declared across documents.", action="Verify latest Class Status and update HVPQ/Q88.")

    # 5. Superintendent gaps
    run_superintendent_rules(findings, fields, ref_date)

    # 6. Certificates: compare HVPQ/Q88/CLASS mapped fields.
    certs = sorted(set([f.field_id.split(".")[1] for f in fields if f.field_id.startswith("cert.") and len(f.field_id.split(".")) >= 3]))
    for cert in certs:
        if cert in ["tonnage"]:
            continue
        for part in ["issue", "expiry"]:
            if cert == "vgp" and part == "expiry":
                continue
            fid = f"cert.{cert}.{part}"
            vals = sources_value(fields, fid)
            if len(vals) < 2:
                continue
            # Prefer Class/Q88 as current cross-check sources for certs. If HVPQ differs from any by date, flag.
            hv = vals.get("HVPQ", "")
            qv = vals.get("Q88", "")
            cv = vals.get("CLASS", "")
            hvd, qvd, cvd = parse_date_any(hv), parse_date_any(qv), parse_date_any(cv)
            source_newer = qvd or cvd
            # Expiry critical if HVPQ expired but other source valid later
            if part == "expiry" and hvd:
                latest = max([d for d in [qvd, cvd] if d] or [hvd])
                if hvd < ref_date <= latest and latest > hvd:
                    add_finding(findings, area="Certificates", check=f"{cert.upper()} expiry outdated/expired in HVPQ", status="MISMATCH", risk="CRITICAL",
                                hvpq_value=hv, q88_value=qv, class_value=cv,
                                reason="HVPQ certificate expiry appears expired or older while another source shows a later valid expiry.",
                                action="Verify latest certificate and update HVPQ immediately.")
                    continue
            # Issue/expiry mismatches larger than tolerance.
            dlist = [("HVPQ", hvd), ("Q88", qvd), ("CLASS", cvd)]
            dlist = [(s, d) for s, d in dlist if d]
            if len(dlist) >= 2:
                dates_only = [d for s, d in dlist]
                if max(dates_only) != min(dates_only):
                    day_gap = (max(dates_only) - min(dates_only)).days
                    if day_gap > (0 if part == "expiry" else 7):
                        risk = "HIGH"
                        if cert in ["cofr", "iopp", "smc", "issc", "class_certificate"] and part == "expiry":
                            risk = "CRITICAL" if min(dates_only) < ref_date else "HIGH"
                        add_finding(findings, area="Certificates", check=f"{cert.upper()} {part} date mismatch", status="MANUAL CHECK", risk=risk,
                                    hvpq_value=hv, q88_value=qv, class_value=cv,
                                    reason=f"Certificate {part} dates differ across mapped sources by {day_gap} days.",
                                    action="Check the latest certificate/Class Status and correct stale declarations.")

    # 7. Certificate sanity: annual before issue etc.
    for src in ["HVPQ", "Q88"]:
        for cert in certs:
            issue = parse_date_any(first_field(fields, src, f"cert.{cert}.issue"))
            ann = parse_date_any(first_field(fields, src, f"cert.{cert}.last_annual"))
            exp = parse_date_any(first_field(fields, src, f"cert.{cert}.expiry"))
            if issue and ann and ann < issue - relativedelta(days=1):
                add_finding(findings, area="Certificates", check=f"{cert.upper()} annual endorsement before issue date", status="MANUAL CHECK", risk="HIGH",
                            hvpq_value=first_field(fields, src, f"cert.{cert}.issue") + " / " + first_field(fields, src, f"cert.{cert}.last_annual") if src == "HVPQ" else "",
                            q88_value=first_field(fields, src, f"cert.{cert}.issue") + " / " + first_field(fields, src, f"cert.{cert}.last_annual") if src == "Q88" else "",
                            reason="Last annual/endorsement date appears earlier than certificate issue date; may be table extraction issue or stale entry.", action="Verify certificate row manually.")
            other_expiries = []
            for osrc in ["HVPQ", "Q88", "CLASS"]:
                if osrc != src:
                    od = parse_date_any(first_field(fields, osrc, f"cert.{cert}.expiry"))
                    if od: other_expiries.append(od)
            if exp and exp < ref_date and cert not in ["vgp", "tonnage"] and not any(od > exp for od in other_expiries):
                add_finding(findings, area="Certificates", check=f"{cert.upper()} appears expired in {src}", status="MISMATCH", risk="CRITICAL",
                            hvpq_value=first_field(fields, src, f"cert.{cert}.expiry") if src == "HVPQ" else "",
                            q88_value=first_field(fields, src, f"cert.{cert}.expiry") if src == "Q88" else "",
                            reason=f"Certificate expiry in {src} is before reference date {ref_date.isoformat()}.", action="Update expired/stale certificate data.")

    # 8. Surveys: IWS next due blank if last IWS exists
    for src in ["HVPQ", "Q88", "CLASS"]:
        last = first_field(fields, src, "surveys.last_iws")
        nxt = first_field(fields, src, "surveys.next_iws_due")
        if last and not nxt:
            add_finding(findings, area="Class / Survey", check=f"IWS next due missing in {src}", status="MANUAL CHECK", risk="HIGH",
                        hvpq_value=last if src == "HVPQ" else "", q88_value=last if src == "Q88" else "", class_value=last if src == "CLASS" else "",
                        reason="Last IWS is declared but next IWS due is blank/not extracted.", action="Confirm whether IWS remains applicable and update HVPQ/Q88 if required.")
    for fid, label in [("surveys.last_drydock", "Last dry dock"), ("surveys.next_drydock_due", "Next dry dock due"), ("surveys.last_special", "Last special survey"), ("surveys.next_special_due", "Next special survey due")]:
        compare_dates_mapped(findings, fields, fid, label)

    # 9. CII/EEXI blank/mismatch checks
    vals = sources_value(fields, "environment.cii_rating")
    if len(vals) >= 2 and len(set(normalize_value(v) for v in vals.values())) > 1:
        add_finding(findings, area="Environmental", check="CII rating mismatch", status="MISMATCH", risk="HIGH",
                    hvpq_value=vals.get("HVPQ", ""), q88_value=vals.get("Q88", ""), reason="CII rating differs across documents.", action="Verify latest CII statement/SEEMP and update.")
    hv_cii_basis = first_field(fields, "HVPQ", "environment.cii_verified_by")
    q_cii_basis = first_field(fields, "Q88", "environment.cii_verified_by")
    q_cii_rating = first_field(fields, "Q88", "environment.cii_rating")
    if q_cii_rating and hv_cii_basis and not q_cii_basis:
        add_finding(findings, area="Environmental", check="Q88 CII verification basis blank", status="MANUAL CHECK", risk="MEDIUM",
                    hvpq_value=hv_cii_basis, q88_value=q_cii_basis,
                    reason="Q88 appears to contain CII rating but verification basis is blank/not extracted while HVPQ has a basis.", action="Verify Q88 CII verification basis.")

    # 10. Incidents
    hv_inc = normalize_bool(first_field(fields, "HVPQ", "incidents.other_incidents"))
    hv_pgca = normalize_bool(first_field(fields, "HVPQ", "incidents.pollution_grounding_collision_allision"))
    piq_inc = normalize_bool(first_field(fields, "PIQ", "incidents.other_incidents"))
    q_inc = normalize_bool(first_field(fields, "Q88", "incidents.other_incidents"))
    # No incidents across docs -> manual positive confirmation
    if all(x in ["", "no"] for x in [hv_inc, hv_pgca, piq_inc, q_inc]) and any(x == "no" for x in [hv_inc, hv_pgca, piq_inc, q_inc]):
        add_finding(findings, area="Incidents", check="No incidents declared", status="MANUAL CHECK", risk="MEDIUM",
                    hvpq_value=first_field(fields, "HVPQ", "incidents.other_incidents") or first_field(fields, "HVPQ", "incidents.pollution_grounding_collision_allision"),
                    piq_value=first_field(fields, "PIQ", "incidents.other_incidents"), q88_value=first_field(fields, "Q88", "incidents.other_incidents"),
                    reason="No incidents appear declared. This is not a mismatch but needs positive confirmation from vessel/office.",
                    action="Confirm no reportable machinery, navigation, mooring, pollution, security, injury or operational incidents in the last 12 months.")
    if "yes" in [hv_inc, hv_pgca, piq_inc, q_inc] and "no" in [hv_inc, hv_pgca, piq_inc, q_inc]:
        add_finding(findings, area="Incidents", check="Incident declaration mismatch", status="MISMATCH", risk="CRITICAL",
                    hvpq_value=str(hv_inc or hv_pgca), piq_value=str(piq_inc), q88_value=str(q_inc),
                    reason="One document indicates incident(s) while another indicates no incidents.", action="Verify incident register and update PIQ/HVPQ/Q88.")

    # 11. PIQ tank inspections due checks
    # User rule: if tank inspection dates are within the required cycle (normally 12 months), do NOT flag.
    # Valid/current tank inspections are reported positively in the Office Summary instead.
    for name, old_fid, freq_fid in [("Cargo/slop", "tank.cargo_slop.oldest", "tank.cargo_slop.freq_months"), ("Ballast", "tank.ballast.oldest", "tank.ballast.freq_months"), ("Void space", "tank.void.oldest", "tank.void.freq_months")]:
        old = parse_date_any(first_field(fields, "PIQ", old_fid))
        freq = first_field(fields, "PIQ", freq_fid)
        try:
            months = int(float(freq)) if freq else 12
        except Exception:
            months = 12
        if old:
            due = old + relativedelta(months=months)
            days = (due - ref_date).days
            if days < 0:
                add_finding(findings, area="Tank Inspection", check=f"{name} tank inspection overdue", status="MISMATCH", risk="HIGH",
                            piq_value=f"Oldest date: {old.isoformat()}, frequency: {months} months, due: {due.isoformat()}",
                            reason="PIQ tank inspection sequence due date is overdue based on oldest inspection date and required frequency.",
                            action="Vessel to confirm completed inspection sequence and update PIQ/supporting records immediately.")
        elif freq:
            add_finding(findings, area="Tank Inspection", check=f"{name} tank oldest inspection date missing", status="MANUAL CHECK", risk="MEDIUM",
                        piq_value=f"Frequency: {months} months; oldest inspection date not extracted",
                        reason="PIQ shows tank inspection frequency but oldest inspection date was blank/not extracted.",
                        action="Vessel to confirm the oldest inspection date in the current sequence and provide records.")

    # 12. MOC/retrofit checklist
    if normalize_bool(first_field(fields, "PIQ", "moc.retrofit")) == "yes" or normalize_bool(first_field(fields, "PIQ", "moc.equipment_replaced")) == "yes":
        add_finding(findings, area="MOC / Retrofit", check="Retrofit/replacement declared in PIQ", status="MANUAL CHECK", risk="MEDIUM",
                    piq_value=(first_field(fields, "PIQ", "moc.retrofit_details") + " " + first_field(fields, "PIQ", "moc.equipment_replaced_details"))[:500],
                    reason="PIQ declares retrofit/replacement. HVPQ/Q88/Class/certificates should reflect any affected statutory/class entries.",
                    action="Confirm MOC close-out, updated certificates/forms, class records and HVPQ/Q88 declarations.")

    # 13. Required blank/missing checks: HVPQ/Q88 core fields blank
    required_hvpq = [
        "vessel.name", "vessel.imo", "vessel.flag", "vessel.type", "classification.class_society",
        "surveys.last_drydock", "surveys.next_drydock_due", "surveys.last_special", "surveys.next_special_due",
        "psc.last_date", "psc.detained_36m", "environment.cii_rating", "environment.cii_verified_by",
        "insurance.pni_club", "classification.conditions_of_class", "classification.memo_of_class",
    ]
    for fid in required_hvpq:
        if not first_field(fields, "HVPQ", fid):
            add_finding(findings, area="Blank / Missing", check=f"HVPQ missing {FIELD_LABELS.get(fid, fid)}", status="MANUAL CHECK", risk="MEDIUM",
                        hvpq_value="blank", reason="Required/commonly observed HVPQ field is blank or not extracted.", action="Verify and complete HVPQ if applicable. HVPQ is the main document to correct.")

    # 13b. Mapped blanks in PIQ and Q88. These are not automatic defects; they are included so no blank/uncertain field is silently missed.
    has_piq = any(f.source == "PIQ" for f in fields)
    has_q88 = any(f.source == "Q88" for f in fields)
    if has_piq:
        required_piq = [
            "vessel.name", "vessel.type", "psc.last_date", "psc.detained_36m",
            "superintendent.technical.1.from", "superintendent.technical.1.to",
            "superintendent.marine.1.from", "superintendent.marine.1.to",
            "tank.cargo_slop.oldest", "tank.ballast.oldest", "tank.void.oldest",
        ]
        for fid in required_piq:
            if not first_field(fields, "PIQ", fid):
                add_finding(findings, area="Blank / Missing", check=f"PIQ missing/not extracted {FIELD_LABELS.get(fid, fid)}", status="MANUAL CHECK", risk="MEDIUM",
                            piq_value="blank/not extracted", reason="Mapped PIQ field is blank or could not be reliably extracted.",
                            action="Verify the PIQ entry manually and correct PIQ if blank/stale/wrong.")
    if has_q88:
        required_q88 = [
            "vessel.name", "vessel.imo", "vessel.type", "classification.class_society",
            "classification.conditions_of_class", "classification.memo_of_class",
            "surveys.last_drydock", "surveys.next_drydock_due",
            "surveys.last_special", "surveys.next_special_due",
            "environment.cii_rating", "environment.cii_verified_by",
            "insurance.pni_club",
        ]
        for fid in required_q88:
            if not first_field(fields, "Q88", fid):
                add_finding(findings, area="Blank / Missing", check=f"Q88 missing/not extracted {FIELD_LABELS.get(fid, fid)}", status="MANUAL CHECK", risk="MEDIUM",
                            q88_value="blank/not extracted", reason="Mapped Q88 value-add field is blank or could not be reliably extracted.",
                            action="Verify Q88 entry manually. If Q88 disagrees with HVPQ, use source evidence/Class Status before changing HVPQ.")

    # 14. Observation library is not added as a standalone checklist anymore.
    # It is merged into HVPQ check reasons later, so items that are already in order are not repeatedly sent to vessel.

    return dedupe_findings(findings)


def compare_dates_mapped(findings: List[Finding], fields: List[FieldRecord], fid: str, label: str):
    vals = sources_value(fields, fid)
    if len(vals) < 2:
        return
    parsed = {s: parse_date_any(v) for s, v in vals.items() if parse_date_any(v)}
    if len(parsed) < 2:
        return
    if max(parsed.values()) != min(parsed.values()):
        day_gap = (max(parsed.values()) - min(parsed.values())).days
        if day_gap > 7:
            add_finding(findings, area="Class / Survey", check=f"{label} date mismatch", status="MANUAL CHECK", risk="HIGH",
                        hvpq_value=vals.get("HVPQ", ""), q88_value=vals.get("Q88", ""), class_value=vals.get("CLASS", ""),
                        reason=f"Mapped survey dates differ across sources by {day_gap} days.", action="Verify latest class status/certificate and update stale declaration.")


def run_superintendent_rules(findings: List[Finding], fields: List[FieldRecord], ref_date: date):
    for kind, max_days, label in [("technical", MONTHS_7_DAYS, "Technical Superintendent"), ("marine", MONTHS_12_DAYS, "Marine Superintendent")]:
        visits = []
        for i in range(1, 5):
            frm = parse_date_any(first_field(fields, "PIQ", f"superintendent.{kind}.{i}.from"))
            to = parse_date_any(first_field(fields, "PIQ", f"superintendent.{kind}.{i}.to"))
            typ = first_field(fields, "PIQ", f"superintendent.{kind}.{i}.type")
            if frm and to:
                visits.append((frm, to, typ))
        visits = sorted(visits, key=lambda x: x[0])
        for idx in range(len(visits) - 1):
            prev_to = visits[idx][1]
            next_from = visits[idx + 1][0]
            gap = (next_from - prev_to).days
            if gap > max_days:
                add_finding(findings, area="Management Oversight", check=f"{label} inspection gap exceeds limit", status="MISMATCH", risk="CRITICAL",
                            piq_value=f"Previous to: {prev_to.isoformat()}, next from: {next_from.isoformat()}, gap: {gap} days",
                            reason=f"{label} inspection gap exceeds locked rule ({'7' if kind=='technical' else '12'} months, strict no tolerance).",
                            action="Office/vessel to provide justification and update inspection plan; flag for pre-vetting review.")
        if visits:
            last_to = visits[-1][1]
            gap = (ref_date - last_to).days
            if gap > max_days:
                add_finding(findings, area="Management Oversight", check=f"{label} last inspection overdue as of reference date", status="MISMATCH", risk="CRITICAL",
                            piq_value=f"Last to: {last_to.isoformat()}, reference: {ref_date.isoformat()}, gap: {gap} days",
                            reason=f"{label} last inspection exceeds locked interval as of reference date.", action="Arrange inspection/verify updated PIQ.")
        else:
            add_finding(findings, area="Management Oversight", check=f"{label} inspection dates not extracted", status="MANUAL CHECK", risk="HIGH",
                        piq_value="Not extracted", reason="PIQ says management oversight should be verified but visit rows were not extracted.", action="Check PIQ superintendent visit table manually.")


def dedupe_findings(findings: List[Finding]) -> List[Finding]:
    seen = set()
    out = []
    priority = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
    for f in findings:
        k = (f.area, f.check, f.hvpq_value, f.piq_value, f.class_value, f.q88_value)
        if k in seen:
            continue
        seen.add(k)
        out.append(f)
    out.sort(key=lambda x: (priority.get(x.risk.upper(), 9), x.area, x.check))
    return out

# ----------------------------- Observation library -----------------------------

def parse_obs_excel(file_obj) -> pd.DataFrame:
    if file_obj is None:
        return pd.DataFrame()
    try:
        xls = pd.ExcelFile(file_obj)
        frames = []
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet)
            df["_sheet"] = sheet
            frames.append(df)
        if not frames:
            return pd.DataFrame()
        df = pd.concat(frames, ignore_index=True)
        df["joined"] = df.fillna("").astype(str).apply(lambda r: " ".join([str(v) for v in r.tolist()]), axis=1)
        return df
    except Exception as e:
        return pd.DataFrame({"joined": [f"Observation Excel parse error: {e}"]})


def observation_checklist_from_excel(obs_df: pd.DataFrame) -> List[Tuple[str, str]]:
    if obs_df is None or obs_df.empty or "joined" not in obs_df.columns:
        return []
    text = "\n".join(obs_df["joined"].astype(str).tolist()).lower()
    checks = []
    patterns = [
        ("Mooring ropes / brake test / split drums", ["mooring", "brake", "rope", "split drum", "rendering"], "Verify mooring winch brake test dates/BHC/rendering load, rope/tail particulars, end-for-end/retirement records, and split drum declarations."),
        ("Cargo/ballast/void tank inspection records", ["tank", "coating", "ballast", "void"], "Verify tank coating/inspection dates, frequency, condition and that all cargo/slop/ballast/void spaces are covered."),
        ("Certificate dates / endorsements", ["certificate", "expiry", "annual", "endorsement"], "Verify certificate issue/expiry/endorsement dates against latest certificates/Class Status."),
        ("PSC history", ["psc", "port state", "deficien"], "Verify last three PSC dates/ports/MOU/deficiencies/detention and OCIMF PSC database entry."),
        ("CII/EEXI/EEDI", ["cii", "eexi", "eedi", "aer"], "Verify latest energy efficiency values, rating and verification basis."),
        ("Publications", ["publication", "edition", "isgott", "meg", "colreg", "mfag"], "Verify listed publication editions against onboard publication list/SMS requirement."),
        ("Firefighting / foam", ["foam", "fire", "fixed firefighting", "sample locker"], "Verify foam type/test date, fixed firefighting systems and sample locker arrangements."),
        ("Pollution prevention / overboard blanks", ["overboard", "scupper", "sea chest", "pressure test", "bunker piping", "cargo piping"], "Verify overboard discharge blanks/testing arrangement, sea chest wording, scupper plugs, and cargo/bunker pressure test records."),
        ("Cargo / IGS / venting", ["igs", "inert gas", "vent", "p/v", "vapour", "tank gauging", "cow"], "Verify IGS, venting/PV valve, vapour return, tank gauging and COW declarations."),
        ("Diagrams / arrangements", ["diagram", "manifold", "fairlead", "chock", "bitt", "bow mooring"], "Verify mooring/manifold/fairlead/chock/bitt/bow arrangement diagrams are attached/current."),
        ("Lifting gear / cranes", ["crane", "lifting", "swl", "hoist"], "Verify lifting gear/crane annual and five-year tests, SWL and certificates."),
    ]
    for title, keys, action in patterns:
        if any(k in text for k in keys):
            checks.append((title, action))
    return checks

# ----------------------------- Export helpers -----------------------------

def df_from_fields(fields: List[FieldRecord]) -> pd.DataFrame:
    return pd.DataFrame([asdict(f) for f in fields]) if fields else pd.DataFrame(columns=["source", "field_id", "label", "value", "date_value", "confidence", "raw"])


def df_from_findings(findings: List[Finding]) -> pd.DataFrame:
    return pd.DataFrame([asdict(f) for f in findings]) if findings else pd.DataFrame(columns=["area", "check", "status", "risk", "hvpq_value", "piq_value", "class_value", "q88_value", "xml_value", "reason", "action"])


def make_excel(findings: List[Finding], fields: List[FieldRecord], office_report: pd.DataFrame, vessel_actions: pd.DataFrame, obs_pinpoints: pd.DataFrame) -> bytes:
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        office_report.to_excel(writer, index=False, sheet_name="Office Summary")
        vessel_actions.to_excel(writer, index=False, sheet_name="Vessel Action Checklist")
        df_from_findings(findings).to_excel(writer, index=False, sheet_name="Detailed Findings")
        obs_pinpoints.to_excel(writer, index=False, sheet_name="Obs Library Pinpoints")
        df_from_fields(fields).to_excel(writer, index=False, sheet_name="Extracted Fields")
        # Light formatting: freeze header row and set column widths.
        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            for col in ws.columns:
                max_len = min(max((len(str(cell.value)) if cell.value is not None else 0) for cell in col) + 2, 70)
                ws.column_dimensions[col[0].column_letter].width = max(12, max_len)
    return bio.getvalue()



# ----------------------------- Office summary and vessel action helpers -----------------------------

QID_ACTION_MAP = {
    "1.1.8": ("Vessel type", "Confirm type of ship wording against Form A/B of IOPP/IOPPC and Q88. Ensure oil/product/chemical wording is consistent and Annex II carriage is correctly declared."),
    "1.1.9": ("Vessel type other-specify", "If HVPQ says 'Other', confirm the free-text vessel type is completed exactly as per certificate/Q88."),
    "1.2.1": ("EEDI", "Verify EEDI applicability, value/reason for exemption and verification basis against latest energy efficiency documentation."),
    "1.2.2": ("EEXI", "Verify EEXI rating/value and verification basis against latest EEXI technical file/certificate."),
    "1.2.3": ("CII/AER", "Verify latest CII rating, AER value and verification basis. If Q88/HVPQ blank or inconsistent, update before vetting."),
    "1.3.1": ("Registered owner", "Confirm registered owner full style, address, country, contact and IMO number against certificate/Q88/Class Status."),
    "1.3.2": ("Technical operator", "Confirm technical operator full style, company IMO number, DPA and emergency contacts against Q88/SMS."),
    "1.3.3": ("Commercial operator", "Confirm commercial/disponent operator details against current chartering data/Q88."),
    "1.4.3": ("Building contract date", "Confirm building contract date against builder/class records; correct HVPQ if date differs."),
    "1.5.1": ("Class society", "Confirm class society name and IACS status against current Class Status."),
    "1.5.2": ("Class notation", "Confirm full class notation is current and includes relevant notations such as IWS, IGS, COW, BWT, EGCS, LG etc. where applicable."),
    "1.5.4": ("Dry dock", "Check last/second-last/next dry dock dates and location against Class Status."),
    "1.5.5": ("IWS", "Check last IWS and next IWS due date. If next due is blank, confirm whether IWS remains applicable or was reset by drydock/renewal."),
    "1.5.6": ("Special survey", "Check last/next special survey dates and ESS/IACS details against Class Status."),
    "1.5.10": ("Thickness measurement", "Check date of last thickness measurement against class/CAP records."),
    "1.5.11": ("Annual survey", "Check last annual survey and next annual survey window against Class Status."),
    "1.5.12": ("Intermediate survey", "Check last intermediate survey date against Class Status."),
    "1.5.14": ("Conditions of Class", "Confirm open Conditions of Class / recommendations exactly as per Class Status; if none, ensure all docs say No/Nil."),
    "1.5.16": ("Memoranda of Class", "Confirm memoranda/notes/recommendations exactly as per Class Status; if none, ensure all docs say No/Nil."),
    "1.5.18": ("Flag/Class dispensation", "Check any flag/class dispensations or exemptions and ensure HVPQ/Q88 match current documents."),
    "1.9.2": ("Unscheduled repairs", "Confirm whether unscheduled repairs since last special survey/drydock are declared where required."),
    "1.9.3": ("Pollution/grounding/collision/allision", "Confirm no pollution, grounding, collision or allision incident in last 12 months, or declare details if any."),
    "1.9.5": ("Other incidents", "Confirm no other reportable incidents in last 12 months, including machinery, injury, mooring, security or operational incidents."),
    "1.9.6": ("Incident details", "If incidents exist, verify date/type/location/details are complete and consistent with PIQ/Q88."),
    "1.9.8": ("PSC", "Verify last PSC date, port, MOU, deficiencies, detention and OCIMF PSC database entry."),
    "2.1.5": ("Certificate dates", "Verify certificate issue/expiry/annual/intermediate/endorsement dates against latest certificates and Class Status."),
    "2.2.1": ("Publications", "Verify onboard publication editions against latest publication list/SMS and update HVPQ if stale."),
    "5.3.1": ("Fixed foam", "Verify foam type, test analysis certificate date and cargo-area fixed foam details."),
    "5.3.2": ("Fixed firefighting systems", "Verify fixed firefighting systems for paint locker, pump room, ER, lockers and other spaces."),
    "6.1.8": ("Sea chest", "Verify sea chest/cargo piping segregation wording and whether cargo sea chest is fitted/not provided."),
    "6.1.10": ("Overboard blanks/testing", "Verify overboard discharge blanks or testing arrangement and supporting records."),
    "6.1.13": ("Cargo piping pressure test", "Verify cargo piping pressure test policy, pressure and latest test record."),
    "6.1.14": ("Bunker piping pressure test", "Verify bunker piping pressure test policy, pressure and latest test record."),
    "7.1.1": ("Cargo tank coating", "Verify all cargo/slop/residual tanks are listed with coating type, extent, condition and latest inspection date."),
    "7.1.3": ("Ballast tank coating", "Verify all ballast tanks are listed with coating type, extent, condition and latest inspection date."),
    "7.1.5": ("Structural inspection programme", "Verify formal inspection programme covers void, hold, cargo and ballast spaces."),
    "10.1.4": ("Mooring winch / brake test", "Verify mooring winch particulars, split drum declaration, BHC/rendering load and latest brake test dates against records."),
    "10.1.7": ("Mooring ropes", "Verify mooring rope/tail particulars, certificates, end-for-end dates, retirement/discard criteria and records."),
    "10.2.1": ("Fairlead/chock/bitt diagram", "Verify fairlead/chock/bitt diagram is attached/current and matches deck arrangement."),
    "10.7.1": ("Bow mooring arrangement", "Verify bow mooring arrangement diagram and equipment details."),
    "10.8.1": ("Manifold arrangement", "Verify manifold arrangement diagram, dimensions and reducers/spools match vessel."),
    "10.9.1": ("Lifting gear / cranes", "Verify crane/lifting gear SWL, annual test and five-year load test dates/certificates."),
}

AREA_QIDS = {
    "Mooring ropes / brake test / split drums": ["10.1.4", "10.1.7"],
    "Cargo/ballast/void tank inspection records": ["7.1.1", "7.1.3", "7.1.5"],
    "Certificate dates / endorsements": ["2.1.5"],
    "PSC history": ["1.9.8"],
    "CII/EEXI/EEDI": ["1.2.1", "1.2.2", "1.2.3"],
    "Publications": ["2.2.1"],
    "Firefighting / foam": ["5.3.1", "5.3.2"],
    "Pollution prevention / overboard blanks": ["6.1.8", "6.1.10", "6.1.13", "6.1.14"],
    "Diagrams / arrangements": ["10.2.1", "10.7.1", "10.8.1"],
    "Lifting gear / cranes": ["10.9.1"],
}

def obs_qids_from_text(text: str) -> List[str]:
    text = clean_text(text)
    # Keep only HVPQ-looking IDs; avoid dates and decimals by requiring at least two dots OR known one-dot fields.
    candidates = re.findall(r"\b(?:item\s*)?(\d{1,2}\.\d{1,4}(?:\.\d{1,4})?)\b", text, flags=re.I)
    out = []
    for q in candidates:
        # Filter obvious decimals/versions; accept known qids or those beginning with likely HVPQ chapters.
        if q in QID_ACTION_MAP or re.match(r"^(1|2|3|4|5|6|7|8|9|10|11|12)\.", q):
            if q not in out:
                out.append(q)
    return out


def observation_pinpoint_rows(obs_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    if obs_df is None or obs_df.empty or "joined" not in obs_df.columns:
        return pd.DataFrame(columns=["Priority", "Question No.", "Topic", "What vessel should check", "Observation basis / example"])
    for _, r in obs_df.iterrows():
        joined = str(r.get("joined", ""))
        header_qids = obs_qids_from_text(" ".join([str(c) for c in obs_df.columns]))
        qids = obs_qids_from_text(joined)
        for q in header_qids + qids:
            if q in QID_ACTION_MAP:
                topic, action = QID_ACTION_MAP[q]
                rows.append({
                    "Priority": "Targeted",
                    "Question No.": q,
                    "Topic": topic,
                    "What vessel should check": action,
                    "Observation basis / example": clean_text(joined)[:450],
                })
    # Add generic area-to-question fallbacks when the observation library mentions the area but exact qid is not captured.
    text = "\n".join(obs_df["joined"].astype(str).tolist()).lower()
    for area, qids in AREA_QIDS.items():
        if area.lower().split(" /")[0] in text or any(k in text for k in normalize_key(area).split()[:2]):
            for q in qids:
                if q in QID_ACTION_MAP:
                    topic, action = QID_ACTION_MAP[q]
                    rows.append({"Priority": "Targeted", "Question No.": q, "Topic": topic, "What vessel should check": action, "Observation basis / example": f"Recurring observation library area: {area}"})
    if not rows:
        return pd.DataFrame(columns=["Priority", "Question No.", "Topic", "What vessel should check", "Observation basis / example"])
    df = pd.DataFrame(rows)
    return df.drop_duplicates(subset=["Question No.", "Topic", "What vessel should check"]).sort_values(["Question No.", "Topic"])


def build_positive_office_checks(fields: List[FieldRecord], findings: List[Finding], ref_date: date) -> pd.DataFrame:
    rows = []
    def has_bad(area_contains: str, check_contains: str = "") -> bool:
        for f in findings:
            if area_contains.lower() in f.area.lower() and check_contains.lower() in f.check.lower() and f.risk.upper() in ["CRITICAL", "HIGH"]:
                return True
        return False
    # Identity
    if not has_bad("Identity"):
        name = first_field(fields, "HVPQ", "vessel.name") or first_field(fields, "PIQ", "vessel.name") or first_field(fields, "Q88", "vessel.name") or first_field(fields, "CLASS", "vessel.name")
        imo = first_field(fields, "HVPQ", "vessel.imo") or first_field(fields, "XML", "vessel.imo") or first_field(fields, "Q88", "vessel.imo")
        if name or imo:
            rows.append({"Section": "Identity", "Status": "In order", "Office summary": f"Vessel identity broadly matches across uploaded documents: {name} / IMO {imo}."})
    # Tank inspections: positive if within cycle
    for title, old_fid, freq_fid in [("Cargo/slop tank inspections", "tank.cargo_slop.oldest", "tank.cargo_slop.freq_months"), ("Ballast tank inspections", "tank.ballast.oldest", "tank.ballast.freq_months"), ("Void space inspections", "tank.void.oldest", "tank.void.freq_months")]:
        old = parse_date_any(first_field(fields, "PIQ", old_fid))
        freq = first_field(fields, "PIQ", freq_fid)
        months = 12
        try:
            months = int(float(freq)) if freq else 12
        except Exception:
            pass
        if old:
            due = old + relativedelta(months=months)
            if due >= ref_date:
                rows.append({"Section": "Tank inspection", "Status": "In order", "Office summary": f"{title}: oldest date {old.isoformat()}, frequency {months} months, next due {due.isoformat()} — within required cycle."})
    # Superintendent
    if not has_bad("Management Oversight", "Technical"):
        v = first_field(fields, "PIQ", "superintendent.technical.1.to") or first_field(fields, "PIQ", "superintendent.technical.1.from")
        if v:
            rows.append({"Section": "Management oversight", "Status": "In order / no high-risk gap detected", "Office summary": f"Technical Superintendent visit data extracted; no high-risk gap detected by locked 7-month rule. Latest extracted date: {v}."})
    if not has_bad("Management Oversight", "Marine"):
        v = first_field(fields, "PIQ", "superintendent.marine.1.to") or first_field(fields, "PIQ", "superintendent.marine.1.from")
        if v:
            rows.append({"Section": "Management oversight", "Status": "In order / no high-risk gap detected", "Office summary": f"Marine Superintendent visit data extracted; no high-risk gap detected by locked 12-month rule. Latest extracted date: {v}."})
    # Certificates validity
    cert_exp = []
    for f in fields:
        if f.field_id.startswith("cert.") and f.field_id.endswith(".expiry"):
            d = parse_date_any(f.value)
            if d:
                cert_exp.append((f.field_id.split(".")[1], f.source, d))
    expired = [(c,s,d) for c,s,d in cert_exp if d < ref_date]
    high_cert_issues = [f.check for f in findings if f.area == "Certificates" and f.risk.upper() in ["CRITICAL", "HIGH"]]
    if cert_exp:
        if not high_cert_issues and not expired:
            rows.append({"Section": "Certificates", "Status": "In order", "Office summary": f"Extracted certificate expiry dates are valid as of {ref_date.isoformat()}; no high-risk certificate issue detected."})
        else:
            rows.append({"Section": "Certificates", "Status": "Exceptions noted", "Office summary": "Certificate review completed; exceptions are listed in the issue summary/vessel tab."})
    # Class conditions / memo
    coc_vals = sources_value(fields, "classification.conditions_of_class")
    memo_vals = sources_value(fields, "classification.memo_of_class")
    if coc_vals and all(normalize_bool(v) in ["no", "", "na"] for v in coc_vals.values()):
        rows.append({"Section": "Class", "Status": "In order", "Office summary": "Conditions of Class appear declared as nil/no in extracted sources."})
    if memo_vals and all(normalize_bool(v) in ["no", "", "na"] for v in memo_vals.values()):
        rows.append({"Section": "Class", "Status": "In order", "Office summary": "Memoranda of Class appear declared as nil/no in extracted sources."})
    return pd.DataFrame(rows, columns=["Section", "Status", "Office summary"])


def build_issue_summary(findings: List[Finding]) -> pd.DataFrame:
    rows = []
    for f in findings:
        if f.risk.upper() in ["CRITICAL", "HIGH"]:
            rows.append({
                "Priority": f.risk,
                "Area": f.area,
                "Issue": f.check,
                "Brief office note": f.reason,
                "Action": f.action,
            })
    return pd.DataFrame(rows, columns=["Priority", "Area", "Issue", "Brief office note", "Action"])


def make_vessel_action_checklist(findings: List[Finding], obs_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for f in findings:
        if f.risk.upper() in ["CRITICAL", "HIGH", "MEDIUM"]:
            qno = ""
            # Infer common qnos from area/check.
            check_norm = normalize_key(f.check + " " + f.area)
            if "cofr" in check_norm or "certificate" in check_norm or f.area == "Certificates": qno = "2.1.5"
            elif "technical superintendent" in check_norm or "marine superintendent" in check_norm: qno = "PIQ 2.2.1001 / 2.2.1002"
            elif "iws" in check_norm: qno = "1.5.5"
            elif "special survey" in check_norm: qno = "1.5.6"
            elif "dry dock" in check_norm: qno = "1.5.4"
            elif "cii" in check_norm: qno = "1.2.3"
            elif "p&i" in check_norm or "pni" in check_norm: qno = "1.1.13"
            elif "owner" in check_norm: qno = "1.3.1"
            elif "incident" in check_norm: qno = "1.9.3 / 1.9.5 / PIQ 5.7"
            elif "tank" in check_norm: qno = "2.3.3001-3003 / 7.1.1 / 7.1.3"
            elif "retrofit" in check_norm or "moc" in check_norm: qno = "PIQ 2.5.1002-1004"
            rows.append({
                "Priority": f.risk,
                "Question / Section": qno,
                "Area": f.area,
                "What to check": f.check,
                "Why flagged": f.reason,
                "Action requested from vessel/office": f.action,
                "HVPQ value": f.hvpq_value,
                "PIQ value": f.piq_value,
                "Class Status value": f.class_value,
                "Q88 value": f.q88_value,
            })
    obs_pin = observation_pinpoint_rows(obs_df)
    for _, r in obs_pin.iterrows():
        rows.append({
            "Priority": "Targeted check",
            "Question / Section": r.get("Question No.", ""),
            "Area": r.get("Topic", "Observation library"),
            "What to check": r.get("What vessel should check", ""),
            "Why flagged": "Recurring historical HVPQ/PIQ observation pattern. This is not a confirmed defect; it is a targeted pre-vetting check.",
            "Action requested from vessel/office": r.get("What vessel should check", ""),
            "HVPQ value": "", "PIQ value": "", "Class Status value": "", "Q88 value": "",
        })
    if not rows:
        return pd.DataFrame(columns=["Priority", "Question / Section", "Area", "What to check", "Why flagged", "Action requested from vessel/office", "HVPQ value", "PIQ value", "Class Status value", "Q88 value"])
    df = pd.DataFrame(rows)
    return df.drop_duplicates(subset=["Priority", "Question / Section", "Area", "What to check"])


def make_office_report_df(valid_df: pd.DataFrame, issues_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    rows.append({"Section": "A. Items appearing in order", "Status/Priority": "", "Summary / Issue": "", "Action": ""})
    if valid_df.empty:
        rows.append({"Section": "A. Items appearing in order", "Status/Priority": "No positive checks extracted", "Summary / Issue": "Review extracted fields and vessel checklist.", "Action": ""})
    else:
        for _, r in valid_df.iterrows():
            rows.append({"Section": r["Section"], "Status/Priority": r["Status"], "Summary / Issue": r["Office summary"], "Action": ""})
    rows.append({"Section": "B. Main issues / exceptions", "Status/Priority": "", "Summary / Issue": "Details and vessel questions are in the Vessel Action Checklist tab.", "Action": ""})
    if issues_df.empty:
        rows.append({"Section": "B. Main issues / exceptions", "Status/Priority": "No critical/high issue detected", "Summary / Issue": "Only manual targeted checks may remain for vessel confirmation.", "Action": ""})
    else:
        for _, r in issues_df.iterrows():
            rows.append({"Section": r["Area"], "Status/Priority": r["Priority"], "Summary / Issue": r["Issue"] + " — " + r["Brief office note"], "Action": r["Action"]})
    return pd.DataFrame(rows, columns=["Section", "Status/Priority", "Summary / Issue", "Action"])



# ----------------------------- v12 confidence, coverage and export helpers -----------------------------

def qno_for_finding(f: Finding) -> str:
    check_norm = normalize_key((f.check or "") + " " + (f.area or ""))
    if "cofr" in check_norm or "certificate" in check_norm or f.area == "Certificates": return "2.1.5"
    if "technical superintendent" in check_norm: return "PIQ 2.2.1001"
    if "marine superintendent" in check_norm: return "PIQ 2.2.1002"
    if "iws" in check_norm: return "1.5.5"
    if "special survey" in check_norm: return "1.5.6"
    if "dry dock" in check_norm or "drydock" in check_norm: return "1.5.4"
    if "annual survey" in check_norm: return "1.5.11"
    if "intermediate" in check_norm: return "1.5.12"
    if "cii" in check_norm or "aer" in check_norm: return "1.2.3"
    if "eexi" in check_norm: return "1.2.2"
    if "p&i" in check_norm or "pni" in check_norm: return "1.1.13"
    if "owner" in check_norm: return "1.3.1"
    if "type" in check_norm: return "1.1.8 / 1.1.9"
    if "incident" in check_norm: return "1.9.3 / 1.9.5 / PIQ 5.7"
    if "psc" in check_norm: return "1.9.8 / PIQ 2.8.2"
    if "tank" in check_norm: return "PIQ 2.3.3001-3003 / HVPQ 7.1.1 / 7.1.3"
    if "retrofit" in check_norm or "moc" in check_norm: return "PIQ 2.5.1002-1004"
    if "conditions of class" in check_norm: return "1.5.14"
    if "memoranda" in check_norm or "memo" in check_norm: return "1.5.16"
    return ""


def raw_text_for_source(page_cache: Dict[str, List[Tuple[int, str]]], source: str) -> str:
    return "\n".join(t for _, t in page_cache.get(source, []))


def section_text_by_qid(text: str, qid: str, max_chars: int = 2600) -> str:
    """Return text around an HVPQ-style question number. Conservative helper for vessel-facing checks."""
    if not text or not qid:
        return ""
    m = re.search(r"(?<!\d)" + re.escape(qid) + r"(?!\d)", text)
    if not m:
        return ""
    start = max(0, m.start())
    end = min(len(text), start + max_chars)
    snippet = text[start:end]
    # stop at a later question number when it is not simply a subline very close to start
    nxt = re.search(r"\n\s*\d{1,2}\.\d{1,2}(?:\.\d{1,4})?\b", snippet[len(qid)+120:])
    if nxt:
        snippet = snippet[:len(qid)+120+nxt.start()]
    return clean_text(snippet)


def latest_date_in_text(text: str) -> str:
    dates = []
    for d in extract_dates(text or ""):
        pd_dt = parse_date_any(d)
        if pd_dt:
            dates.append((pd_dt, d))
    if not dates:
        return ""
    dates.sort(reverse=True)
    return dates[0][1]


def all_dates_in_text(text: str) -> List[date]:
    out = []
    for d in extract_dates(text or ""):
        pd_dt = parse_date_any(d)
        if pd_dt:
            out.append(pd_dt)
    return out


def add_section_and_operational_fields(fields: List[FieldRecord], source: str, text: str):
    """Add raw section snippets and best-effort operational dates for mooring/brake/ropes/tails.
    These are deliberately marked as deterministic/manual-grade fields; rules will not overclaim if absent.
    """
    if not text:
        return
    qids = ["10.1.4", "10.1.7", "7.1.1", "7.1.3", "2.1.5", "1.9.8", "5.3.1", "5.3.2", "6.1.13", "6.1.14"]
    for q in qids:
        sec = section_text_by_qid(text, q)
        if sec:
            add_field(fields, source, f"section.{q}", sec, label=f"Section {q}", raw=sec, confidence="section-snippet")
    # Brake test: look around brake test keywords, fall back to HVPQ 10.1.4 section
    brake_windows = []
    for m in re.finditer(r"brake\s+test|brake\s+holding|BHC|rendering", text, re.I):
        brake_windows.append(text[max(0, m.start()-800):min(len(text), m.end()+1800)])
    if not brake_windows:
        sec = section_text_by_qid(text, "10.1.4")
        if sec:
            brake_windows.append(sec)
    if brake_windows:
        joined = " ".join(brake_windows[:3])
        dt = latest_date_in_text(joined)
        if dt:
            add_field(fields, source, "mooring.brake_test_date", dt, label="Latest brake test date found", raw=clean_text(joined[:1600]), confidence="best-effort")
        add_field(fields, source, "mooring.brake_section", clean_text(joined[:2000]), label="Brake/mooring section", raw=clean_text(joined[:2000]), confidence="section-snippet")
    # Rope / tail windows
    rope_windows = []
    for m in re.finditer(r"mooring\s+rope|rope\s+certificate|date\s+of\s+installation|end[- ]?for[- ]?end|tail|pennant", text, re.I):
        rope_windows.append(text[max(0, m.start()-700):min(len(text), m.end()+1700)])
    if not rope_windows:
        sec = section_text_by_qid(text, "10.1.7")
        if sec:
            rope_windows.append(sec)
    if rope_windows:
        joined = " ".join(rope_windows[:4])
        dates = all_dates_in_text(joined)
        if dates:
            # Store newest visible installation/service date; summary will use this only as a broad confidence check.
            newest = max(dates)
            add_field(fields, source, "mooring.ropes.latest_visible_date", newest.isoformat(), label="Latest rope/tail visible date", raw=clean_text(joined[:1800]), confidence="best-effort")
        add_field(fields, source, "mooring.ropes_section", clean_text(joined[:2200]), label="Rope/tail section", raw=clean_text(joined[:2200]), confidence="section-snippet")


def hvpq_qid_status_df(obs_df: pd.DataFrame, hvpq_text: str) -> pd.DataFrame:
    base_cols = ["Priority", "Question No.", "Topic", "HVPQ check status", "What vessel/office should check", "HVPQ evidence excerpt", "Observation basis / example"]
    obs_pin = observation_pinpoint_rows(obs_df)
    if obs_pin.empty:
        return pd.DataFrame(columns=base_cols)
    rows = []
    text = hvpq_text or ""
    low = text.lower()
    for _, r in obs_pin.iterrows():
        q = str(r.get("Question No.", "")).strip()
        topic = str(r.get("Topic", "")).strip()
        action = str(r.get("What vessel should check", "")).strip()
        basis = str(r.get("Observation basis / example", "")).strip()
        status = "Could not reliably check - exact question number not found in HVPQ extraction"
        excerpt = ""
        if q:
            m = re.search(r"(?<!\d)" + re.escape(q) + r"(?!\d)", text)
            if m:
                start = max(0, m.start() - 80)
                end = min(len(text), m.start() + 900)
                snippet = clean_text(text[start:end])
                # cut before next major qid after the current one, if visible
                rel = snippet.find(q)
                if rel >= 0:
                    after = snippet[rel+len(q):]
                    nxt = re.search(r"\b\d{1,2}\.\d{1,2}(?:\.\d{1,4})?\b", after)
                    if nxt and nxt.start() > 80:
                        snippet = snippet[:rel+len(q)+nxt.start()]
                excerpt = snippet[:700]
                # basic blank/answer heuristic
                local = excerpt.lower()
                answer_tokens = [" yes", " no", "not applicable", " n/a", " na", " date", " issued", " expires", "expiry", "good", "class", "annual", "months", "202", "201", "bar", "mt", "usd"]
                if any(tok in local for tok in answer_tokens) and len(excerpt) > 80:
                    status = "Found in HVPQ extraction - value/answer appears present; verify exact correctness"
                else:
                    status = "Found in HVPQ extraction but answer could not be reliably confirmed; manual check required"
        rows.append({
            "Priority": "Observation-led check",
            "Question No.": q,
            "Topic": topic,
            "HVPQ check status": status,
            "What vessel/office should check": action,
            "HVPQ evidence excerpt": excerpt,
            "Observation basis / example": basis,
        })
    df = pd.DataFrame(rows, columns=base_cols)
    return df.drop_duplicates(subset=["Question No.", "Topic", "What vessel/office should check"])


def field_exists(fields: List[FieldRecord], field_id: str, sources: Optional[List[str]] = None) -> bool:
    for f in fields:
        if f.field_id == field_id and str(f.value).strip():
            if sources is None or f.source in sources:
                return True
    return False


def count_fields(fields: List[FieldRecord], prefix: str, sources: Optional[List[str]] = None) -> int:
    return sum(1 for f in fields if f.field_id.startswith(prefix) and str(f.value).strip() and (sources is None or f.source in sources))


def build_coverage_matrix(fields: List[FieldRecord], findings: List[Finding], obs_df: pd.DataFrame, hvpq_text: str) -> pd.DataFrame:
    rows = []
    def add(area, check, status, basis, manual_action=""):
        rows.append({"Area": area, "Check performed": check, "Status": status, "Basis / extracted evidence": basis, "Manual action if not reliable": manual_action})
    # Core document extraction
    add("Source extraction", "HVPQ identity and core fields", "Checked" if field_exists(fields,"vessel.imo",["HVPQ","XML"]) else "Could not reliably check", first_field(fields,"HVPQ","vessel.name") or first_field(fields,"XML","vessel.name") or "", "Confirm correct HVPQ was uploaded and PDF/XML extraction succeeded.")
    add("Source extraction", "PIQ fields", "Checked" if field_exists(fields,"vessel.name",["PIQ"]) else "Could not reliably check", first_field(fields,"PIQ","vessel.name"), "Upload PIQ PDF or check parser output.")
    add("Source extraction", "Class Status certificate/survey fields", "Checked" if count_fields(fields,"cert.",["CLASS"]) >= 3 or field_exists(fields,"classification.class_society",["CLASS"]) else "Could not reliably check", f"Class cert fields extracted: {count_fields(fields,'cert.',['CLASS'])}", "Class Status is authority for certificate/survey dates; verify manually if parser did not extract rows.")
    add("Source extraction", "Q88 value-add fields", "Checked" if field_exists(fields,"vessel.imo",["Q88"]) or count_fields(fields,"cert.",["Q88"]) >= 3 else "Not uploaded / could not reliably check", f"Q88 cert fields extracted: {count_fields(fields,'cert.',['Q88'])}", "Q88 is value-add only. Upload Q88 if available; mismatches should be highlighted separately.")
    # HVPQ as primary correction target
    add("HVPQ correction focus", "Certificate dates in HVPQ compared against Class Status where available", "Checked" if any(f.area=="Certificates" for f in findings) or count_fields(fields,"cert.",["HVPQ"]) else "Could not reliably check", f"HVPQ cert fields: {count_fields(fields,'cert.',['HVPQ'])}; Class cert fields: {count_fields(fields,'cert.',['CLASS'])}", "If Class Status rows are not parsed, check certificate table manually against HVPQ 2.1.5.")
    add("HVPQ correction focus", "Conditions/Memoranda/Dispensation declarations", "Checked" if any(field_exists(fields,fid,["HVPQ","CLASS","Q88"]) for fid in ["classification.conditions_of_class","classification.memo_of_class","classification.flag_dispensation"]) else "Could not reliably check", "Compared HVPQ/Q88/Class Status where extracted.", "Manually check Class Status for open conditions, memoranda, recommendations, exemptions, dispensations.")
    # PIQ-focused checks
    add("PIQ consistency", "Technical/Marine superintendent intervals", "Checked" if any(f.field_id.startswith("superintendent.") for f in fields) else "Could not reliably check", "Uses locked TS 7-month and MS 12-month strict rules.", "Check PIQ 2.2.1001 / 2.2.1002 visit rows manually.")
    add("PIQ consistency", "Tank inspection cycle", "Checked" if any(f.field_id.startswith("tank.") for f in fields) else "Could not reliably check", "Tank dates within cycle are not flagged as findings.", "Check PIQ 2.3.3001-3003 manually if dates not extracted.")
    add("PIQ/HVPQ/Q88", "Incident declarations / no-incident positive confirmation", "Checked" if any(f.field_id.startswith("incident.") for f in fields) else "Could not reliably check", "If all declare no incident, a positive vessel confirmation row is created.", "Confirm no reportable injury, machinery, mooring, navigation, pollution, security or operational incident in last 12 months.")
    # Observation library
    obs_q = hvpq_qid_status_df(obs_df, hvpq_text)
    if obs_df is None or obs_df.empty:
        add("Observation library", "Historical observation question-number checks", "Not uploaded", "Built-in observation library is active.", "Built-in observation priorities are already available.")
    elif obs_q.empty:
        add("Observation library", "Historical observation question-number checks", "Could not reliably check", "No exact question numbers captured from observation library.", "Review the embedded observation knowledge base if additional question numbers are required.")
    else:
        missing = int(obs_q["HVPQ check status"].astype(str).str.contains("not found|could not reliably", case=False, regex=True).sum())
        add("Observation library", "Historical observation question-number checks", "Checked with manual gaps" if missing else "Checked", f"Generated {len(obs_q)} targeted HVPQ question checks; {missing} require manual confirmation.", "Review the Observation QID Checks tab for exact HVPQ question-level checks.")
    return pd.DataFrame(rows)


def make_hvpq_correction_register(findings: List[Finding]) -> pd.DataFrame:
    rows = []
    for f in findings:
        # Primary correction register: HVPQ vs authoritative/PIQ mismatches and high/medium manual checks.
        if f.risk.upper() not in ["CRITICAL", "HIGH", "MEDIUM"]:
            continue
        source_logic = ""
        action = f.action
        if f.class_value:
            source_logic = "Class Status is authoritative for certificate/class/survey data"
            if "update" not in action.lower() and f.hvpq_value:
                action = "If Class Status is current, correct HVPQ accordingly. " + action
        elif f.piq_value:
            source_logic = "PIQ cross-check against HVPQ / operational declaration"
        elif f.q88_value:
            source_logic = "Q88 value-add cross-check; not automatically authoritative"
        else:
            source_logic = "Manual verification required"
        rows.append({
            "Priority": f.risk,
            "HVPQ / PIQ question or section": qno_for_finding(f),
            "Area": f.area,
            "Issue / check": f.check,
            "Status": f.status,
            "Source logic": source_logic,
            "HVPQ value to review/correct": f.hvpq_value,
            "Class Status value (authoritative where applicable)": f.class_value,
            "PIQ value": f.piq_value,
            "Q88 value-add value": f.q88_value,
            "Reason": f.reason,
            "Required action": action,
        })
    cols = ["Priority","HVPQ / PIQ question or section","Area","Issue / check","Status","Source logic","HVPQ value to review/correct","Class Status value (authoritative where applicable)","PIQ value","Q88 value-add value","Reason","Required action"]
    return pd.DataFrame(rows, columns=cols)


def make_q88_value_add(findings: List[Finding]) -> pd.DataFrame:
    rows = []
    for f in findings:
        if f.q88_value and (not f.class_value or not f.hvpq_value or not semantically_equivalent(f.hvpq_value, f.q88_value, f.check)):
            rows.append({
                "Priority": f.risk,
                "Question / Section": qno_for_finding(f),
                "Area": f.area,
                "Q88 value-add issue": f.check,
                "HVPQ value": f.hvpq_value,
                "Q88 value": f.q88_value,
                "Class Status value": f.class_value,
                "Interpretation": "Q88 is a value-add cross-check. Do not blindly change HVPQ based on Q88 alone unless Class Status/certificate/source evidence supports it.",
                "Action": f.action,
            })
    return pd.DataFrame(rows, columns=["Priority","Question / Section","Area","Q88 value-add issue","HVPQ value","Q88 value","Class Status value","Interpretation","Action"])


def make_manual_unchecked(coverage_df: pd.DataFrame, obs_qid_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    if not coverage_df.empty:
        bad = coverage_df[coverage_df["Status"].astype(str).str.contains("Could not reliably|Not uploaded", case=False, regex=True, na=False)]
        for _, r in bad.iterrows():
            rows.append({"Priority": "Manual verification", "Area": r["Area"], "Item not reliably checked": r["Check performed"], "Why": r["Basis / extracted evidence"], "What to do": r["Manual action if not reliable"]})
    if not obs_qid_df.empty:
        bad2 = obs_qid_df[obs_qid_df["HVPQ check status"].astype(str).str.contains("could not reliably|not found|manual", case=False, regex=True, na=False)]
        for _, r in bad2.iterrows():
            rows.append({"Priority": "Observation-led manual check", "Area": r.get("Topic",""), "Item not reliably checked": f"HVPQ {r.get('Question No.','')}", "Why": r.get("HVPQ check status",""), "What to do": r.get("What vessel/office should check","")})
    return pd.DataFrame(rows, columns=["Priority","Area","Item not reliably checked","Why","What to do"])


def make_vessel_action_checklist(findings: List[Finding], obs_df: pd.DataFrame, hvpq_text: str = "") -> pd.DataFrame:
    rows = []
    for f in findings:
        if f.risk.upper() in ["CRITICAL", "HIGH", "MEDIUM"]:
            rows.append({
                "Priority": f.risk,
                "Question / Section": qno_for_finding(f),
                "Area": f.area,
                "What vessel/office should check": f.check,
                "Why it is being asked": f.reason,
                "Action requested": f.action,
                "HVPQ value": f.hvpq_value,
                "PIQ value": f.piq_value,
                "Class Status value": f.class_value,
                "Q88 value": f.q88_value,
            })
    obs_qid = hvpq_qid_status_df(obs_df, hvpq_text)
    for _, r in obs_qid.iterrows():
        rows.append({
            "Priority": "Targeted check",
            "Question / Section": r.get("Question No.", ""),
            "Area": r.get("Topic", "Observation library"),
            "What vessel/office should check": r.get("What vessel/office should check", ""),
            "Why it is being asked": f"Historical observation pattern. HVPQ extraction status: {r.get('HVPQ check status','')}",
            "Action requested": "Verify the HVPQ answer and supporting evidence; correct HVPQ if wrong/blank/stale.",
            "HVPQ value": r.get("HVPQ evidence excerpt", ""),
            "PIQ value": "", "Class Status value": "", "Q88 value": "",
        })
    if not rows:
        return pd.DataFrame(columns=["Priority", "Question / Section", "Area", "What vessel/office should check", "Why it is being asked", "Action requested", "HVPQ value", "PIQ value", "Class Status value", "Q88 value"])
    df = pd.DataFrame(rows)
    return df.drop_duplicates(subset=["Priority", "Question / Section", "Area", "What vessel/office should check"])


def make_excel(findings: List[Finding], fields: List[FieldRecord], vessel_actions: pd.DataFrame, obs_qids: pd.DataFrame, coverage_df: pd.DataFrame, manual_df: pd.DataFrame, q88_df: pd.DataFrame, hvpq_register: pd.DataFrame) -> bytes:
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        vessel_actions.to_excel(writer, index=False, sheet_name="Vessel Register")
        hvpq_register.to_excel(writer, index=False, sheet_name="HVPQ PIQ Issues")
        manual_df.to_excel(writer, index=False, sheet_name="Manual Confirmation")
        q88_df.to_excel(writer, index=False, sheet_name="Q88 Value Add")
        coverage_df.to_excel(writer, index=False, sheet_name="Coverage Matrix")
        obs_qids.to_excel(writer, index=False, sheet_name="Observation Q Checks")
        df_from_findings(findings).to_excel(writer, index=False, sheet_name="Detailed Findings")
        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            # Header style
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            widths = {
                "A": 18, "B": 22, "C": 24, "D": 34, "E": 22, "F": 44, "G": 44, "H": 44, "I": 44, "J": 44, "K": 52, "L": 52
            }
            for col_idx, col in enumerate(ws.columns, start=1):
                letter = get_column_letter(col_idx)
                max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
                ws.column_dimensions[letter].width = min(max(widths.get(letter, 12), min(max_len + 2, 55)), 60)
            # readable row heights for wrapped text
            for r in range(2, min(ws.max_row, 250) + 1):
                ws.row_dimensions[r].height = 48
    return bio.getvalue()


# ----------------------------- Simple UI summary helpers -----------------------------

def _risk_n(findings: List[Finding], risk: str) -> int:
    return sum(1 for f in findings if f.risk.upper() == risk.upper())


def _area_has_issue(findings: List[Finding], area_keyword: str, risks=("CRITICAL", "HIGH", "MEDIUM")) -> bool:
    k = area_keyword.lower()
    return any(k in f.area.lower() and f.risk.upper() in risks for f in findings)


def build_human_review_summary(fields: List[FieldRecord], findings: List[Finding], coverage_df: pd.DataFrame, manual_df: pd.DataFrame, obs_qid_df: pd.DataFrame, ref_date: date) -> Dict[str, str]:
    checked = []
    if field_exists(fields, "vessel.imo", ["HVPQ", "XML"]): checked.append("HVPQ identity/core particulars")
    if count_fields(fields, "cert.", ["HVPQ"]): checked.append("HVPQ certificate table")
    if count_fields(fields, "cert.", ["CLASS"]): checked.append("Class Status certificate/survey dates")
    if field_exists(fields, "vessel.imo", ["Q88"]) or count_fields(fields, "cert.", ["Q88"]): checked.append("Q88 value-add fields")
    if any(f.field_id.startswith("superintendent.") for f in fields): checked.append("PIQ superintendent visit gaps")
    if any(f.field_id.startswith("tank.") for f in fields): checked.append("PIQ tank inspection cycles")
    if any(f.field_id.startswith("incidents.") for f in fields): checked.append("incident/no-incident declarations")
    if not obs_qid_df.empty: checked.append("historical observation question numbers")
    checked_text = ", ".join(checked) if checked else "the uploaded documents, subject to extraction quality"

    ok_bits = []
    # Tank inspections ok if no high tank issue and PIQ dates exist within cycle
    tank_ok = []
    for title, old_fid, freq_fid in [("cargo/slop", "tank.cargo_slop.oldest", "tank.cargo_slop.freq_months"), ("ballast", "tank.ballast.oldest", "tank.ballast.freq_months"), ("void", "tank.void.oldest", "tank.void.freq_months")]:
        old = parse_date_any(first_field(fields, "PIQ", old_fid))
        freq = first_field(fields, "PIQ", freq_fid)
        months = 12
        try: months = int(float(freq)) if freq else 12
        except Exception: pass
        if old and old + relativedelta(months=months) >= ref_date:
            tank_ok.append(title)
    if tank_ok: ok_bits.append(f"tank inspection cycles appear within required interval for {', '.join(tank_ok)} tanks")
    if not _area_has_issue(findings, "Management Oversight", ("CRITICAL", "HIGH")) and any(f.field_id.startswith("superintendent.") for f in fields):
        ok_bits.append("no high-risk superintendent interval breach was detected from extracted PIQ dates")
    if not _area_has_issue(findings, "Certificates", ("CRITICAL", "HIGH")) and count_fields(fields, "cert.", ["HVPQ"]):
        ok_bits.append("no high-risk certificate mismatch/expiry issue was detected from extracted certificate rows")
    if not _area_has_issue(findings, "Class", ("CRITICAL", "HIGH")) and any(field_exists(fields, fid, ["HVPQ", "CLASS", "Q88"]) for fid in ["classification.conditions_of_class", "classification.memo_of_class"]):
        ok_bits.append("Class Conditions/Memoranda were checked where extractable")
    ok_text = "; ".join(ok_bits) if ok_bits else "no clean 'in-order' conclusion is shown where extraction was insufficient; those items are moved to manual confirmation instead"

    main_issues = [f for f in findings if f.risk.upper() in ["CRITICAL", "HIGH"]]
    if main_issues:
        issue_text = "; ".join([f"{f.area}: {f.check}" for f in main_issues[:8]])
        if len(main_issues) > 8: issue_text += f"; plus {len(main_issues)-8} more high-priority item(s)"
    else:
        issue_text = "no critical/high-priority mismatch was detected from the extracted mapped fields"

    if not manual_df.empty:
        manual_text = f"{len(manual_df)} item(s) could not be reliably checked or need positive confirmation. These are included in the Manual Confirmation and Vessel Action Checklist tabs so they are not missed."
    else:
        manual_text = "no major extraction-gap/manual-confirmation item was generated from the current upload."

    if obs_qid_df.empty:
        obs_text = "No observation-library question-number checks were generated. The embedded observation library did not generate question checks for this item."
    else:
        review = obs_qid_df[obs_qid_df["HVPQ check status"].astype(str).str.contains("could not reliably|not found|manual", case=False, regex=True, na=False)]
        ok = len(obs_qid_df) - len(review)
        obs_text = f"{len(obs_qid_df)} repeated-observation HVPQ question check(s) were generated from the observation sheet. {ok} were located with an apparent answer/excerpt; {len(review)} need manual review or clearer evidence."

    return {
        "checked": f"Checked: {checked_text}.",
        "ok": f"Appearing in order: {ok_text}.",
        "issues": f"Not satisfactory / requires correction: {issue_text}.",
        "manual": f"Could not reliably check / manual confirmation: {manual_text}",
        "obs": f"Repeat-observation coverage: {obs_text}",
    }


def build_major_repeat_summary(obs_qid_df: pd.DataFrame) -> pd.DataFrame:
    cols = ["Question No.", "Topic", "Review status", "What to check", "HVPQ evidence excerpt"]
    if obs_qid_df is None or obs_qid_df.empty:
        return pd.DataFrame(columns=cols)
    rows = []
    for _, r in obs_qid_df.iterrows():
        stt = str(r.get("HVPQ check status", ""))
        if re.search(r"could not reliably|not found|manual", stt, flags=re.I):
            status = "Needs review"
        else:
            status = "Located in HVPQ - verify correctness"
        rows.append({
            "Question No.": r.get("Question No.", ""),
            "Topic": r.get("Topic", ""),
            "Review status": status,
            "What to check": r.get("What vessel/office should check", ""),
            "HVPQ evidence excerpt": r.get("HVPQ evidence excerpt", ""),
        })
    return pd.DataFrame(rows, columns=cols)


def style_priority_dataframe(df: pd.DataFrame):
    if df is None or df.empty:
        return df
    def row_style(row):
        val = " ".join(str(x).upper() for x in row.values)
        if "CRITICAL" in val:
            return ['background-color: #fde2e1'] * len(row)
        if "HIGH" in val:
            return ['background-color: #fff1cc'] * len(row)
        if "NEEDS REVIEW" in val or "MANUAL" in val or "COULD NOT" in val:
            return ['background-color: #eef4ff'] * len(row)
        return [''] * len(row)
    return df.style.apply(row_style, axis=1)


# ----------------------------- v14 source-specific registers -----------------------------

def status_rank(status: str) -> int:
    s = str(status).lower()
    if any(x in s for x in ["critical", "not satisfactory", "issue", "mismatch", "overdue", "expired"]): return 0
    if any(x in s for x in ["manual", "could not", "review", "blank"]): return 1
    if any(x in s for x in ["in order", "ok", "checked"]): return 2
    return 3


def _ok_or_manual_date(label: str, source: str, field_id: str, fields: List[FieldRecord], ref_date: date, months: int, qno: str, area: str) -> Dict[str, str]:
    val = first_field(fields, source, field_id)
    d = parse_date_any(val)
    if not d:
        return {"Priority":"Manual", "Question / Section":qno, "Area":area, "Check":label, "Status":"Could not reliably check", "Document value":val or "Not extracted/blank", "Reference value":"", "Finding / interpretation":f"{label} date was not reliably extracted.", "Action requested":"Vessel/office to verify supporting record and update HVPQ/PIQ/Q88 if blank or stale."}
    due = d + relativedelta(months=months)
    if due >= ref_date:
        return {"Priority":"OK", "Question / Section":qno, "Area":area, "Check":label, "Status":"In order", "Document value":str(val), "Reference value":f"Due not before {due.isoformat()}", "Finding / interpretation":f"{label} appears within {months} months as of {ref_date.isoformat()}.", "Action requested":"No correction indicated from extracted data; keep evidence ready onboard."}
    return {"Priority":"High", "Question / Section":qno, "Area":area, "Check":label, "Status":"Not satisfactory", "Document value":str(val), "Reference value":f"Due {due.isoformat()}", "Finding / interpretation":f"{label} appears older than permitted {months}-month interval.", "Action requested":"Verify latest record. If no newer record exists, complete/arrange check and update HVPQ/PIQ/Q88."}


def cert_validity_rows(fields: List[FieldRecord], ref_date: date) -> List[Dict[str, str]]:
    rows=[]
    cert_keys=set()
    for f in fields:
        m=re.match(r"cert\.([^.]+)\.expiry$", f.field_id)
        if m and f.source in ["HVPQ","CLASS"]:
            cert_keys.add(m.group(1))
    expired=[]; valid=[]; manual=[]
    for key in sorted(cert_keys):
        # Class Status is reference/authority when available, else HVPQ.
        cval = first_field(fields,"CLASS",f"cert.{key}.expiry")
        hval = first_field(fields,"HVPQ",f"cert.{key}.expiry")
        use = cval or hval
        d=parse_date_any(use)
        label=key.upper().replace('_',' ')
        if not d:
            manual.append(label); continue
        if d < ref_date:
            expired.append((label,use,cval,hval))
        else:
            valid.append(label)
        if cval and hval and parse_date_any(cval) and parse_date_any(hval) and parse_date_any(cval)!=parse_date_any(hval):
            rows.append({"Priority":"High", "Question / Section":"2.1.5", "Area":"Certificates", "Check":f"{label} expiry mismatch HVPQ vs Class Status", "Status":"Not satisfactory", "Document value":hval, "Reference value":cval, "Finding / interpretation":"HVPQ certificate expiry differs from Class Status. Class Status is the reference for certificate/survey dates.", "Action requested":"Verify latest certificate/Class Status and correct HVPQ if stale/wrong."})
    if expired:
        for label,use,cval,hval in expired:
            rows.append({"Priority":"Critical", "Question / Section":"2.1.5", "Area":"Certificates", "Check":f"{label} certificate validity", "Status":"Not satisfactory", "Document value":hval or use, "Reference value":cval or use, "Finding / interpretation":"Certificate expiry appears before the review date.", "Action requested":"Verify certificate immediately and update HVPQ/Q88 as applicable."})
    elif valid:
        rows.append({"Priority":"OK", "Question / Section":"2.1.5", "Area":"Certificates", "Check":"Certificate validity", "Status":"In order", "Document value":f"{len(valid)} certificate expiry date(s) checked", "Reference value":"Class Status where available, otherwise HVPQ", "Finding / interpretation":"No expired certificate detected from extracted certificate expiry dates.", "Action requested":"Keep latest certificates/Class Status ready for inspection."})
    if manual:
        rows.append({"Priority":"Manual", "Question / Section":"2.1.5", "Area":"Certificates", "Check":"Certificate rows not fully extracted", "Status":"Could not reliably check", "Document value":", ".join(manual[:20]), "Reference value":"", "Finding / interpretation":"Some certificate expiry dates were not reliably extracted.", "Action requested":"Manually verify those certificate dates against Class Status/latest certificates."})
    return rows


def build_piq_checks(fields: List[FieldRecord], findings: List[Finding], ref_date: date) -> pd.DataFrame:
    rows=[]
    # TS / MS intervals from findings and explicit OK rows if no breach.
    mgmt_findings=[f for f in findings if f.area == "Management Oversight"]
    for f in mgmt_findings:
        rows.append({"Priority":f.risk, "Question / Section":qno_for_finding(f), "Area":f.area, "Check":f.check, "Status":f.status, "Document value":f.piq_value, "Reference value":"TS max 7 months / MS max 12 months", "Finding / interpretation":f.reason, "Action requested":f.action})
    if any(f.field_id.startswith("superintendent.technical") for f in fields) and not any("Technical" in f.check and f.risk.upper() in ["CRITICAL","HIGH"] for f in mgmt_findings):
        rows.append({"Priority":"OK", "Question / Section":"PIQ 2.2.1001", "Area":"Management Oversight", "Check":"Technical Superintendent visit interval", "Status":"In order", "Document value":first_field(fields,"PIQ","superintendent.technical.raw")[:350], "Reference value":"Maximum gap 7 months", "Finding / interpretation":"No breach detected from extracted Technical Superintendent visit rows.", "Action requested":"No correction indicated; keep visit reports available."})
    if any(f.field_id.startswith("superintendent.marine") for f in fields) and not any("Marine" in f.check and f.risk.upper() in ["CRITICAL","HIGH"] for f in mgmt_findings):
        rows.append({"Priority":"OK", "Question / Section":"PIQ 2.2.1002", "Area":"Management Oversight", "Check":"Marine Superintendent visit interval", "Status":"In order", "Document value":first_field(fields,"PIQ","superintendent.marine.raw")[:350], "Reference value":"Maximum gap 12 months", "Finding / interpretation":"No breach detected from extracted Marine Superintendent visit rows.", "Action requested":"No correction indicated; keep visit reports available."})
    # Tank intervals
    for title, old_fid, freq_fid, qno in [("Cargo/slop tank inspections", "tank.cargo_slop.oldest", "tank.cargo_slop.freq_months", "PIQ 2.3.3001"), ("Ballast tank inspections", "tank.ballast.oldest", "tank.ballast.freq_months", "PIQ 2.3.3002"), ("Void space inspections", "tank.void.oldest", "tank.void.freq_months", "PIQ 2.3.3003")]:
        old = parse_date_any(first_field(fields,"PIQ",old_fid)); freq = first_field(fields,"PIQ",freq_fid)
        months=12
        try: months=int(float(freq)) if freq else 12
        except Exception: pass
        if old:
            due=old+relativedelta(months=months)
            rows.append({"Priority":"OK" if due>=ref_date else "High", "Question / Section":qno, "Area":"Tank inspection", "Check":title, "Status":"In order" if due>=ref_date else "Not satisfactory", "Document value":str(first_field(fields,"PIQ",old_fid)), "Reference value":f"Frequency {months} months; due {due.isoformat()}", "Finding / interpretation":f"Oldest inspection date is {'within' if due>=ref_date else 'outside'} the required interval.", "Action requested":"No correction indicated; keep inspection records ready." if due>=ref_date else "Verify completed tank inspection sequence and update PIQ/supporting records."})
        else:
            rows.append({"Priority":"Manual", "Question / Section":qno, "Area":"Tank inspection", "Check":title, "Status":"Could not reliably check", "Document value":"Not extracted/blank", "Reference value":"Expected frequency/date in PIQ", "Finding / interpretation":"Oldest inspection date could not be reliably extracted.", "Action requested":"Manually check PIQ 2.3.3001-3003 and tank inspection records."})
    # MOC/retrofit, PSC, incidents, PIQ general fields/blanks
    for f in findings:
        if f.piq_value and f.area not in ["Management Oversight", "Tank Inspection"] and f.risk.upper() in ["CRITICAL","HIGH","MEDIUM"]:
            rows.append({"Priority":f.risk, "Question / Section":qno_for_finding(f), "Area":f.area, "Check":f.check, "Status":f.status, "Document value":f.piq_value, "Reference value":f.hvpq_value or f.q88_value or f.class_value, "Finding / interpretation":f.reason, "Action requested":f.action})
    for fid, label, qno in [("piq.static_nav_assessment","Static navigational assessment","PIQ 3.2.1"),("piq.dynamic_nav_assessment_shore","Dynamic navigational assessment - shore staff","PIQ 3.2.2"),("piq.cargo_audit","Comprehensive cargo audit","PIQ 3.2.5"),("piq.engineering_audit","Comprehensive engineering audit","PIQ 3.2.6"),("piq.mooring_anchoring_audit","Comprehensive mooring/anchoring audit","PIQ 3.2.7"),("moc.retrofit","MOC retrofit declaration","PIQ 2.5.1002"),("psc.last_date","Last PSC declaration","PIQ 2.8.2")]:
        val=first_field(fields,"PIQ",fid)
        rows.append({"Priority":"OK" if val else "Manual", "Question / Section":qno, "Area":"PIQ completeness", "Check":label, "Status":"Checked" if val else "Could not reliably check", "Document value":val, "Reference value":"", "Finding / interpretation":"PIQ value was extracted for review." if val else "PIQ value was not reliably extracted or may be blank.", "Action requested":"Verify entry is current and supported by evidence."})
    df=pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=["Priority","Question / Section","Area","Check","Status","Document value","Reference value","Finding / interpretation","Action requested"])
    df=df.drop_duplicates()
    df["_rank"]=df["Priority"].map(lambda x: {"Critical":0,"CRITICAL":0,"High":1,"HIGH":1,"Medium":2,"MEDIUM":2,"Manual":3,"OK":4}.get(str(x),5))
    return df.sort_values(["_rank","Area","Question / Section"]).drop(columns="_rank")


def build_hvpq_checks(fields: List[FieldRecord], findings: List[Finding], ref_date: date) -> pd.DataFrame:
    rows=[]
    for f in findings:
        if f.risk.upper() not in ["CRITICAL","HIGH","MEDIUM"]: continue
        # HVPQ target items: all class/cert/q88/hvpq mismatches and blanks, except pure PIQ operational rows handled separately.
        if f.hvpq_value or f.class_value or (f.area in ["Certificates","Class / Survey","Classification","Blank / Missing","Environment","Ownership / Operation","Insurance","Vessel Type"]):
            rows.append({"Priority":f.risk,"Question / Section":qno_for_finding(f),"Area":f.area,"Check":f.check,"Status":f.status,"HVPQ value":f.hvpq_value,"Reference source":"Class Status" if f.class_value else ("Q88" if f.q88_value else "PIQ/manual"),"Reference value":f.class_value or f.q88_value or f.piq_value,"Finding / interpretation":f.reason,"Action requested":f.action})
    # Certificate validity positive / mismatch rows
    for r in cert_validity_rows(fields, ref_date):
        rows.append({"Priority":r["Priority"],"Question / Section":r["Question / Section"],"Area":r["Area"],"Check":r["Check"],"Status":r["Status"],"HVPQ value":r["Document value"],"Reference source":"Class Status/latest certificate","Reference value":r["Reference value"],"Finding / interpretation":r["Finding / interpretation"],"Action requested":r["Action requested"]})
    # HVPQ operational recurring checks
    rows.append(_hvpq_ops_row(fields, ref_date, "Brake testing", "mooring.brake_test_date", 12, "10.1.4", "Mooring", "Latest brake test date found in HVPQ/Q88 text"))
    rows.append(_hvpq_ops_row(fields, ref_date, "Mooring ropes age / visible date", "mooring.ropes.latest_visible_date", 60, "10.1.7", "Mooring", "Latest rope/tail visible date found; verify every rope individually"))
    # Tails are harder; do not overclaim unless tail text is present.
    tail_sec = first_field(fields,"HVPQ","mooring.ropes_section") or first_field(fields,"Q88","mooring.ropes_section")
    if re.search(r"tail|pennant", tail_sec, re.I):
        rows.append(_hvpq_ops_row(fields, ref_date, "Tails within 18 months of installation", "mooring.ropes.latest_visible_date", 18, "10.1.7", "Mooring", "Tail/pennant keyword found; verify each tail date individually"))
    else:
        rows.append({"Priority":"Manual","Question / Section":"10.1.7","Area":"Mooring","Check":"Tails within 18 months of installation","Status":"Could not reliably check","HVPQ value":"Tail/pennant details not reliably extracted","Reference source":"Vessel records","Reference value":"","Finding / interpretation":"The app could not confirm tail installation/service dates from extracted text.","Action requested":"Vessel to confirm tail certificates/installation dates and whether all tails are within the applicable service interval."})
    df=pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=["Priority","Question / Section","Area","Check","Status","HVPQ value","Reference source","Reference value","Finding / interpretation","Action requested"])
    df=df.drop_duplicates()
    df["_rank"]=df["Priority"].map(lambda x: {"Critical":0,"CRITICAL":0,"High":1,"HIGH":1,"Medium":2,"MEDIUM":2,"Manual":3,"OK":4}.get(str(x),5))
    return df.sort_values(["_rank","Area","Question / Section"]).drop(columns="_rank")


def _hvpq_ops_row(fields, ref_date, label, fid, months, qno, area, interp):
    val = first_field(fields,"HVPQ",fid) or first_field(fields,"Q88",fid)
    source = "HVPQ" if first_field(fields,"HVPQ",fid) else ("Q88" if first_field(fields,"Q88",fid) else "")
    d=parse_date_any(val)
    if not d:
        return {"Priority":"Manual","Question / Section":qno,"Area":area,"Check":label,"Status":"Could not reliably check","HVPQ value":val or "Not extracted/blank","Reference source":"Vessel records","Reference value":"","Finding / interpretation":f"{label} could not be reliably verified from extracted HVPQ/Q88 text.","Action requested":"Vessel/office to verify supporting records and update HVPQ if blank/stale/wrong."}
    due=d+relativedelta(months=months)
    ok=due>=ref_date
    return {"Priority":"OK" if ok else "High","Question / Section":qno,"Area":area,"Check":label,"Status":"In order" if ok else "Not satisfactory","HVPQ value":val,"Reference source":source,"Reference value":f"Due not before {due.isoformat()} for {months}-month check","Finding / interpretation":interp + ("; appears in order from extracted date." if ok else "; appears outside expected interval."),"Action requested":"Keep evidence ready onboard." if ok else "Verify latest record and update HVPQ if stale/wrong."}


def build_q88_checks(fields: List[FieldRecord], findings: List[Finding]) -> pd.DataFrame:
    rows=[]
    for f in findings:
        if f.q88_value and f.risk.upper() in ["CRITICAL","HIGH","MEDIUM"]:
            rows.append({"Priority":f.risk,"Question / Section":qno_for_finding(f),"Area":f.area,"Check":f.check,"Status":f.status,"Q88 value":f.q88_value,"HVPQ value":f.hvpq_value,"Class/PIQ reference":f.class_value or f.piq_value,"Finding / interpretation":f.reason,"Action requested":"Q88 is value-add only. Verify source evidence; if Q88 is correct and HVPQ is stale, correct HVPQ. If Q88 is wrong, update Q88."})
    for fid,label,qno in [("environment.cii_verified_by","CII verification basis","1.2.3 / Q88 CII"),("cert.cofr.expiry","COFR expiry","Q88 2.x"),("cert.pni_cover.expiry","P&I cover expiry","Q88 1.15"),("classification.conditions_of_class","Conditions of Class","Q88 1.20"),("classification.memo_of_class","Memoranda of Class","Q88 1.20a")]:
        val=first_field(fields,"Q88",fid)
        rows.append({"Priority":"OK" if val else "Manual","Question / Section":qno,"Area":"Q88 completeness","Check":label,"Status":"Checked" if val else "Could not reliably check / blank","Q88 value":val,"HVPQ value":first_field(fields,"HVPQ",fid),"Class/PIQ reference":first_field(fields,"CLASS",fid),"Finding / interpretation":"Q88 value was extracted for cross-check." if val else "Q88 value was not reliably extracted or appears blank.","Action requested":"Verify Q88 entry if used for chartering/vetting consistency."})
    df=pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=["Priority","Question / Section","Area","Check","Status","Q88 value","HVPQ value","Class/PIQ reference","Finding / interpretation","Action requested"])
    df=df.drop_duplicates()
    df["_rank"]=df["Priority"].map(lambda x: {"Critical":0,"CRITICAL":0,"High":1,"HIGH":1,"Medium":2,"MEDIUM":2,"Manual":3,"OK":4}.get(str(x),5))
    return df.sort_values(["_rank","Area","Question / Section"]).drop(columns="_rank")


def build_repeat_obs_checks(obs_qid_df: pd.DataFrame) -> pd.DataFrame:
    if obs_qid_df is None or obs_qid_df.empty:
        return pd.DataFrame(columns=["Priority","Question / Section","Topic","HVPQ check status","What to check","HVPQ evidence excerpt","Observation basis"])
    df=obs_qid_df.rename(columns={"Question No.":"Question / Section","What vessel/office should check":"What to check","Observation basis / example":"Observation basis"}).copy()
    df["Priority"] = df["HVPQ check status"].astype(str).apply(lambda x: "Manual" if re.search(r"could not|not found|manual", x, re.I) else "Targeted check")
    return df[["Priority","Question / Section","Topic","HVPQ check status","What to check","HVPQ evidence excerpt","Observation basis"]]


def build_summary_paragraphs(hvpq_df: pd.DataFrame, q88_df: pd.DataFrame, piq_df: pd.DataFrame, obs_df: pd.DataFrame) -> Dict[str,str]:
    def collect_ok(df, n=8):
        if df is None or df.empty: return []
        ok = df[df["Status"].astype(str).str.contains("in order|checked", case=False, na=False)] if "Status" in df.columns else pd.DataFrame()
        return [f"{r.get('Area','')}: {r.get('Check','')}" for _,r in ok.head(n).iterrows()]
    def collect_bad(df, n=8):
        if df is None or df.empty: return []
        textcols = [c for c in ["Priority","Status"] if c in df.columns]
        mask = pd.Series(False, index=df.index)
        for c in textcols:
            mask = mask | df[c].astype(str).str.contains("Critical|High|Medium|Manual|not satisfactory|could not|blank|mismatch|overdue|expired", case=False, na=False)
        bad = df[mask]
        return [f"{r.get('Area', r.get('Topic',''))}: {r.get('Check', r.get('What to check',''))}" for _,r in bad.head(n).iterrows()]
    ok_items = collect_ok(hvpq_df) + collect_ok(piq_df) + collect_ok(q88_df, 4)
    bad_items = collect_bad(hvpq_df) + collect_bad(piq_df) + collect_bad(q88_df) + collect_bad(obs_df, 4)
    ok_txt = "; ".join(ok_items[:12]) if ok_items else "No positive 'in order' conclusion was generated where extraction was insufficient. Such items are listed as manual checks instead."
    bad_txt = "; ".join(bad_items[:14]) if bad_items else "No critical/high/manual gap was generated from the mapped checks."
    obs_count = 0 if obs_df is None or obs_df.empty else len(obs_df)
    return {
        "checked": "The review separately checked HVPQ correction items, Q88 value-add consistency, PIQ operational declarations, and repeat-observation question numbers from the observation sheet. Class Status was used only as the authority/reference for certificate and survey dates plus Conditions/Memoranda/dispensations.",
        "ok": "From the extracted data, the following checks appear in order: " + ok_txt,
        "bad": "Items requiring correction/review/manual confirmation include: " + bad_txt,
        "obs": f"Repeat-observation coverage generated {obs_count} HVPQ question-level check row(s). These are targeted checks, not automatic defects; vessel/office should verify supporting evidence for each question.",
    }


def make_excel_v14(hvpq_df: pd.DataFrame, q88_df: pd.DataFrame, piq_df: pd.DataFrame, obs_df: pd.DataFrame) -> bytes:
    bio=io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        sheets=[("HVPQ Checks", hvpq_df), ("Q88 Checks", q88_df), ("PIQ Checks", piq_df), ("Repeat Obs Questions", obs_df)]
        for name, df in sheets:
            df.to_excel(writer, index=False, sheet_name=name)
        for ws in writer.book.worksheets:
            ws.freeze_panes="A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font=Font(bold=True, color="FFFFFF")
                cell.fill=PatternFill("solid", fgColor="1F4E78")
                cell.alignment=Alignment(wrap_text=True, vertical="center")
            for row in ws.iter_rows(min_row=2):
                pr = str(row[0].value).upper() if row and row[0].value is not None else ""
                fill = None
                if "CRITICAL" in pr or "HIGH" in pr:
                    fill=PatternFill("solid", fgColor="FCE4D6")
                elif "MEDIUM" in pr or "MANUAL" in pr:
                    fill=PatternFill("solid", fgColor="EAF2F8")
                elif "OK" in pr:
                    fill=PatternFill("solid", fgColor="E2F0D9")
                for cell in row:
                    cell.alignment=Alignment(wrap_text=True, vertical="top")
                    if fill: cell.fill=fill
            widths={"A":14,"B":18,"C":22,"D":34,"E":18,"F":42,"G":42,"H":44,"I":54,"J":54}
            for idx, col in enumerate(ws.columns, start=1):
                letter=get_column_letter(idx)
                max_len=max((len(str(c.value)) if c.value is not None else 0) for c in col)
                ws.column_dimensions[letter].width=min(max(widths.get(letter, 12), min(max_len+2, 55)), 62)
            for r in range(2, min(ws.max_row, 300)+1):
                ws.row_dimensions[r].height=55
    return bio.getvalue()



# ----------------------------- v15 observation-informed and rules-based output -----------------------------

DEFAULT_COMPARISON_RULES_TEXT = 'I will give you 2 documents to analyze HVPQ (PDF) and PIQ Excel , we need to audit individually for errors and then compare both for inconsistencies. The audit needs to be done \nin a human like reading fashion. Here are the rules\n\nHVPQ File checks\n1. In HVPQ, Section 13 Combination carriers can be left blank.\n2. In HVPQ Section 9, subsection 6 LNG bunkers can be left blank .\n3. This is Oil ship so Chemical sections can be left blank.\n3. Remaining questions and sub questions apply your logical reasoning if parent question is NO, sub-questions can be NA or blank.\n4. Other than this, No questions are left blank, please flag if anything.\n5. No certificates should be expired.\n6. In section 2 certificates, Any certificate issued more than 1 year ago must have annual endorsement date within 1 year. Last annual date of any certificate should match with response in 1.5.11\n7. In section 2 certificates, Any certificate issued more than 2.5 year ago must have intermediate endorsement date within 2.5 years. Last intermediate date of any certificate should match with response in 1.5.12\n8. 1.1.13.4 response must be yes\n9. Vessel age basis 1.4.7 if more than 15 years, then CAP rating is applicable and must be 1.\n10. 1.5.1.2 must be "Yes"\n11. 1.5.4.1 date of last dry dock should not be more than 5 years old and must match response of 1.5.6.1\n12. If 1.9.1 is "No", 1.9.2 cannot be "No" or blank.\n13. Response for 3.2.1 should be "No"\n14. Courses mention in 3.3.4 must include something like "Engine Room Resource Management using an Engine Room Simulator", "Cargo simulator", "Ship handling".\n\n16. Response in 5.3.1.4 should not be more than 1 year old.\n17. Section 7.1.1 date of coating inspections by ship staff should not be expired basis the frequency of inspections mentioned.\n18. Section 7.1.3 date of coating inspections by competent person should not be expired basis the frequency of inspections mentioned.\n19. 7.1.4.5 response should be more than 0 and in % unit.\n20. If 1.6.1 is greater than 200, than pump type in 9.6.2 is Centrifugal else Deepwell.\n21. 10.1.4 Date of last brake test should not be expired basis frequency of testing brakes.\n22. 10.1.7 if type is wire installed date should not be older than 10 years, if type is Tails installed date should not be older than 18 months, if type is Ropes installed date should not be older than 5 years.\n23. 10.9.1 last annual test and last 5 year test dates should not be expired.\n\nPIQ File Checks\n1. Vessel Type in PIQ 1.1.1 must match with HVPQ response in 2.1.4\n3. In PIQ Chapter 3.2.1 and 3.2.2 static and Dynamic navigational assessment both cannot be Yes or NO, must be opposite response.\n4. Last navigational assessment in either of 3.2.1 or 3.2.2 should not be older than 12 months.\n5. Chapter 3.5,3.6,3.7 if marked yes, dates of last audits should not be older than 12 months.\n6. Chapter 3 subsection 3.3.1,3.3.3,3.3.4 crew training, courses mentioned must also be included in HVPQ 3.3.4\n7. Chapter 3.4.2001 date must be within 12 months and Chapter 3.4.2002 date must be within 3 months\n8. Section 5.2.4 response must match with HVPQ response 5.3.2 sub question 4.\n9. Chapter 5.7.1001 to 5.7.1029, all questions response must align with HVPQ responses 1.9.1 to 1.9.7, please analyze. Also check if PIQ Chapter 2.1.1 Purpose of visit is Damage or Other / Occasional then there may be a damage or repair verification survey here this should align with other responses.\n10. Chapter 2 sub-section 8, General information on PSC must align with HVPQ responses for PSC in 1.9.8 and 1.9.9\n11. Section 8 sub section 3 response for various questions must align with HVPQ responses in HVPQ Chapter 9 sub section 9 - Vapor emission control and 10 - Venting, please analyze.\n12. Section 10 subsection 2 question 1 response must match with HVPQ 11.3.3\n13. Section 10 subsection 2 question 3 response must match with HVPQ 11.9.1\n14. Chapter 2 Subsection 2 ques 1001 last technical suptt visit must not be older than 7 months and there should not be a gap of more than 7 months for successive visits.\n15. Chapter 2 Subsection 2 ques 1002 last marine suptt visit must not be older than 12 months and there should not be a gap of more than 12 months for successive visits.\n16. Chapter 2 Subsection 3 ques 3001 required frequency of inspection of cargo tanks and last date must align if HVPQ response 7.1.1 for frequency and inspection dates.\n17. Chapter 2 Subsection 3 ques 3002 required frequency of inspection of ballast tanks and last date must align if HVPQ response 7.1.3 for frequency and inspection dates.\n\n\n\n\n\n\n'


# ----------------------------- Embedded machine-readable knowledge base v18 -----------------------------
# This replaces the older manual upload of observation libraries, incident libraries and rule TXT files.
# Observations are repeat-risk signals only; they never create a finding by themselves.
EMBEDDED_KNOWLEDGE_BASE = {'schema_version': '2026-05-17.v18', 'description': 'Machine-readable embedded observation and validation knowledge base for HVPQ/PIQ/Q88 verifier. Observations are priority signals only; validation rules create findings only when evidence supports them.', 'observation_library': [{'question_no': '10.1.4', 'repeat_count': 35, 'priority': 'HIGH_REPEAT', 'category': 'Mooring / lifting / SPM / ETA', 'topic': 'Mooring brake test date / brake holding capacity', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following HVPQ Entries were not updated – The date of last winch test (10.1.4): 05-Aug-2025. There were no restrictions noted for the hose handling crane’s capability to maintain it’s design SWL when plumbing a point one metre outboard from the ship’s side over the full length of the manifold (10.9.1). the vessel was capable of carrying out operations at SBM /CBM but the arrangement at the vapor manifold area did not have arrangements for securing floating hoses. (10.9.4)', "HVPQ entries in the following sections were incorrect. 1. HVPQ 1.1.8 - Type of vessel in IOPP form B 1.11.4 stated as an oil tanker, however form B states as crude oil product tanker. 2. HVPQ 2.1.4 - Stated as an oil tanker, it should be crude oil product tanker. 3. HVPQ 10.1.4 - Vessel's mooring winch drums were split drum type. The entry in the HVPQ was incorrect and stated as 'No' to the split drums.", 'The following typo errors were seen in the latest HVPQ uploaded in the OCIMF repository, dated 12 May 2025: 1) HVPQ 7.1.1: cargo tanks inspection interval quarterly instead annually. 2) HVPQ 7.1.3: last 5 side starboard ballast tank inspection dated 12 July 2024 instead 12 March 2025. 3) HVPQ 9.16.18: missing the entry bleed and block valves. 4) HVPQ 10.1.3: missing entry SDMBL 36.5 tonnes. 5) HVPQ 10.1.4: the details of only one mooring winch was recorded. The vessel was fitted with 8 mooring w']}, {'question_no': '7.1.3', 'repeat_count': 34, 'priority': 'HIGH_REPEAT', 'category': 'Structural assessment / tank coating', 'topic': 'Tank coating / structural inspection dates and frequency', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following information was either not completed or was not accurately updated in the HVPQ dated 14-Sep-2025: (i) Date of last in water survey (1. 5. 5. 1 ): 16-Jan-2025; (ii) Assigned dead weight 4 1.e. 34,999 MT that was available was not included (1.8.6); (iii) The following ballast tank inspection dates were not updated (7.1.3): 1P on 08-Jul-2025, 2P on 08-Jul-2025, 3P on 08-Jul-2025, 5S on 7 7-Aug-2025; (iv) Accommodation ladder wire renewal date (10.10.2). 05-Jun-2025; (v) The renewal da', 'As per the HVPQ dated 02 Sep 2024, the following information was inaccurate:\n1. Section 1.5.11 Date of last annual survey was 12 Aug 2024, whereas this was the date of the last intermediate survey.\n2. Section 2.1.5 Certificate dates. a. The date of last annual for the statutory certificates was 12 Aug 2024, whereas this was the\ndate of last intermediate survey.\nb. The dates of last endorsement of the statutory and applicable certificates were blank.\n3. Section 7.1.1 Cargo tank coating. The last ', 'The following typo errors were seen in the latest HVPQ uploaded in the OCIMF repository, dated 12 May 2025: 1) HVPQ 7.1.1: cargo tanks inspection interval quarterly instead annually. 2) HVPQ 7.1.3: last 5 side starboard ballast tank inspection dated 12 July 2024 instead 12 March 2025. 3) HVPQ 9.16.18: missing the entry bleed and block valves. 4) HVPQ 10.1.3: missing entry SDMBL 36.5 tonnes. 5) HVPQ 10.1.4: the details of only one mooring winch was recorded. The vessel was fitted with 8 mooring w']}, {'question_no': '10.9.1', 'repeat_count': 29, 'priority': 'HIGH_REPEAT', 'category': 'Mooring / lifting / SPM / ETA', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['Two items were incorrectly declared in the latest HVPQ6 dated 3 October 2024 asf: 1-Item 1.4.3- Date of building contract 20-January-2019; Actually on 18-May-2018: 2-item 10.9.1 – last 5yr test for one of two cranes : 20-January-2011: Actually on 20-January-2022. Reportedly online updated immediately.', 'The following HVPQ Entries were not updated – The date of last winch test (10.1.4): 05-Aug-2025. There were no restrictions noted for the hose handling crane’s capability to maintain it’s design SWL when plumbing a point one metre outboard from the ship’s side over the full length of the manifold (10.9.1). the vessel was capable of carrying out operations at SBM /CBM but the arrangement at the vapor manifold area did not have arrangements for securing floating hoses. (10.9.4)', 'On the uploaded HVPQ the following items were recorded with inaccurate information: 1.9.(5,6) (LTI on 04-Jun-2025), 1.9.8 (Last PSC inspection was at Nha Be, Vietnam on 24-Jun-2025), 10.9.1 (Last 5 yearly test was on 04-Jun-2025).']}, {'question_no': '7.1.1', 'repeat_count': 22, 'priority': 'HIGH_REPEAT', 'category': 'Structural assessment / tank coating', 'topic': 'Tank coating / structural inspection dates and frequency', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['As per the HVPQ dated 02 Sep 2024, the following information was inaccurate:\n1. Section 1.5.11 Date of last annual survey was 12 Aug 2024, whereas this was the date of the last intermediate survey.\n2. Section 2.1.5 Certificate dates. a. The date of last annual for the statutory certificates was 12 Aug 2024, whereas this was the\ndate of last intermediate survey.\nb. The dates of last endorsement of the statutory and applicable certificates were blank.\n3. Section 7.1.1 Cargo tank coating. The last ', 'The following typo errors were seen in the latest HVPQ uploaded in the OCIMF repository, dated 12 May 2025: 1) HVPQ 7.1.1: cargo tanks inspection interval quarterly instead annually. 2) HVPQ 7.1.3: last 5 side starboard ballast tank inspection dated 12 July 2024 instead 12 March 2025. 3) HVPQ 9.16.18: missing the entry bleed and block valves. 4) HVPQ 10.1.3: missing entry SDMBL 36.5 tonnes. 5) HVPQ 10.1.4: the details of only one mooring winch was recorded. The vessel was fitted with 8 mooring w', 'There were minor errors having no impact on the inspection. 7.1.1 Tank Type 2G is a gas tanker The tank construction is SS cladding, which is not an option in the HVPQ drop-down menu. 9.1.1 Tank plan cross-section schematic is wrong. 9.5.3 Says that the pump cannot be by pass pump during loading; this is incorrect. 12.1.13 The vessel has closed chocks and bollards at the manifold, the distances are missing.']}, {'question_no': '10.1.3', 'repeat_count': 20, 'priority': 'HIGH_REPEAT', 'category': 'Mooring / lifting / SPM / ETA', 'topic': 'Venting / P-V valves / IGS', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ published to the OCIMF website on 24 Sept 2024 was randomly reviewed on 25 Sept 2024. Some questions were not responded and some questions were wrongly responded. For eg - 1.1.1, 1.2.4, 1.5.4, 1.9.8, 2.2.1, 5.3.2, 6.1.12, 7.1.6, 9.59.5, 9.68.6, 10.1.3, 10.4.2, 10.10.6, 11.10.1, 11.10.2, 12.4.1', 'HVPQ submitted through the CVIQ was found to be not completed, wrongly filled, not answered for the following sections as below: 1.1. 13, 1 .2.1, 1.2.4, 1.3.1, 1.3.2, 2. 1.4, 3.3.1, 4. 1. 1, 7. 1 .1, 7. 1.3, 9. 1 .1, 9.8.22, 9. 16.6, 9.17.1, 9.34.7, 10.1.3, 10. 1.4, 10.2.1, 10.4.2, 10.7.1, 10.8.1, 12. 1.14. For example: 1.1. 13 P and I Club# 3 Amount of P&I Cover; 9.16.6 What is the capacity of the IGS?-not responded; 1.3. 1 Registered Owner# 9 Number of years this ship has been owned by Registe', 'The following discrepancies were noted in the HVPQ: - Item 1.5.4.2: Lack of necessary information. - Item 1.5.5.1: Error in the due date recording, listed incorrectly as 03 April 2024 instead of 03 April 2022. - Item 1.5.11: Lack of necessary information. - Item 1.5.12: Lack of necessary information. - Item 1.5.16: An error discovered in the recording of the answer, which was incorrectly marked as NO instead of YES. - Item 1.5.17: Lack of necessary information. - Item 3.1.3: The number of rating']}, {'question_no': '10.7.1', 'repeat_count': 20, 'priority': 'HIGH_REPEAT', 'category': 'Mooring / lifting / SPM / ETA', 'topic': 'Venting / P-V valves / IGS', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['HVPQ submitted through the CVIQ was found to be not completed, wrongly filled, not answered for the following sections as below: 1.1. 13, 1 .2.1, 1.2.4, 1.3.1, 1.3.2, 2. 1.4, 3.3.1, 4. 1. 1, 7. 1 .1, 7. 1.3, 9. 1 .1, 9.8.22, 9. 16.6, 9.17.1, 9.34.7, 10.1.3, 10. 1.4, 10.2.1, 10.4.2, 10.7.1, 10.8.1, 12. 1.14. For example: 1.1. 13 P and I Club# 3 Amount of P&I Cover; 9.16.6 What is the capacity of the IGS?-not responded; 1.3. 1 Registered Owner# 9 Number of years this ship has been owned by Registe', 'The following discrepancies were noted in the HVPQ: - Item 1.5.4.2: Lack of necessary information. - Item 1.5.5.1: Error in the due date recording, listed incorrectly as 03 April 2024 instead of 03 April 2022. - Item 1.5.11: Lack of necessary information. - Item 1.5.12: Lack of necessary information. - Item 1.5.16: An error discovered in the recording of the answer, which was incorrectly marked as NO instead of YES. - Item 1.5.17: Lack of necessary information. - Item 3.1.3: The number of rating', 'The following HVPQ items were not corrected from the previous observations provided in the PIQ:- 1.5.11 / 1.5.12 / 9.1.1 /10.1.3.2.1 / 10.2.1 & 10.7.1. They were diagrams, and OP reported that operator was not able to upload']}, {'question_no': '10.8.1', 'repeat_count': 15, 'priority': 'MEDIUM_REPEAT', 'category': 'Mooring / lifting / SPM / ETA', 'topic': 'Venting / P-V valves / IGS', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['HVPQ submitted through the CVIQ was found to be not completed, wrongly filled, not answered for the following sections as below: 1.1. 13, 1 .2.1, 1.2.4, 1.3.1, 1.3.2, 2. 1.4, 3.3.1, 4. 1. 1, 7. 1 .1, 7. 1.3, 9. 1 .1, 9.8.22, 9. 16.6, 9.17.1, 9.34.7, 10.1.3, 10. 1.4, 10.2.1, 10.4.2, 10.7.1, 10.8.1, 12. 1.14. For example: 1.1. 13 P and I Club# 3 Amount of P&I Cover; 9.16.6 What is the capacity of the IGS?-not responded; 1.3. 1 Registered Owner# 9 Number of years this ship has been owned by Registe', 'The following items in the HVPQ were not correctly completed: 10.7.1; 10.8.1; 12.1.3 and 12.1.4.', 'Some erroneous or missing information was noted within the HVPQ under the following sections: 10.1.3.2, 10.7.1, 10.8.1.']}, {'question_no': '1.5.5', 'repeat_count': 15, 'priority': 'MEDIUM_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Certificates / class survey dates / endorsements', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following discrepancies were noted in the HVPQ: - Item 1.5.4.2: Lack of necessary information. - Item 1.5.5.1: Error in the due date recording, listed incorrectly as 03 April 2024 instead of 03 April 2022. - Item 1.5.11: Lack of necessary information. - Item 1.5.12: Lack of necessary information. - Item 1.5.16: An error discovered in the recording of the answer, which was incorrectly marked as NO instead of YES. - Item 1.5.17: Lack of necessary information. - Item 3.1.3: The number of rating', "The following incorrect information was listed in the ship's HVPQ (dated 28 April 2025).\nItem 1.5.5.1\nLast IWS date 16 Nov 2022 was not correct since last SSH was also undertaken on this date, as per Item 1.5.6.1.", '(i) The HVPQ entries inside section Ref: 1.5.5.2/ 9.16.18/ 9.32.2(Dehumidifier not fitted)/ 11.3.1(KW incorrect) /11.7.1/ 11.7.4/ 12.1.4 did not accurately reflect the information relating to the ship at the time of inspection. The incorrect info & particulars were brought to the attention of Master.\n(ii) The vessel designation as recorded as per IOPP certificate was not declared inside the PIQ.']}, {'question_no': '2.2.1', 'repeat_count': 13, 'priority': 'MEDIUM_REPEAT', 'category': 'Environmental / certificates', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ published to the OCIMF website on 24 Sept 2024 was randomly reviewed on 25 Sept 2024. Some questions were not responded and some questions were wrongly responded. For eg - 1.1.1, 1.2.4, 1.5.4, 1.9.8, 2.2.1, 5.3.2, 6.1.12, 7.1.6, 9.59.5, 9.68.6, 10.1.3, 10.4.2, 10.10.6, 11.10.1, 11.10.2, 12.4.1', 'Not all the publications required under HVPQ 2.2.1 were observed listed on the last Operator’s HVPQ dated 29th July 2025.', 'The following items in HVPQ completed on 9 Dec.2024 were inaccurate. 2.1.5 DOC last annual, CLC wreck removal and bunker 2.2.1 Publication Edition Number - ICS Bridge Procedure Guide 3.1.1 The minimum number of certificate officers to be carried as record in the MSMD 3.1.9 The minimum number of ratings to be carried as specified in the MSMD 9.16.5 Fixed O2 alarm fitted in inert gas generating or storage spaces 10.1.7 Details for Mooring Ropes - Renewal date 10.9.1 Crane last annual test']}, {'question_no': '10.2.1', 'repeat_count': 13, 'priority': 'MEDIUM_REPEAT', 'category': 'Mooring / lifting / SPM / ETA', 'topic': 'Venting / P-V valves / IGS', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['HVPQ submitted through the CVIQ was found to be not completed, wrongly filled, not answered for the following sections as below: 1.1. 13, 1 .2.1, 1.2.4, 1.3.1, 1.3.2, 2. 1.4, 3.3.1, 4. 1. 1, 7. 1 .1, 7. 1.3, 9. 1 .1, 9.8.22, 9. 16.6, 9.17.1, 9.34.7, 10.1.3, 10. 1.4, 10.2.1, 10.4.2, 10.7.1, 10.8.1, 12. 1.14. For example: 1.1. 13 P and I Club# 3 Amount of P&I Cover; 9.16.6 What is the capacity of the IGS?-not responded; 1.3. 1 Registered Owner# 9 Number of years this ship has been owned by Registe', 'The following discrepancies were noted in the HVPQ: - Item 1.5.4.2: Lack of necessary information. - Item 1.5.5.1: Error in the due date recording, listed incorrectly as 03 April 2024 instead of 03 April 2022. - Item 1.5.11: Lack of necessary information. - Item 1.5.12: Lack of necessary information. - Item 1.5.16: An error discovered in the recording of the answer, which was incorrectly marked as NO instead of YES. - Item 1.5.17: Lack of necessary information. - Item 3.1.3: The number of rating', 'The following HVPQ items were not corrected from the previous observations provided in the PIQ:- 1.5.11 / 1.5.12 / 9.1.1 /10.1.3.2.1 / 10.2.1 & 10.7.1. They were diagrams, and OP reported that operator was not able to upload']}, {'question_no': '1.5.6', 'repeat_count': 12, 'priority': 'MEDIUM_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Certificates / class survey dates / endorsements', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following information related to the uploaded HVPQ were missing or inaccurate; 1.5.6.5 The date of the next special survey was stated as 07.09.2026 instead of 07.09.2030. 1.9.8 Port state control dates were missing. 4.2.10 Can the radio transmit the helicopter homing signal on 410 KHz states as Yes instead of No.', "The following incorrect information was listed in the ship's HVPQ (dated 28 April 2025).\nItem 1.5.5.1\nLast IWS date 16 Nov 2022 was not correct since last SSH was also undertaken on this date, as per Item 1.5.6.1.", 'A few errors were noted in the HVPQ. The correct data was:\n1.5.4 Date of last drydock was 12 June 2024.\n1.5.6 Date of last special survey was 12 June 2024.\n9.16.19 The nitrogen system had one segregation.']}, {'question_no': '1.3.1', 'repeat_count': 11, 'priority': 'MEDIUM_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Certificates / class survey dates / endorsements', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['HVPQ submitted through the CVIQ was found to be not completed, wrongly filled, not answered for the following sections as below: 1.1. 13, 1 .2.1, 1.2.4, 1.3.1, 1.3.2, 2. 1.4, 3.3.1, 4. 1. 1, 7. 1 .1, 7. 1.3, 9. 1 .1, 9.8.22, 9. 16.6, 9.17.1, 9.34.7, 10.1.3, 10. 1.4, 10.2.1, 10.4.2, 10.7.1, 10.8.1, 12. 1.14. For example: 1.1. 13 P and I Club# 3 Amount of P&I Cover; 9.16.6 What is the capacity of the IGS?-not responded; 1.3. 1 Registered Owner# 9 Number of years this ship has been owned by Registe', 'In the last HVPQ dated on 03 January 2025 was observed missing information: Ownership and Operation 1.3.1.5 IMO Number register owner. In the item 3 firefighting and lifesaving equipment it was stated last date test foam analysis on 24 November 2023. However, on board was noted valid certificate dated on 09 September 2024.', 'The following discrepancies were noted in the HVPQ document: - Item 1.3.1.6: Lack of the necessary information. - Item 1.5.10: Error in the recording of the date, listed incorrectly as 28 September 2022 instead of 19 September 2022.. - Item 1.5.12: Lack of the necessary information. - Item 3.1.1: Error in the recording of the quantity, listed incorrectly as 07 instead of 06. - Item 3.1.9: Error in the recording of the quantity, listed incorrectly as 08 instead of 07. - Item 9.10.4: Error in the ']}, {'question_no': '1.5.12', 'repeat_count': 11, 'priority': 'MEDIUM_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Certificates / class survey dates / endorsements', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following discrepancies were noted in the HVPQ: - Item 1.5.4.2: Lack of necessary information. - Item 1.5.5.1: Error in the due date recording, listed incorrectly as 03 April 2024 instead of 03 April 2022. - Item 1.5.11: Lack of necessary information. - Item 1.5.12: Lack of necessary information. - Item 1.5.16: An error discovered in the recording of the answer, which was incorrectly marked as NO instead of YES. - Item 1.5.17: Lack of necessary information. - Item 3.1.3: The number of rating', 'The following HVPQ items were not corrected from the previous observations provided in the PIQ:- 1.5.11 / 1.5.12 / 9.1.1 /10.1.3.2.1 / 10.2.1 & 10.7.1. They were diagrams, and OP reported that operator was not able to upload', 'The following discrepancies were noted in the HVPQ document: - Item 1.3.1.6: Lack of the necessary information. - Item 1.5.10: Error in the recording of the date, listed incorrectly as 28 September 2022 instead of 19 September 2022.. - Item 1.5.12: Lack of the necessary information. - Item 3.1.1: Error in the recording of the quantity, listed incorrectly as 07 instead of 06. - Item 3.1.9: Error in the recording of the quantity, listed incorrectly as 08 instead of 07. - Item 9.10.4: Error in the ']}, {'question_no': '5.3.1', 'repeat_count': 11, 'priority': 'MEDIUM_REPEAT', 'category': 'Safety / firefighting / lifeboats', 'topic': 'Mooring line / wire / tail installation age', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['HVPQ had some minor errors as follows: (1.1.7/3) Ship\'s email entered was ineffective as senders\' emails may be rejected & not received by vessel. (Only by adding "master." was ship receiving all emails. For information, using email given in HVPQ, Inspector was not able to contact ship for pre-boarding formalities); (3.2.1) Incorrectly entered YES when senior officers in actuality, did not return to same ship on rotational basis; (5.3.1.4) Foam supplied or tested date was entered as 11 November ', '1.3.1.10 Date when registered with the current owner was incorrectly recorded. 1.3.2.10 Date current operator assumed technical control of the vessel was not documented. 1.5.5 IWS details were incorrectly recorded. 1.5.11 Date of last annual survey was not documented. 5.3.1.4 Date of last foam test analysis certificate was incorrectly recorded.', 'The following discrepancies were found in the HVPQ document:\nItem 1.5.3.1: An error was discovered in the recording of the answer, which was incorrectly marked as NO instead of YES.\nItem 1.5.3.2: Lack of the necessary information.\nItem 1.5.3.3: Lack of the necessary information.\nItem 1.5.5.1: Lack of the necessary information.\nItem 1.5.5.2: Lack of the necessary information.\nItem 1.5.10: Error in recording the date, listed incorrectly 26 March 2024 instead 12 March 2024.\nItem 1.5.12: Lack of the']}, {'question_no': '1.9.8', 'repeat_count': 10, 'priority': 'MEDIUM_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'PSC inspection / detention / deficiency declaration', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ published to the OCIMF website on 24 Sept 2024 was randomly reviewed on 25 Sept 2024. Some questions were not responded and some questions were wrongly responded. For eg - 1.1.1, 1.2.4, 1.5.4, 1.9.8, 2.2.1, 5.3.2, 6.1.12, 7.1.6, 9.59.5, 9.68.6, 10.1.3, 10.4.2, 10.10.6, 11.10.1, 11.10.2, 12.4.1', 'The following information related to the uploaded HVPQ were missing or inaccurate; 1.5.6.5 The date of the next special survey was stated as 07.09.2026 instead of 07.09.2030. 1.9.8 Port state control dates were missing. 4.2.10 Can the radio transmit the helicopter homing signal on 410 KHz states as Yes instead of No.', 'On the uploaded HVPQ the following items were recorded with inaccurate information: 1.9.(5,6) (LTI on 04-Jun-2025), 1.9.8 (Last PSC inspection was at Nha Be, Vietnam on 24-Jun-2025), 10.9.1 (Last 5 yearly test was on 04-Jun-2025).']}, {'question_no': '3.1.1', 'repeat_count': 10, 'priority': 'MEDIUM_REPEAT', 'category': 'Crew / training / operator assessments', 'topic': 'Manning / crew declarations', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['Question 3.1.1 in the HVPQ wrongly stated the minimum number of certified officers to be carried as 6 instead of 7 as per the Minimum Safe Manning Document as well as the information filled in the PIQ. Question 9.16.2 in the HVPQ was answered as a No for the vessel not being fitted with a P/V Breaker. The vessel was however with a P/V breaker in the form of two higher capacity P/V valves installed on the main N2 line.', 'The HVPQ last updated on 05 Jun 2025 and uploaded to the document repository on 06 Jun 2025 was reviewed and the following discrepancies were noted, viz., 1.5.2 date next In Water Survey due, 3.1.1 minimum number of certified officers to be carried as recorded in Minimum Safe Manning document was incorrect, 3.1.9 minimum number of ratings to be carried as specified in the Minimum Safe Manning document was incorrect.', 'The following discrepancies were noted in the HVPQ document: - Item 1.3.1.6: Lack of the necessary information. - Item 1.5.10: Error in the recording of the date, listed incorrectly as 28 September 2022 instead of 19 September 2022.. - Item 1.5.12: Lack of the necessary information. - Item 3.1.1: Error in the recording of the quantity, listed incorrectly as 07 instead of 06. - Item 3.1.9: Error in the recording of the quantity, listed incorrectly as 08 instead of 07. - Item 9.10.4: Error in the ']}, {'question_no': '10.1.7', 'repeat_count': 10, 'priority': 'MEDIUM_REPEAT', 'category': 'Mooring / lifting / SPM / ETA', 'topic': 'Tank coating / structural inspection dates and frequency', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following information was either not completed or was not accurately updated in the HVPQ dated 14-Sep-2025: (i) Date of last in water survey (1. 5. 5. 1 ): 16-Jan-2025; (ii) Assigned dead weight 4 1.e. 34,999 MT that was available was not included (1.8.6); (iii) The following ballast tank inspection dates were not updated (7.1.3): 1P on 08-Jul-2025, 2P on 08-Jul-2025, 3P on 08-Jul-2025, 5S on 7 7-Aug-2025; (iv) Accommodation ladder wire renewal date (10.10.2). 05-Jun-2025; (v) The renewal da', 'The following typo errors were seen in the latest HVPQ uploaded in the OCIMF repository, dated 12 May 2025: 1) HVPQ 7.1.1: cargo tanks inspection interval quarterly instead annually. 2) HVPQ 7.1.3: last 5 side starboard ballast tank inspection dated 12 July 2024 instead 12 March 2025. 3) HVPQ 9.16.18: missing the entry bleed and block valves. 4) HVPQ 10.1.3: missing entry SDMBL 36.5 tonnes. 5) HVPQ 10.1.4: the details of only one mooring winch was recorded. The vessel was fitted with 8 mooring w', 'The following entries were either incorrect or incomplete :- 1.2.4.3, 1.3.1.7 & 8, 1.4.2, 1.5.4.1, 1.5.6, 3.1.6.2, 3.2.1, 3.2.2, 6.1.8, 7.1.1, 9.15.1, 9.15.3.5, 9.16.10, 9.16.29.2, 10.1.7 and 12.1.9.']}, {'question_no': '2.1.5', 'repeat_count': 10, 'priority': 'MEDIUM_REPEAT', 'category': 'Environmental / certificates', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['As per the HVPQ dated 02 Sep 2024, the following information was inaccurate:\n1. Section 1.5.11 Date of last annual survey was 12 Aug 2024, whereas this was the date of the last intermediate survey.\n2. Section 2.1.5 Certificate dates. a. The date of last annual for the statutory certificates was 12 Aug 2024, whereas this was the\ndate of last intermediate survey.\nb. The dates of last endorsement of the statutory and applicable certificates were blank.\n3. Section 7.1.1 Cargo tank coating. The last ', 'HVPQ provided with CVIQ had incorrect information under 2.1.5 (Expiry date of safety radio certificate was wrongly entered as 14\nNovember 2028 instead of 15 December 2024.)', 'The following items in HVPQ completed on 9 Dec.2024 were inaccurate. 2.1.5 DOC last annual, CLC wreck removal and bunker 2.2.1 Publication Edition Number - ICS Bridge Procedure Guide 3.1.1 The minimum number of certificate officers to be carried as record in the MSMD 3.1.9 The minimum number of ratings to be carried as specified in the MSMD 9.16.5 Fixed O2 alarm fitted in inert gas generating or storage spaces 10.1.7 Details for Mooring Ropes - Renewal date 10.9.1 Crane last annual test']}, {'question_no': '1.5.4', 'repeat_count': 9, 'priority': 'MEDIUM_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Certificates / class survey dates / endorsements', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ published to the OCIMF website on 24 Sept 2024 was randomly reviewed on 25 Sept 2024. Some questions were not responded and some questions were wrongly responded. For eg - 1.1.1, 1.2.4, 1.5.4, 1.9.8, 2.2.1, 5.3.2, 6.1.12, 7.1.6, 9.59.5, 9.68.6, 10.1.3, 10.4.2, 10.10.6, 11.10.1, 11.10.2, 12.4.1', 'The following discrepancies were noted in the HVPQ: - Item 1.5.4.2: Lack of necessary information. - Item 1.5.5.1: Error in the due date recording, listed incorrectly as 03 April 2024 instead of 03 April 2022. - Item 1.5.11: Lack of necessary information. - Item 1.5.12: Lack of necessary information. - Item 1.5.16: An error discovered in the recording of the answer, which was incorrectly marked as NO instead of YES. - Item 1.5.17: Lack of necessary information. - Item 3.1.3: The number of rating', 'A few errors were noted in the HVPQ. The correct data was:\n1.5.4 Date of last drydock was 12 June 2024.\n1.5.6 Date of last special survey was 12 June 2024.\n9.16.19 The nitrogen system had one segregation.']}, {'question_no': '9.1.1', 'repeat_count': 9, 'priority': 'MEDIUM_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'Mooring line / wire / tail installation age', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following discrepancies were noted in the HVPQ: - Item 1.5.4.2: Lack of necessary information. - Item 1.5.5.1: Error in the due date recording, listed incorrectly as 03 April 2024 instead of 03 April 2022. - Item 1.5.11: Lack of necessary information. - Item 1.5.12: Lack of necessary information. - Item 1.5.16: An error discovered in the recording of the answer, which was incorrectly marked as NO instead of YES. - Item 1.5.17: Lack of necessary information. - Item 3.1.3: The number of rating', 'The following HVPQ items were not corrected from the previous observations provided in the PIQ:- 1.5.11 / 1.5.12 / 9.1.1 /10.1.3.2.1 / 10.2.1 & 10.7.1. They were diagrams, and OP reported that operator was not able to upload', "The HVPQ data entered either in error or not populated for the following HVPQ numbers were noted:\n9.1.1 tank plan incomplete\n9.2.1 1 & 2 not populated\n9.15.1 heat exchanger internal, heating coils were fitted.\n10.1.6.4 All mooring ropes had indicator strands\n10.9.4 not populated\n11.1.9 not populated\n11.3.1 A/E fuel HFO only. (The HVPQ only allows one selection) the A/E's can use either HFO or MDO.\nRECTIFIED, where possible, during the inspection."]}, {'question_no': '9.16.5', 'repeat_count': 9, 'priority': 'MEDIUM_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['HVPQ item 6.1.8&10, regarding cargo chest and item 9.16.5, regarding IGS space/room oxygen sensors installation were marked wrongly.', 'The following items in HVPQ completed on 9 Dec.2024 were inaccurate. 2.1.5 DOC last annual, CLC wreck removal and bunker 2.2.1 Publication Edition Number - ICS Bridge Procedure Guide 3.1.1 The minimum number of certificate officers to be carried as record in the MSMD 3.1.9 The minimum number of ratings to be carried as specified in the MSMD 9.16.5 Fixed O2 alarm fitted in inert gas generating or storage spaces 10.1.7 Details for Mooring Ropes - Renewal date 10.9.1 Crane last annual test', 'Below sections of the HVPQ were not correctly updated :\n1.3.2.5 : IMO number of the Technical operator was 1677771.\n1.3.2.10 : Date current operator assumed responsibility was 08 Feb 2025 as per Safety Management Certificate.\n1.9.8 : Last PSC inspection was performed at New Orleans on 27 Jan 2025.\n2.1.5 : All statutory certificates except for Safety Equipment Certificate were issued on 08 Feb 2025 and were valid until 06 Nov 2028. Safety Equipment Certificate was issued on 08 Feb 2025 and was va']}, {'question_no': '1.1.13', 'repeat_count': 9, 'priority': 'MEDIUM_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['In HVPQ under question 1.1.13 for information on the vessel\'s P&I club the option of "Other (Specify)" had been selected and then\nthe name of the ships P&I had been typed in, instead of selecting the P&I club from the drop-down list.', 'The HVPQ showed erroneous dates and data as follows:\n1.1.13.4: P&I cover include wreck removal no response\n1.5.5.2: incorrect date\n1.5.6.5: next special date no response\n1.9.8.1: PSC date incorrect\n3.1.6.1: should be yes\n3.1.6.2: should be no\n3.1.10: same as for officers no response needed\n4.1.1: type of navigation equipments not completed\n10.9.1: crane annual test date incorrect\n11.1.13.1: quick closing valves fitted\n11.1.13.2: no response\n11.4.3: no response\n11.10.1: no response\n12.1.1: should', 'The name of P and I club on VPQ 1.1.13 was not provided however the entry certificate for P and I club was available onboard at the time of inspection.']}, {'question_no': '4.1.1', 'repeat_count': 9, 'priority': 'MEDIUM_REPEAT', 'category': 'Navigation and communication', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['HVPQ submitted through the CVIQ was found to be not completed/ wrongly filled/ not answered for the following sections as below: 1.5.1,1.5.5,1.8.7, 4.1.1, 5.3.2, 7.1.6, 9.59.7, 10.2.1, 10.4.1, 10.7.1, 10.8.1, 10.10.7, 11.1.9, 11.2.2, 11.9.3, 12.4.1. For example, 1.5.1 Classification society #2 is Classification society is an IACS member, # does the ship have dual class ? 9.59.7 Compressors # Are they oil free ? There were no responses provided. 4.1.1 Navigational equipment fitted on board? # Glo', 'The HVPQ showed erroneous dates and data as follows:\n1.1.13.4: P&I cover include wreck removal no response\n1.5.5.2: incorrect date\n1.5.6.5: next special date no response\n1.9.8.1: PSC date incorrect\n3.1.6.1: should be yes\n3.1.6.2: should be no\n3.1.10: same as for officers no response needed\n4.1.1: type of navigation equipments not completed\n10.9.1: crane annual test date incorrect\n11.1.13.1: quick closing valves fitted\n11.1.13.2: no response\n11.4.3: no response\n11.10.1: no response\n12.1.1: should', 'The HVPQ submitted through the CVIQ was found to be not completed/wrongly filled/ not answered for the following sections as\nbelow: 1.5.5, 1.5.19, 1.8.1, 1.9.8, 2.2.1, 3.1.9, 4.1.1, 4.2.2, 10.1.3, 10.4.2, 10.6.3, 10.6.6, 10.6.7, 10.7.1, 10.8.1, 10.8.2-10.8.4, 11.1.8,\n11.1.9, 11.1.13, 11.2.2, 11.3.1, 11.3.4, 11.9.1, 12.1.7, 12.1.8, 12.1.14, 12.2.4, 12.4.1. For example: 1.9.8 Port State Control # 1 Date of\nlast Port State Control inspection and # 2 Port of last Port State Control inspection - were']}, {'question_no': '11.3.1', 'repeat_count': 8, 'priority': 'MEDIUM_REPEAT', 'category': 'Engine room / machinery / generators', 'topic': 'Mooring line / wire / tail installation age', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ dated 21 May 2025 had the following errors: 5.3.2.8 - N/A Not Fitted. 5.3.8.2 - N/A Not Fitted. 5.3.8.3 - The Port Lifeboat was the dedicated rescue boat. 9.16.18 - Block & Bleed Arrangement. 10.6.8 - No - the SPM bracket (SWL 204MT) was fitted on the starboard side of the forward main deck with an approx 60 degree off set from the center lead and then via one pedestal roller with a wrap angle of approx 60 degrees to the winch pick up drum. 11.3.1 The vessel was equipped with 3 power ge', "The HVPQ data entered either in error or not populated for the following HVPQ numbers were noted:\n9.1.1 tank plan incomplete\n9.2.1 1 & 2 not populated\n9.15.1 heat exchanger internal, heating coils were fitted.\n10.1.6.4 All mooring ropes had indicator strands\n10.9.4 not populated\n11.1.9 not populated\n11.3.1 A/E fuel HFO only. (The HVPQ only allows one selection) the A/E's can use either HFO or MDO.\nRECTIFIED, where possible, during the inspection.", '(i) The HVPQ entries inside section Ref: 1.5.5.2/ 9.16.18/ 9.32.2(Dehumidifier not fitted)/ 11.3.1(KW incorrect) /11.7.1/ 11.7.4/ 12.1.4 did not accurately reflect the information relating to the ship at the time of inspection. The incorrect info & particulars were brought to the attention of Master.\n(ii) The vessel designation as recorded as per IOPP certificate was not declared inside the PIQ.']}, {'question_no': '1.5.11', 'repeat_count': 8, 'priority': 'MEDIUM_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Tank coating / structural inspection dates and frequency', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following discrepancies were noted in the HVPQ: - Item 1.5.4.2: Lack of necessary information. - Item 1.5.5.1: Error in the due date recording, listed incorrectly as 03 April 2024 instead of 03 April 2022. - Item 1.5.11: Lack of necessary information. - Item 1.5.12: Lack of necessary information. - Item 1.5.16: An error discovered in the recording of the answer, which was incorrectly marked as NO instead of YES. - Item 1.5.17: Lack of necessary information. - Item 3.1.3: The number of rating', 'The following HVPQ items were not corrected from the previous observations provided in the PIQ:- 1.5.11 / 1.5.12 / 9.1.1 /10.1.3.2.1 / 10.2.1 & 10.7.1. They were diagrams, and OP reported that operator was not able to upload', 'As per the HVPQ dated 02 Sep 2024, the following information was inaccurate:\n1. Section 1.5.11 Date of last annual survey was 12 Aug 2024, whereas this was the date of the last intermediate survey.\n2. Section 2.1.5 Certificate dates. a. The date of last annual for the statutory certificates was 12 Aug 2024, whereas this was the\ndate of last intermediate survey.\nb. The dates of last endorsement of the statutory and applicable certificates were blank.\n3. Section 7.1.1 Cargo tank coating. The last ']}, {'question_no': '3.1.9', 'repeat_count': 8, 'priority': 'MEDIUM_REPEAT', 'category': 'Crew / training / operator assessments', 'topic': 'Manning / crew declarations', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ last updated on 05 Jun 2025 and uploaded to the document repository on 06 Jun 2025 was reviewed and the following discrepancies were noted, viz., 1.5.2 date next In Water Survey due, 3.1.1 minimum number of certified officers to be carried as recorded in Minimum Safe Manning document was incorrect, 3.1.9 minimum number of ratings to be carried as specified in the Minimum Safe Manning document was incorrect.', 'The following discrepancies were noted in the HVPQ document: - Item 1.3.1.6: Lack of the necessary information. - Item 1.5.10: Error in the recording of the date, listed incorrectly as 28 September 2022 instead of 19 September 2022.. - Item 1.5.12: Lack of the necessary information. - Item 3.1.1: Error in the recording of the quantity, listed incorrectly as 07 instead of 06. - Item 3.1.9: Error in the recording of the quantity, listed incorrectly as 08 instead of 07. - Item 9.10.4: Error in the ', 'On the uploaded HVPQ, some entries were either missing or not accurately answered: - Items 6.1.8, 9.6.6, 11.8.3, 12.1.3 & 12.1.4 had blank responses. - The date entered in the item 1.1.4.4 was one day ahead of the actual date of change of flag. - Item 1.2.3 stated that a CII rating of A was obtained and verified by the Class. Actually, the vessel was delivered on 18 September 2025 and was not due to receive its initial CII rating. - The numbers of certified officers and ratings recorded in the i']}, {'question_no': '1.1.8', 'repeat_count': 8, 'priority': 'MEDIUM_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Certificates / class survey dates / endorsements', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ["HVPQ entries in the following sections were incorrect. 1. HVPQ 1.1.8 - Type of vessel in IOPP form B 1.11.4 stated as an oil tanker, however form B states as crude oil product tanker. 2. HVPQ 2.1.4 - Stated as an oil tanker, it should be crude oil product tanker. 3. HVPQ 10.1.4 - Vessel's mooring winch drums were split drum type. The entry in the HVPQ was incorrect and stated as 'No' to the split drums.", 'A review of the uploaded HVPQ dated 18 Feb 2025 indicated that the Q1.1.8 was not correctly updated as per the IOPPC which was stated that the vessel was an Oil Tanker. The HVPQ question Q12.1.4 and Q12.1.5 was answered in affirmative whereby generated a question in CVIQ for this inspection.', 'Review of the uploaded HVPQ dated 14 Apr 2025 indicated the following were not correctly updated. 1. HVPQ Q1.1.8 was not correctly updated to indicate the type of ship as an "Oil Tanker," as specified in the International Oil Pollution Prevention Certificate (IOPPC). 2. HVPQ Q10.4.1 was answered in negative and the question required to ignore the remainder of the section, but Item Q10.4.7 and Q10.5.3 was provided with information which generated a question in CVIQ for emergency towing equipment ']}, {'question_no': '1.5.19', 'repeat_count': 8, 'priority': 'MEDIUM_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Certificates / class survey dates / endorsements', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ submitted through the CVIQ was found to be not completed/wrongly filled/ not answered for the following sections as\nbelow: 1.5.5, 1.5.19, 1.8.1, 1.9.8, 2.2.1, 3.1.9, 4.1.1, 4.2.2, 10.1.3, 10.4.2, 10.6.3, 10.6.6, 10.6.7, 10.7.1, 10.8.1, 10.8.2-10.8.4, 11.1.8,\n11.1.9, 11.1.13, 11.2.2, 11.3.1, 11.3.4, 11.9.1, 12.1.7, 12.1.8, 12.1.14, 12.2.4, 12.4.1. For example: 1.9.8 Port State Control # 1 Date of\nlast Port State Control inspection and # 2 Port of last Port State Control inspection - were', 'HVPQ dated 16th October 2024 on 1.5.19, 1.9.5, 7.1.3 and 10.1.4 were not correct.', 'HVPQ - Items 1.5.19 and 7.1.1 were incorrect. PIQ - item 5.7.1028.1 was incorrect. (A crew member injury was recorded onboard on 25 September 2024).']}, {'question_no': '1.3.2', 'repeat_count': 7, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'PSC inspection / detention / deficiency declaration', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['HVPQ submitted through the CVIQ was found to be not completed, wrongly filled, not answered for the following sections as below: 1.1. 13, 1 .2.1, 1.2.4, 1.3.1, 1.3.2, 2. 1.4, 3.3.1, 4. 1. 1, 7. 1 .1, 7. 1.3, 9. 1 .1, 9.8.22, 9. 16.6, 9.17.1, 9.34.7, 10.1.3, 10. 1.4, 10.2.1, 10.4.2, 10.7.1, 10.8.1, 12. 1.14. For example: 1.1. 13 P and I Club# 3 Amount of P&I Cover; 9.16.6 What is the capacity of the IGS?-not responded; 1.3. 1 Registered Owner# 9 Number of years this ship has been owned by Registe', 'Below sections of the HVPQ were not correctly updated :\n1.3.2.5 : IMO number of the Technical operator was 1677771.\n1.3.2.10 : Date current operator assumed responsibility was 08 Feb 2025 as per Safety Management Certificate.\n1.9.8 : Last PSC inspection was performed at New Orleans on 27 Jan 2025.\n2.1.5 : All statutory certificates except for Safety Equipment Certificate were issued on 08 Feb 2025 and were valid until 06 Nov 2028. Safety Equipment Certificate was issued on 08 Feb 2025 and was va', 'The technical operator was not changed, but the question in the HVPQ (1.3.2.10) date current operator assumed control of the ship was on 05 August 2024.']}, {'question_no': '9.15.1', 'repeat_count': 7, 'priority': 'LOW_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'Mooring line / wire / tail installation age', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ["The HVPQ data entered either in error or not populated for the following HVPQ numbers were noted:\n9.1.1 tank plan incomplete\n9.2.1 1 & 2 not populated\n9.15.1 heat exchanger internal, heating coils were fitted.\n10.1.6.4 All mooring ropes had indicator strands\n10.9.4 not populated\n11.1.9 not populated\n11.3.1 A/E fuel HFO only. (The HVPQ only allows one selection) the A/E's can use either HFO or MDO.\nRECTIFIED, where possible, during the inspection.", 'The following entries were either incorrect or incomplete :- 1.2.4.3, 1.3.1.7 & 8, 1.4.2, 1.5.4.1, 1.5.6, 3.1.6.2, 3.2.1, 3.2.2, 6.1.8, 7.1.1, 9.15.1, 9.15.3.5, 9.16.10, 9.16.29.2, 10.1.7 and 12.1.9.', 'The HVPQ updated on 04 March 2025 was not accurately completed with respect to following items:\n1.5.9/1.5.10/2.2.1/5.3.2.10/8.2.3.3/9.6.6/9.8.14.3/9.15.1/9.30.6/10.4.2/10.9.1/11.1.6']}, {'question_no': '4.2.2', 'repeat_count': 7, 'priority': 'LOW_REPEAT', 'category': 'Navigation and communication', 'topic': 'Tank coating / structural inspection dates and frequency', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ submitted through the CVIQ was found to be not completed/wrongly filled/ not answered for the following sections as\nbelow: 1.5.5, 1.5.19, 1.8.1, 1.9.8, 2.2.1, 3.1.9, 4.1.1, 4.2.2, 10.1.3, 10.4.2, 10.6.3, 10.6.6, 10.6.7, 10.7.1, 10.8.1, 10.8.2-10.8.4, 11.1.8,\n11.1.9, 11.1.13, 11.2.2, 11.3.1, 11.3.4, 11.9.1, 12.1.7, 12.1.8, 12.1.14, 12.2.4, 12.4.1. For example: 1.9.8 Port State Control # 1 Date of\nlast Port State Control inspection and # 2 Port of last Port State Control inspection - were', 'The following discrepancies were noted 3.1.3 Number of ratings declared incorrectly as 24 instead of an board 13. SDO stated the same would be rectified 4.1.1 Software and firmware data was not populated, SDO stated that during the last annuals the technician had not left behind such data while the vessel was contracted for annual software upgrade 4.2.2 EPIRB data was not populated 7.1.1 Cargo tank coating inspection interval was incorrectly stated as 30 months instead of 60 months as per operat', 'On the uploaded HVPQ dated 16 June 2025, a few entries were either missing or not accurately answered: - In item 2.2.1, the column of edition number for all publications was marked as Yes or No rather than a particular version of the published book as applicable. - Item 4.2.2 was missing information about communications equipment software and firmware version as applicable. - In item 5.2.1, the diameter of the circle was recorded as 10.0 meters instead of 20.0 meters which was marked on the main']}, {'question_no': '1.5.18', 'repeat_count': 7, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Certificates / class survey dates / endorsements', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['Noted following HVPQ item were wrongly marked: Item 1.5.18 - Yes, but the vessel currently did not have any Class dispensation. Item 6.1.7&8 - Yes, but Cargo sea chest was not fitted onboard. Therefore invalid question 6.2.2 generated.', 'The VPQ were errors as following; (a) 1.5.14 CLASS CONDITION: “No”-> Yes (The following class condition was appended on the last class survey status report dated 12 November 2024; (1) VESSEL INFORMED THE NO. 2 STEERING GEAR SYSTEM OF MALFUNCTIONING DUE TO A BURNT MOTOR. SAME TO BE RENEWED AND SPECIALLY SATISFACTION OF ATTENDING SURVEYOR. IMPOSED DATE: 10 NOVEMBER 2024 / DUE DATE: 15 NOVEMBER 2024) (2) THE FOLLOWING DEFICIENCIES REMAIN TO BE DEALT WITH X-BAND RADAR MALFUNCTIONING. IMPOSED DATE: 0', 'Below sections of the HVPQ were not correctly updated :\n1.5.18 : Vessel was in possession of flag state dispensation valid until 16 May 2025 for the malfunction of remote ullage and\npressure sensors of Slop port cargo tank.\n9.16.5 : Fixed oxygen detection was not provided.']}, {'question_no': '1.2.4', 'repeat_count': 6, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Certificates / class survey dates / endorsements', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ published to the OCIMF website on 24 Sept 2024 was randomly reviewed on 25 Sept 2024. Some questions were not responded and some questions were wrongly responded. For eg - 1.1.1, 1.2.4, 1.5.4, 1.9.8, 2.2.1, 5.3.2, 6.1.12, 7.1.6, 9.59.5, 9.68.6, 10.1.3, 10.4.2, 10.10.6, 11.10.1, 11.10.2, 12.4.1', 'HVPQ submitted through the CVIQ was found to be not completed, wrongly filled, not answered for the following sections as below: 1.1. 13, 1 .2.1, 1.2.4, 1.3.1, 1.3.2, 2. 1.4, 3.3.1, 4. 1. 1, 7. 1 .1, 7. 1.3, 9. 1 .1, 9.8.22, 9. 16.6, 9.17.1, 9.34.7, 10.1.3, 10. 1.4, 10.2.1, 10.4.2, 10.7.1, 10.8.1, 12. 1.14. For example: 1.1. 13 P and I Club# 3 Amount of P&I Cover; 9.16.6 What is the capacity of the IGS?-not responded; 1.3. 1 Registered Owner# 9 Number of years this ship has been owned by Registe', 'information provided in the HVPQ by the vessel\'s operator completed on 30 Jun 2025 downloaded on 30 Jun 2025 from OCIMF\nwebsite was inacurate. Under HVPQ No. 1.2.3.2 CII rating was defined as "A" instead of "B" as per certificate number DCS-9637088-\n2024-242665. Under HVPQ No. 1.2.4.2 EIV rating was defined as "5.96" instead of "5.544" as per certificate number DCS-9637088-\n2024-242665.']}, {'question_no': '10.4.2', 'repeat_count': 6, 'priority': 'LOW_REPEAT', 'category': 'Mooring / lifting / SPM / ETA', 'topic': 'Venting / P-V valves / IGS', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ published to the OCIMF website on 24 Sept 2024 was randomly reviewed on 25 Sept 2024. Some questions were not responded and some questions were wrongly responded. For eg - 1.1.1, 1.2.4, 1.5.4, 1.9.8, 2.2.1, 5.3.2, 6.1.12, 7.1.6, 9.59.5, 9.68.6, 10.1.3, 10.4.2, 10.10.6, 11.10.1, 11.10.2, 12.4.1', 'HVPQ submitted through the CVIQ was found to be not completed, wrongly filled, not answered for the following sections as below: 1.1. 13, 1 .2.1, 1.2.4, 1.3.1, 1.3.2, 2. 1.4, 3.3.1, 4. 1. 1, 7. 1 .1, 7. 1.3, 9. 1 .1, 9.8.22, 9. 16.6, 9.17.1, 9.34.7, 10.1.3, 10. 1.4, 10.2.1, 10.4.2, 10.7.1, 10.8.1, 12. 1.14. For example: 1.1. 13 P and I Club# 3 Amount of P&I Cover; 9.16.6 What is the capacity of the IGS?-not responded; 1.3. 1 Registered Owner# 9 Number of years this ship has been owned by Registe', 'The HVPQ updated on 04 March 2025 was not accurately completed with respect to following items:\n1.5.9/1.5.10/2.2.1/5.3.2.10/8.2.3.3/9.6.6/9.8.14.3/9.15.1/9.30.6/10.4.2/10.9.1/11.1.6']}, {'question_no': '10.10.6', 'repeat_count': 6, 'priority': 'LOW_REPEAT', 'category': 'Mooring / lifting / SPM / ETA', 'topic': 'Mooring line / wire / tail installation age', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ published to the OCIMF website on 24 Sept 2024 was randomly reviewed on 25 Sept 2024. Some questions were not responded and some questions were wrongly responded. For eg - 1.1.1, 1.2.4, 1.5.4, 1.9.8, 2.2.1, 5.3.2, 6.1.12, 7.1.6, 9.59.5, 9.68.6, 10.1.3, 10.4.2, 10.10.6, 11.10.1, 11.10.2, 12.4.1', 'Question no. 4.1.1, 10.1.3.2, 10.1.6, 10.2.1, 10.8, 10.10.6, 10.10.7 & 11.3 were incomplete or blank.', 'The following items of the HVPQ dated 30 November 2024 were not updated : 1.3.1.5 (IMO of Owner); 1.3.1.10(Date when registered); 1.5.10(Date of last thickness measurements); 1.6.13-15, 1.6.17-1.6.19, 1.6.21-25, 1.6.28 (Distances and parallel body); 4.2.2(Communication equipment on board);9.16.18, 9.16.21 (Type of the deck seal and non-return valve);9.17.1 (Details of cargo pump); 10.1.8(Retirement policy); 10.2.2-3 (Details of bollards and fairleads); 10.4.2 ( Details of ETA); 10.8.1-4 (Manifol']}, {'question_no': '6.1.8', 'repeat_count': 6, 'priority': 'LOW_REPEAT', 'category': 'Pollution prevention', 'topic': 'Venting / P-V valves / IGS', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['HVPQ item 6.1.8&10, regarding cargo chest and item 9.16.5, regarding IGS space/room oxygen sensors installation were marked wrongly.', 'The vessel was not fitted with a cargo sea chest. HVPQ question 6.1.8 listed the type of valve fitted in the cargo sea chest.', 'On the uploaded HVPQ, some entries were either missing or not accurately answered: - Items 6.1.8, 9.6.6, 11.8.3, 12.1.3 & 12.1.4 had blank responses. - The date entered in the item 1.1.4.4 was one day ahead of the actual date of change of flag. - Item 1.2.3 stated that a CII rating of A was obtained and verified by the Class. Actually, the vessel was delivered on 18 September 2025 and was not due to receive its initial CII rating. - The numbers of certified officers and ratings recorded in the i']}, {'question_no': '1.5.14', 'repeat_count': 6, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Certificates / class survey dates / endorsements', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['Following were noted in HVPQ last updated 07 Feb 2025: - Entries missing in 1.5.14 & 10.1.8. Incorrect entries in 9.30.6, 10.6.3 & 10.9.1', 'The VPQ were errors as following; (a) 1.5.14 CLASS CONDITION: “No”-> Yes (The following class condition was appended on the last class survey status report dated 12 November 2024; (1) VESSEL INFORMED THE NO. 2 STEERING GEAR SYSTEM OF MALFUNCTIONING DUE TO A BURNT MOTOR. SAME TO BE RENEWED AND SPECIALLY SATISFACTION OF ATTENDING SURVEYOR. IMPOSED DATE: 10 NOVEMBER 2024 / DUE DATE: 15 NOVEMBER 2024) (2) THE FOLLOWING DEFICIENCIES REMAIN TO BE DEALT WITH X-BAND RADAR MALFUNCTIONING. IMPOSED DATE: 0', 'From the review of the HVPQ dated 13 May 2025, it was noted that the minimum safe manning certificate information (Section 3.1.1) was incorrect (7 Officers and 7 Ratings), while the MSM certificate required 8 Officers and 8 Ratings). In addition, the HVPQ section 1.1.13 indicated as P&I Club the North Standard Limited, while the vessel had UK P&I Club. Furthermore, HVPQ section 1.5.14 Condition of Class was answered as "No" while the vessel had a condition of Class issued for a shell plating ind']}, {'question_no': '12.1.4', 'repeat_count': 6, 'priority': 'LOW_REPEAT', 'category': 'Ice / special operations / VOC', 'topic': 'Certificates / class survey dates / endorsements', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['A review of the uploaded HVPQ dated 18 Feb 2025 indicated that the Q1.1.8 was not correctly updated as per the IOPPC which was stated that the vessel was an Oil Tanker. The HVPQ question Q12.1.4 and Q12.1.5 was answered in affirmative whereby generated a question in CVIQ for this inspection.', '(i) The HVPQ entries inside section Ref: 1.5.5.2/ 9.16.18/ 9.32.2(Dehumidifier not fitted)/ 11.3.1(KW incorrect) /11.7.1/ 11.7.4/ 12.1.4 did not accurately reflect the information relating to the ship at the time of inspection. The incorrect info & particulars were brought to the attention of Master.\n(ii) The vessel designation as recorded as per IOPP certificate was not declared inside the PIQ.', 'On the uploaded HVPQ, some entries were either missing or not accurately answered: - Items 6.1.8, 9.6.6, 11.8.3, 12.1.3 & 12.1.4 had blank responses. - The date entered in the item 1.1.4.4 was one day ahead of the actual date of change of flag. - Item 1.2.3 stated that a CII rating of A was obtained and verified by the Class. Actually, the vessel was delivered on 18 September 2025 and was not due to receive its initial CII rating. - The numbers of certified officers and ratings recorded in the i']}, {'question_no': '5.3.2', 'repeat_count': 5, 'priority': 'LOW_REPEAT', 'category': 'Safety / firefighting / lifeboats', 'topic': 'Safety / firefighting / lifeboats', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ published to the OCIMF website on 24 Sept 2024 was randomly reviewed on 25 Sept 2024. Some questions were not responded and some questions were wrongly responded. For eg - 1.1.1, 1.2.4, 1.5.4, 1.9.8, 2.2.1, 5.3.2, 6.1.12, 7.1.6, 9.59.5, 9.68.6, 10.1.3, 10.4.2, 10.10.6, 11.10.1, 11.10.2, 12.4.1', 'The HVPQ dated 21 May 2025 had the following errors: 5.3.2.8 - N/A Not Fitted. 5.3.8.2 - N/A Not Fitted. 5.3.8.3 - The Port Lifeboat was the dedicated rescue boat. 9.16.18 - Block & Bleed Arrangement. 10.6.8 - No - the SPM bracket (SWL 204MT) was fitted on the starboard side of the forward main deck with an approx 60 degree off set from the center lead and then via one pedestal roller with a wrap angle of approx 60 degrees to the winch pick up drum. 11.3.1 The vessel was equipped with 3 power ge', 'HVPQ submitted through the CVIQ was found to be not completed/ wrongly filled/ not answered for the following sections as below: 1.5.1,1.5.5,1.8.7, 4.1.1, 5.3.2, 7.1.6, 9.59.7, 10.2.1, 10.4.1, 10.7.1, 10.8.1, 10.10.7, 11.1.9, 11.2.2, 11.9.3, 12.4.1. For example, 1.5.1 Classification society #2 is Classification society is an IACS member, # does the ship have dual class ? 9.59.7 Compressors # Are they oil free ? There were no responses provided. 4.1.1 Navigational equipment fitted on board? # Glo']}, {'question_no': '11.10.1', 'repeat_count': 5, 'priority': 'LOW_REPEAT', 'category': 'Engine room / machinery / generators', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ published to the OCIMF website on 24 Sept 2024 was randomly reviewed on 25 Sept 2024. Some questions were not responded and some questions were wrongly responded. For eg - 1.1.1, 1.2.4, 1.5.4, 1.9.8, 2.2.1, 5.3.2, 6.1.12, 7.1.6, 9.59.5, 9.68.6, 10.1.3, 10.4.2, 10.10.6, 11.10.1, 11.10.2, 12.4.1', 'The HVPQ showed erroneous dates and data as follows:\n1.1.13.4: P&I cover include wreck removal no response\n1.5.5.2: incorrect date\n1.5.6.5: next special date no response\n1.9.8.1: PSC date incorrect\n3.1.6.1: should be yes\n3.1.6.2: should be no\n3.1.10: same as for officers no response needed\n4.1.1: type of navigation equipments not completed\n10.9.1: crane annual test date incorrect\n11.1.13.1: quick closing valves fitted\n11.1.13.2: no response\n11.4.3: no response\n11.10.1: no response\n12.1.1: should', 'The HVPQ showed erroneous data and dates as below: 1.1.13 the name of P&I was not correct 3.1.9 rating as per MSMC only 8 instead of 9 10.9.1 cranes annual test incorrect 11.10.1 not filled.']}, {'question_no': '9.16.18', 'repeat_count': 5, 'priority': 'LOW_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'Tank coating / structural inspection dates and frequency', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ dated 21 May 2025 had the following errors: 5.3.2.8 - N/A Not Fitted. 5.3.8.2 - N/A Not Fitted. 5.3.8.3 - The Port Lifeboat was the dedicated rescue boat. 9.16.18 - Block & Bleed Arrangement. 10.6.8 - No - the SPM bracket (SWL 204MT) was fitted on the starboard side of the forward main deck with an approx 60 degree off set from the center lead and then via one pedestal roller with a wrap angle of approx 60 degrees to the winch pick up drum. 11.3.1 The vessel was equipped with 3 power ge', 'The type of deck seal was wrongly declared as semi-dry in HVPQ item 9.16.18. The vessel was fitted with a wet type deck seal instead.', 'The following typo errors were seen in the latest HVPQ uploaded in the OCIMF repository, dated 12 May 2025: 1) HVPQ 7.1.1: cargo tanks inspection interval quarterly instead annually. 2) HVPQ 7.1.3: last 5 side starboard ballast tank inspection dated 12 July 2024 instead 12 March 2025. 3) HVPQ 9.16.18: missing the entry bleed and block valves. 4) HVPQ 10.1.3: missing entry SDMBL 36.5 tonnes. 5) HVPQ 10.1.4: the details of only one mooring winch was recorded. The vessel was fitted with 8 mooring w']}, {'question_no': '1.2.2', 'repeat_count': 5, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Environmental indices', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['1.2.2 Of the HVPQ indicated the vessel was not assigned an EEXI Rating. The vessel had been assigned an EEXI Rating', 'Some erroneous and/or missing information noted under the following sections: 1.2.2.1; 1.5.4.2; 1.5.5; 7.1.3; 10.1.4.', 'As per Supplement to the IEEC, this vessel EEXI rating was 1.97. In HVPQ, under 1.2.2 "Does the vessel has EEXI rating" was answered "No". (EEXI rating 1.97 was mentioned in the following HVPQ question).']}, {'question_no': '6.1.1', 'repeat_count': 5, 'priority': 'LOW_REPEAT', 'category': 'Pollution prevention', 'topic': 'Cargo pump details', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ uploaded by the Operator on 12 Mar 2025 had errors including:\na) 6.1.1.4- Incorrect distance for coaming of height 260mm.\nb) 9.32.2.1- The vessel has a dehumidifier.\nc) 10.6.2- The vessel was fitted with only 1 bow stopper.', 'It was observed that the HVPQ uploaded by the Operator on 09 Dec 2024 had significant errors at various locations including a)\n6.1.1- Various coaming heights , b) 9.5.3- Loading through pump cannot be bypassed, c) 9.7.3- All valves could be operated from\nthe CCR and e) 9.16.19- Vessel had seven IG segregations', 'The following items of the HVPQ incorrect information was noted:\n1.2.3.- Stated CII E, but actual rating was B.\n6.1.1/4.- "How far forward the athwartships coaming is this height maintained" stated wrong value 12800 m.\n9.11.2.- Manifold hydraulic valves time 10 seconds; but the manifold had not hydraulic valves.\nChapter 13 was filled however the ship was not a Combination Carrier']}, {'question_no': '1.8.7', 'repeat_count': 5, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'General / certificates / PSC / ownership', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['HVPQ submitted through the CVIQ was found to be not completed/ wrongly filled/ not answered for the following sections as below: 1.5.1,1.5.5,1.8.7, 4.1.1, 5.3.2, 7.1.6, 9.59.7, 10.2.1, 10.4.1, 10.7.1, 10.8.1, 10.10.7, 11.1.9, 11.2.2, 11.9.3, 12.4.1. For example, 1.5.1 Classification society #2 is Classification society is an IACS member, # does the ship have dual class ? 9.59.7 Compressors # Are they oil free ? There were no responses provided. 4.1.1 Navigational equipment fitted on board? # Glo', 'In the HVPQ dated 16 January 2025 uploaded in the system was observed: 8. Load Line Information : 1.8.6 Assigned dead weight 1 : 49999 MT, however on board was noted the correct one : 49795.7 MT. 1.8.7 What is the current in use assigned dwt: it was stated: 497796.00 MT. The correct is 49795.7 MT.', 'The HVPQ was shown erroneous data below:\n1.5.6.3/1.8.7/6.1.14.2']}, {'question_no': '1.9.5', 'repeat_count': 5, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The vessel reported two incidents in Nov-2024 and the item 1.9.5 in the HVPQ stated that she was not involved in any incidents during the past 12 months.', "On HVPQ, several items were recorded with inaccurate information: \n- The HVPQ declared a No to the item 1.9.5, however the vessel had three reported incidents on record in the previous 12 months. \n- The HVPQ declared a Yes to the items 6.2.1, 6.2.2 & 6.2.3, but the vessel didn't hold a valid USCG VRP Approval Letter or COFR. \n- The HVPQ declared a Yes to the item 9.7.3, actually all manifold valves were manual. \n- The safe working load (SWL) of each provision crane and engine room crane was ente", 'HVPQ dated 16th October 2024 on 1.5.19, 1.9.5, 7.1.3 and 10.1.4 were not correct.']}, {'question_no': '11.4.3', 'repeat_count': 5, 'priority': 'LOW_REPEAT', 'category': 'Engine room / machinery / generators', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ showed erroneous dates and data as follows:\n1.1.13.4: P&I cover include wreck removal no response\n1.5.5.2: incorrect date\n1.5.6.5: next special date no response\n1.9.8.1: PSC date incorrect\n3.1.6.1: should be yes\n3.1.6.2: should be no\n3.1.10: same as for officers no response needed\n4.1.1: type of navigation equipments not completed\n10.9.1: crane annual test date incorrect\n11.1.13.1: quick closing valves fitted\n11.1.13.2: no response\n11.4.3: no response\n11.10.1: no response\n12.1.1: should', 'The following entries were either incomplete or incorrect - 1.5.1, 1.5.12, 6.1.5, 9.6.2, 9.6.6, 9.8.17, 9.15.1, 9.16.11, 9.16.12, 10.9.1, 11.3.4, 11.4.3, 11.5.5.2, 9.10.9.2 to 4.', 'A review of the uploaded HVPQ indicated that the below items were found to be omitted or incorrectly declared: 1) 1.1.8/IOPP designation was wrongly listed as “Others” instead of the correct option “GAS” which could be selected from the drop down options: 2) 1.1.13.1 P&I Club was wrongly listed as “Other” instead of the correct club – “North Standard Limited” which could be selected from the drop down opions 3)11.4.3/Motive power of emergency compressor was not declared.']}, {'question_no': '6.1.14', 'repeat_count': 5, 'priority': 'LOW_REPEAT', 'category': 'Pollution prevention', 'topic': 'Mooring line / wire / tail installation age', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ was shown erroneous data below:\n1.5.6.3/1.8.7/6.1.14.2', 'PROCESS - The following was noted not updated accurately as per the HVPQ dated 26-Nov-2024: a) 2.2.1 - The edition numbers for the publications were not provided as required. b) 6.1.14 - The bunker pipeline annual pressure test was incorrectly stated as 6.60 bars instead of 4.40 bars. c) 7.1.3 - The FPT, 1W and 3W ballast tanks annual inspection dates were not updated and indicated as overdue instead. d) 10.1.3.2 - The diagram for the mooring winch layout was not provided. e) 10.7.1 - The bow mo', 'The HVPQ was last updated by the operator on 27th March 2025 and the followings were noted - Items 6.1.14: The bunker pipeline annual pressure test was incorrectly stated as 6.60 bars, instead of 4.40 bars. - Item 10.1.3.1: The diagram for the mooring winch layout was not provided. Item 12.1.8 - details for the bollards, chock, rollers and fairleads used for the STS operation indicated that the winches / capstan driven by the electric power, instead of hydraulic power.']}, {'question_no': '1.8.1', 'repeat_count': 5, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Tank coating / structural inspection dates and frequency', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ submitted through the CVIQ was found to be not completed/wrongly filled/ not answered for the following sections as\nbelow: 1.5.5, 1.5.19, 1.8.1, 1.9.8, 2.2.1, 3.1.9, 4.1.1, 4.2.2, 10.1.3, 10.4.2, 10.6.3, 10.6.6, 10.6.7, 10.7.1, 10.8.1, 10.8.2-10.8.4, 11.1.8,\n11.1.9, 11.1.13, 11.2.2, 11.3.1, 11.3.4, 11.9.1, 12.1.7, 12.1.8, 12.1.14, 12.2.4, 12.4.1. For example: 1.9.8 Port State Control # 1 Date of\nlast Port State Control inspection and # 2 Port of last Port State Control inspection - were', 'The HVPQ uploaded by the Operator on 21st April 2025 had a few errors as follow – \na) 1.8.7 and 1.8.1 – load line information incorrect \nb) 7.1.3 – Incorrect dates for ballast tank inspections.', "Review of the ship's operator latest HVPQ completed on 14 Feb 2025 and uploaded on 15 Feb 2025 the following inaccuracies were observed: (1) Item 1.5.5 The date of last In Water Survey (IWS) was entered as 23 July 2024, however, from record this was the date of dry docking. (2) Item 1.8.1 The vessel had no multiple dead weight, vessel maximum summer dead weight was entered as 297 187.6, however, as recorded in ClasNK certificates and Certificate of Registry, the dead weight was 297 572 MT. (3) I"]}, {'question_no': '12.1.8', 'repeat_count': 5, 'priority': 'LOW_REPEAT', 'category': 'Ice / special operations / VOC', 'topic': 'Mooring line / wire / tail installation age', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ submitted through the CVIQ was found to be not completed/wrongly filled/ not answered for the following sections as\nbelow: 1.5.5, 1.5.19, 1.8.1, 1.9.8, 2.2.1, 3.1.9, 4.1.1, 4.2.2, 10.1.3, 10.4.2, 10.6.3, 10.6.6, 10.6.7, 10.7.1, 10.8.1, 10.8.2-10.8.4, 11.1.8,\n11.1.9, 11.1.13, 11.2.2, 11.3.1, 11.3.4, 11.9.1, 12.1.7, 12.1.8, 12.1.14, 12.2.4, 12.4.1. For example: 1.9.8 Port State Control # 1 Date of\nlast Port State Control inspection and # 2 Port of last Port State Control inspection - were', 'The following errors were noted in HVPQ uploaded on SIRE. - 3. 1. 1 - Minimum officers required as \nper MSMD were 6. - 9. 16. 10. 1 I 2 - Details not provided. - 9. 17. 1 - The cargo pump details of slop wing tanks were not recorded. - 9.30.1- Vessel was assigned IMO Type 2 I 3 chemical tanker notation. - 9.31. 10-IGS composition details not completed. - 10.1.4- The mooring winches were \nfitted with split drums. - 10.8. 1 - Manifold arrangement diagram not completed. - 10. 9. 1 - Last periodical', 'PROCESS - The following was noted not updated accurately as per the HVPQ dated 26-Nov-2024: a) 2.2.1 - The edition numbers for the publications were not provided as required. b) 6.1.14 - The bunker pipeline annual pressure test was incorrectly stated as 6.60 bars instead of 4.40 bars. c) 7.1.3 - The FPT, 1W and 3W ballast tanks annual inspection dates were not updated and indicated as overdue instead. d) 10.1.3.2 - The diagram for the mooring winch layout was not provided. e) 10.7.1 - The bow mo']}, {'question_no': '11.3.4', 'repeat_count': 4, 'priority': 'LOW_REPEAT', 'category': 'Engine room / machinery / generators', 'topic': 'Generators / emergency power', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ dated 21 May 2025 had the following errors: 5.3.2.8 - N/A Not Fitted. 5.3.8.2 - N/A Not Fitted. 5.3.8.3 - The Port Lifeboat was the dedicated rescue boat. 9.16.18 - Block & Bleed Arrangement. 10.6.8 - No - the SPM bracket (SWL 204MT) was fitted on the starboard side of the forward main deck with an approx 60 degree off set from the center lead and then via one pedestal roller with a wrap angle of approx 60 degrees to the winch pick up drum. 11.3.1 The vessel was equipped with 3 power ge', 'The HVPQ submitted through the CVIQ was found to be not completed/wrongly filled/ not answered for the following sections as\nbelow: 1.5.5, 1.5.19, 1.8.1, 1.9.8, 2.2.1, 3.1.9, 4.1.1, 4.2.2, 10.1.3, 10.4.2, 10.6.3, 10.6.6, 10.6.7, 10.7.1, 10.8.1, 10.8.2-10.8.4, 11.1.8,\n11.1.9, 11.1.13, 11.2.2, 11.3.1, 11.3.4, 11.9.1, 12.1.7, 12.1.8, 12.1.14, 12.2.4, 12.4.1. For example: 1.9.8 Port State Control # 1 Date of\nlast Port State Control inspection and # 2 Port of last Port State Control inspection - were', 'The following entries were either incomplete or incorrect - 1.5.1, 1.5.12, 6.1.5, 9.6.2, 9.6.6, 9.8.17, 9.15.1, 9.16.11, 9.16.12, 10.9.1, 11.3.4, 11.4.3, 11.5.5.2, 9.10.9.2 to 4.']}, {'question_no': '9.17.1', 'repeat_count': 4, 'priority': 'LOW_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'Tank coating / structural inspection dates and frequency', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['HVPQ submitted through the CVIQ was found to be not completed, wrongly filled, not answered for the following sections as below: 1.1. 13, 1 .2.1, 1.2.4, 1.3.1, 1.3.2, 2. 1.4, 3.3.1, 4. 1. 1, 7. 1 .1, 7. 1.3, 9. 1 .1, 9.8.22, 9. 16.6, 9.17.1, 9.34.7, 10.1.3, 10. 1.4, 10.2.1, 10.4.2, 10.7.1, 10.8.1, 12. 1.14. For example: 1.1. 13 P and I Club# 3 Amount of P&I Cover; 9.16.6 What is the capacity of the IGS?-not responded; 1.3. 1 Registered Owner# 9 Number of years this ship has been owned by Registe', 'The following items of the HVPQ dated 30 November 2024 were not updated : 1.3.1.5 (IMO of Owner); 1.3.1.10(Date when registered); 1.5.10(Date of last thickness measurements); 1.6.13-15, 1.6.17-1.6.19, 1.6.21-25, 1.6.28 (Distances and parallel body); 4.2.2(Communication equipment on board);9.16.18, 9.16.21 (Type of the deck seal and non-return valve);9.17.1 (Details of cargo pump); 10.1.8(Retirement policy); 10.2.2-3 (Details of bollards and fairleads); 10.4.2 ( Details of ETA); 10.8.1-4 (Manifol', 'The HVPQ uploaded had the following discrepancies: 1.2.4 - EIV rating stated available in 2024, not entered. 1.5.14 - Open conditions of class not entered 7.1.1 - Cargo tank inspections frequency stated as annual, last inspection on 11 & 12 May 2024 except for cargo tanks 4S & 5W. 7.1.3 - Ballast tank inspections frequency stated as annual, apart from forepeak tank, all others were last inspected on 18, 19 & 25 May 2024. 9.6.2, 9.6.6 & 9.17.1- Cargo pump types stated "centrifugal", whereas deep ']}, {'question_no': '3.1.3', 'repeat_count': 4, 'priority': 'LOW_REPEAT', 'category': 'Crew / training / operator assessments', 'topic': 'Tank coating / structural inspection dates and frequency', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following discrepancies were noted in the HVPQ: - Item 1.5.4.2: Lack of necessary information. - Item 1.5.5.1: Error in the due date recording, listed incorrectly as 03 April 2024 instead of 03 April 2022. - Item 1.5.11: Lack of necessary information. - Item 1.5.12: Lack of necessary information. - Item 1.5.16: An error discovered in the recording of the answer, which was incorrectly marked as NO instead of YES. - Item 1.5.17: Lack of necessary information. - Item 3.1.3: The number of rating', 'The following discrepancies were noted 3.1.3 Number of ratings declared incorrectly as 24 instead of an board 13. SDO stated the same would be rectified 4.1.1 Software and firmware data was not populated, SDO stated that during the last annuals the technician had not left behind such data while the vessel was contracted for annual software upgrade 4.2.2 EPIRB data was not populated 7.1.1 Cargo tank coating inspection interval was incorrectly stated as 30 months instead of 60 months as per operat', 'The following discrepancies were found in the HVPQ document:\nItem 1.5.3.1: An error was discovered in the recording of the answer, which was incorrectly marked as NO instead of YES.\nItem 1.5.3.2: Lack of the necessary information.\nItem 1.5.3.3: Lack of the necessary information.\nItem 1.5.5.1: Lack of the necessary information.\nItem 1.5.5.2: Lack of the necessary information.\nItem 1.5.10: Error in recording the date, listed incorrectly 26 March 2024 instead 12 March 2024.\nItem 1.5.12: Lack of the']}, {'question_no': '1.2.3', 'repeat_count': 4, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Certificates / class survey dates / endorsements', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['information provided in the HVPQ by the vessel\'s operator completed on 30 Jun 2025 downloaded on 30 Jun 2025 from OCIMF\nwebsite was inacurate. Under HVPQ No. 1.2.3.2 CII rating was defined as "A" instead of "B" as per certificate number DCS-9637088-\n2024-242665. Under HVPQ No. 1.2.4.2 EIV rating was defined as "5.96" instead of "5.544" as per certificate number DCS-9637088-\n2024-242665.', 'On the uploaded HVPQ, some entries were either missing or not accurately answered: - Items 6.1.8, 9.6.6, 11.8.3, 12.1.3 & 12.1.4 had blank responses. - The date entered in the item 1.1.4.4 was one day ahead of the actual date of change of flag. - Item 1.2.3 stated that a CII rating of A was obtained and verified by the Class. Actually, the vessel was delivered on 18 September 2025 and was not due to receive its initial CII rating. - The numbers of certified officers and ratings recorded in the i', '(i) The HVPQ entries inside section Ref: 1.2.3.2/4 (CII-B by Class ABS)/ 1.5.14 (Condition of Class on LRIT conformance test)/ 1.5.5.2/\n9.30.17.2/ 9.30.1( IMO Type 2) / 10.9.1(Last Class annual examination) / 10.10.6/ 121.4 did not accurately reflect the information\nrelating to the ship at the time of inspection. The incorrect info and particulars were brought to the attention of the Master.\n(ii) The vessel designation as recorded as per IOPP cert ., and the name of vessel P&I club were not decl']}, {'question_no': '1.5.10', 'repeat_count': 4, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Certificates / class survey dates / endorsements', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following discrepancies were noted in the HVPQ document: - Item 1.3.1.6: Lack of the necessary information. - Item 1.5.10: Error in the recording of the date, listed incorrectly as 28 September 2022 instead of 19 September 2022.. - Item 1.5.12: Lack of the necessary information. - Item 3.1.1: Error in the recording of the quantity, listed incorrectly as 07 instead of 06. - Item 3.1.9: Error in the recording of the quantity, listed incorrectly as 08 instead of 07. - Item 9.10.4: Error in the ', 'The HVPQ updated on 04 March 2025 was not accurately completed with respect to following items:\n1.5.9/1.5.10/2.2.1/5.3.2.10/8.2.3.3/9.6.6/9.8.14.3/9.15.1/9.30.6/10.4.2/10.9.1/11.1.6', 'The following discrepancies were found in the HVPQ document:\nItem 1.5.3.1: An error was discovered in the recording of the answer, which was incorrectly marked as NO instead of YES.\nItem 1.5.3.2: Lack of the necessary information.\nItem 1.5.3.3: Lack of the necessary information.\nItem 1.5.5.1: Lack of the necessary information.\nItem 1.5.5.2: Lack of the necessary information.\nItem 1.5.10: Error in recording the date, listed incorrectly 26 March 2024 instead 12 March 2024.\nItem 1.5.12: Lack of the']}, {'question_no': '10.6.3', 'repeat_count': 4, 'priority': 'LOW_REPEAT', 'category': 'Mooring / lifting / SPM / ETA', 'topic': 'Environmental indices', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['Following were noted in HVPQ last updated 07 Feb 2025: - Entries missing in 1.5.14 & 10.1.8. Incorrect entries in 9.30.6, 10.6.3 & 10.9.1', 'On the uploaded HVPQ, some entries were either missing or not accurately answered: - Items 6.1.8, 9.6.6, 11.8.3, 12.1.3 & 12.1.4 had blank responses. - The date entered in the item 1.1.4.4 was one day ahead of the actual date of change of flag. - Item 1.2.3 stated that a CII rating of A was obtained and verified by the Class. Actually, the vessel was delivered on 18 September 2025 and was not due to receive its initial CII rating. - The numbers of certified officers and ratings recorded in the i', 'The HVPQ submitted through the CVIQ was found to be not completed/wrongly filled/ not answered for the following sections as\nbelow: 1.5.5, 1.5.19, 1.8.1, 1.9.8, 2.2.1, 3.1.9, 4.1.1, 4.2.2, 10.1.3, 10.4.2, 10.6.3, 10.6.6, 10.6.7, 10.7.1, 10.8.1, 10.8.2-10.8.4, 11.1.8,\n11.1.9, 11.1.13, 11.2.2, 11.3.1, 11.3.4, 11.9.1, 12.1.7, 12.1.8, 12.1.14, 12.2.4, 12.4.1. For example: 1.9.8 Port State Control # 1 Date of\nlast Port State Control inspection and # 2 Port of last Port State Control inspection - were']}, {'question_no': '9.6.6', 'repeat_count': 4, 'priority': 'LOW_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'Cargo pump details', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['On the uploaded HVPQ, some entries were either missing or not accurately answered: - Items 6.1.8, 9.6.6, 11.8.3, 12.1.3 & 12.1.4 had blank responses. - The date entered in the item 1.1.4.4 was one day ahead of the actual date of change of flag. - Item 1.2.3 stated that a CII rating of A was obtained and verified by the Class. Actually, the vessel was delivered on 18 September 2025 and was not due to receive its initial CII rating. - The numbers of certified officers and ratings recorded in the i', 'The HVPQ updated on 04 March 2025 was not accurately completed with respect to following items:\n1.5.9/1.5.10/2.2.1/5.3.2.10/8.2.3.3/9.6.6/9.8.14.3/9.15.1/9.30.6/10.4.2/10.9.1/11.1.6', 'The following entries were either incomplete or incorrect - 1.5.1, 1.5.12, 6.1.5, 9.6.2, 9.6.6, 9.8.17, 9.15.1, 9.16.11, 9.16.12, 10.9.1, 11.3.4, 11.4.3, 11.5.5.2, 9.10.9.2 to 4.']}, {'question_no': '12.1.3', 'repeat_count': 4, 'priority': 'LOW_REPEAT', 'category': 'Ice / special operations / VOC', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['On the uploaded HVPQ, some entries were either missing or not accurately answered: - Items 6.1.8, 9.6.6, 11.8.3, 12.1.3 & 12.1.4 had blank responses. - The date entered in the item 1.1.4.4 was one day ahead of the actual date of change of flag. - Item 1.2.3 stated that a CII rating of A was obtained and verified by the Class. Actually, the vessel was delivered on 18 September 2025 and was not due to receive its initial CII rating. - The numbers of certified officers and ratings recorded in the i', 'The following items in the HVPQ were not correctly completed: 10.7.1; 10.8.1; 12.1.3 and 12.1.4.', 'The OP uploaded the HVPQ dated 22 September 2025 were inaccuracy information as: a. the previous management company name in column (1.3.2 (12). b. the vessel was not certified for a person\'s transferring by deck crane, but in column (12.1.3) was answered with "YES".']}, {'question_no': '10.10.7', 'repeat_count': 4, 'priority': 'LOW_REPEAT', 'category': 'Mooring / lifting / SPM / ETA', 'topic': 'Mooring / lifting / SPM / ETA', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['HVPQ submitted through the CVIQ was found to be not completed/ wrongly filled/ not answered for the following sections as below: 1.5.1,1.5.5,1.8.7, 4.1.1, 5.3.2, 7.1.6, 9.59.7, 10.2.1, 10.4.1, 10.7.1, 10.8.1, 10.10.7, 11.1.9, 11.2.2, 11.9.3, 12.4.1. For example, 1.5.1 Classification society #2 is Classification society is an IACS member, # does the ship have dual class ? 9.59.7 Compressors # Are they oil free ? There were no responses provided. 4.1.1 Navigational equipment fitted on board? # Glo', 'Question no. 4.1.1, 10.1.3.2, 10.1.6, 10.2.1, 10.8, 10.10.6, 10.10.7 & 11.3 were incomplete or blank.', 'The uploaded HVPQ had the following discrepancies 2.1.5 – SMC & ISSC date of endorsement 10.1.4 – split drums stated no, vessel fitted with split drums 10.2.4 – Panama or closed chocks as per MEG 4 10.9.1 – lifting appliances last annual test on 04th Feb 2024, vessel carried out annual surveys on 15th May 2025. 10.10.7 shore gangway designated landing area.']}, {'question_no': '11.1.13', 'repeat_count': 4, 'priority': 'LOW_REPEAT', 'category': 'Engine room / machinery / generators', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ showed erroneous dates and data as follows:\n1.1.13.4: P&I cover include wreck removal no response\n1.5.5.2: incorrect date\n1.5.6.5: next special date no response\n1.9.8.1: PSC date incorrect\n3.1.6.1: should be yes\n3.1.6.2: should be no\n3.1.10: same as for officers no response needed\n4.1.1: type of navigation equipments not completed\n10.9.1: crane annual test date incorrect\n11.1.13.1: quick closing valves fitted\n11.1.13.2: no response\n11.4.3: no response\n11.10.1: no response\n12.1.1: should', 'The HVPQ submitted through the CVIQ was found to be not completed/wrongly filled/ not answered for the following sections as\nbelow: 1.5.5, 1.5.19, 1.8.1, 1.9.8, 2.2.1, 3.1.9, 4.1.1, 4.2.2, 10.1.3, 10.4.2, 10.6.3, 10.6.6, 10.6.7, 10.7.1, 10.8.1, 10.8.2-10.8.4, 11.1.8,\n11.1.9, 11.1.13, 11.2.2, 11.3.1, 11.3.4, 11.9.1, 12.1.7, 12.1.8, 12.1.14, 12.2.4, 12.4.1. For example: 1.9.8 Port State Control # 1 Date of\nlast Port State Control inspection and # 2 Port of last Port State Control inspection - were', '"The following were either not completed or wrongly stated in the HVPQ:\n1.3.1. address of the owner did not match the address on certificates such as, CLC, BCLC, certificate of registry, CSSR and CSR.\n9.1.1 - tank plan\n9.16.5 - O2 alarms in IG spaces\n10.1.3.2 - mooring winch layout diagram\n10.2.1 - mooring fairleads/chocks and bollards / bitts diagram\n10.7.1 - bow mooring arrangement diagram\n10.8.1 - manifold arrangement diagram\n11.1.13 - quick closing valves"']}, {'question_no': '1.4.3', 'repeat_count': 3, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['Two items were incorrectly declared in the latest HVPQ6 dated 3 October 2024 asf: 1-Item 1.4.3- Date of building contract 20-January-2019; Actually on 18-May-2018: 2-item 10.9.1 – last 5yr test for one of two cranes : 20-January-2011: Actually on 20-January-2022. Reportedly online updated immediately.', 'The following items were incorrectly declared in the latest HVPQ6 dated 07 November 2024: 1-1.4.3 Date of building contract: 23 August 2019 instead of actual 22 August 2019; 2-10.9.1 Last annual test for 5 cranes: 12 November 2023 instead of actual 19 June 2024. Reportedly all on-line updated immediately.', 'Certain information in the HVPQ downloaded on 23 January 2026 was inaccurate or not updated in the following sections; \nQ. 1.4.3 (To be 02 Nov 2020) / \n2.1.5 (Issue dates of Safety Radio Certificate and Fitness Certificate) / \n3.1.3 / 3.1.9 / 4.1.1 (Model, software and firmware versions) / \n4.2.2 (Serial number, software and firmware versions) / \n6.1.1.4 / 6.1.14.2 (To be 5 bar) / \n7.1.3 (Date when tank was coated) / \n9.67.3 / 10.9.1 (Last annual test) / \n11.4.3 / 12.2.4.']}, {'question_no': '12.4.1', 'repeat_count': 3, 'priority': 'LOW_REPEAT', 'category': 'Ice / special operations / VOC', 'topic': 'Ice / special operations / VOC', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ published to the OCIMF website on 24 Sept 2024 was randomly reviewed on 25 Sept 2024. Some questions were not responded and some questions were wrongly responded. For eg - 1.1.1, 1.2.4, 1.5.4, 1.9.8, 2.2.1, 5.3.2, 6.1.12, 7.1.6, 9.59.5, 9.68.6, 10.1.3, 10.4.2, 10.10.6, 11.10.1, 11.10.2, 12.4.1', 'HVPQ submitted through the CVIQ was found to be not completed/ wrongly filled/ not answered for the following sections as below: 1.5.1,1.5.5,1.8.7, 4.1.1, 5.3.2, 7.1.6, 9.59.7, 10.2.1, 10.4.1, 10.7.1, 10.8.1, 10.10.7, 11.1.9, 11.2.2, 11.9.3, 12.4.1. For example, 1.5.1 Classification society #2 is Classification society is an IACS member, # does the ship have dual class ? 9.59.7 Compressors # Are they oil free ? There were no responses provided. 4.1.1 Navigational equipment fitted on board? # Glo', 'The HVPQ submitted through the CVIQ was found to be not completed/wrongly filled/ not answered for the following sections as\nbelow: 1.5.5, 1.5.19, 1.8.1, 1.9.8, 2.2.1, 3.1.9, 4.1.1, 4.2.2, 10.1.3, 10.4.2, 10.6.3, 10.6.6, 10.6.7, 10.7.1, 10.8.1, 10.8.2-10.8.4, 11.1.8,\n11.1.9, 11.1.13, 11.2.2, 11.3.1, 11.3.4, 11.9.1, 12.1.7, 12.1.8, 12.1.14, 12.2.4, 12.4.1. For example: 1.9.8 Port State Control # 1 Date of\nlast Port State Control inspection and # 2 Port of last Port State Control inspection - were']}, {'question_no': '5.3.8', 'repeat_count': 3, 'priority': 'LOW_REPEAT', 'category': 'Safety / firefighting / lifeboats', 'topic': 'Venting / P-V valves / IGS', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ dated 21 May 2025 had the following errors: 5.3.2.8 - N/A Not Fitted. 5.3.8.2 - N/A Not Fitted. 5.3.8.3 - The Port Lifeboat was the dedicated rescue boat. 9.16.18 - Block & Bleed Arrangement. 10.6.8 - No - the SPM bracket (SWL 204MT) was fitted on the starboard side of the forward main deck with an approx 60 degree off set from the center lead and then via one pedestal roller with a wrap angle of approx 60 degrees to the winch pick up drum. 11.3.1 The vessel was equipped with 3 power ge', 'Incorrect information was provided in the HVPQ section 5.3.8 indicating that the vessel was fitted with a dedicated rescue boat, but the vessel was fitted with conventional lifeboats only, which port side was the designated rescue boat.', "Various questions in sections 1.2.2 to 1.2.4 of HVPQ were left blank. (2) as per section 5.3.8 of HVPQ vessel was provided with dedicated rescue boat. In actual vessel's starboard side davit launched conventional lifeboat was the designated rescue boat."]}, {'question_no': '2.1', 'repeat_count': 3, 'priority': 'LOW_REPEAT', 'category': 'Environmental / certificates', 'topic': 'Venting / P-V valves / IGS', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['HVPQ submitted through the CVIQ was found to be not completed, wrongly filled, not answered for the following sections as below: 1.1. 13, 1 .2.1, 1.2.4, 1.3.1, 1.3.2, 2. 1.4, 3.3.1, 4. 1. 1, 7. 1 .1, 7. 1.3, 9. 1 .1, 9.8.22, 9. 16.6, 9.17.1, 9.34.7, 10.1.3, 10. 1.4, 10.2.1, 10.4.2, 10.7.1, 10.8.1, 12. 1.14. For example: 1.1. 13 P and I Club# 3 Amount of P&I Cover; 9.16.6 What is the capacity of the IGS?-not responded; 1.3. 1 Registered Owner# 9 Number of years this ship has been owned by Registe', 'The following HVPQ items were not corrected from the previous observations provided in the PIQ:- 1.5.11 / 1.5.12 / 9.1.1 /10.1.3.2.1 / 10.2.1 & 10.7.1. They were diagrams, and OP reported that operator was not able to upload', 'According to the attached HVPQ, the following items were found: \n1) 1.3.1.10 the owner took over from Dec 2019; \n2) 1.5.16/ 1.5.17/1.5.18/ 1.5.19 were blank, one exemption was kept at the time of inspection (9 nos. level gauges of cargo tanks); \n3) 2.1 parts of the annual survey were marked on 13 Nov 2023, including SCC/ IOPP/ IBWM; \n4) 2.2.1 Parts of publications were not the latest editions: ICS Guide to helicopter/ Ship Operations, Guidance Manual for Tanker structures; \n5) 7.1.3 Parts of bal']}, {'question_no': '1.5.16', 'repeat_count': 3, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following discrepancies were noted in the HVPQ: - Item 1.5.4.2: Lack of necessary information. - Item 1.5.5.1: Error in the due date recording, listed incorrectly as 03 April 2024 instead of 03 April 2022. - Item 1.5.11: Lack of necessary information. - Item 1.5.12: Lack of necessary information. - Item 1.5.16: An error discovered in the recording of the answer, which was incorrectly marked as NO instead of YES. - Item 1.5.17: Lack of necessary information. - Item 3.1.3: The number of rating', 'Following discrepancies noted.\n1.3.11 Total number of this type of ships was incorrectly stated as 500.\n1.5.16 was marked as NO while M05 and M08 were noted on CSSR.\n4.1.1 serial numbers of the navigation equipment were. ot populated.\n4.2.2 EPIRB and SART data was not populated.\n10.1.7 Data for mooring tails and shackle were not populated.\n10.9.1 Date was last crane annual inspection was incorrectly stated as 29 March 2025.\n7.1.3 While 1P/S TST and D/B were inspected in May 2025 while the other ', 'According to the attached HVPQ, the following items were found: \n1) 1.3.1.10 the owner took over from Dec 2019; \n2) 1.5.16/ 1.5.17/1.5.18/ 1.5.19 were blank, one exemption was kept at the time of inspection (9 nos. level gauges of cargo tanks); \n3) 2.1 parts of the annual survey were marked on 13 Nov 2023, including SCC/ IOPP/ IBWM; \n4) 2.2.1 Parts of publications were not the latest editions: ICS Guide to helicopter/ Ship Operations, Guidance Manual for Tanker structures; \n5) 7.1.3 Parts of bal']}, {'question_no': '1.5.17', 'repeat_count': 3, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Tank coating / structural inspection dates and frequency', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following discrepancies were noted in the HVPQ: - Item 1.5.4.2: Lack of necessary information. - Item 1.5.5.1: Error in the due date recording, listed incorrectly as 03 April 2024 instead of 03 April 2022. - Item 1.5.11: Lack of necessary information. - Item 1.5.12: Lack of necessary information. - Item 1.5.16: An error discovered in the recording of the answer, which was incorrectly marked as NO instead of YES. - Item 1.5.17: Lack of necessary information. - Item 3.1.3: The number of rating', 'The HVPQ uploaded contained the following discrepancies: -Item 1.5.17: Memoranda described under Observations in 2.1.1 (hardware observations) were not detailed in the HVPQ. -Item 7. T. 1: Inspection intervals of the cargo tanks was 60 months (every class special survey) as per OperatorS procedures. The cargo tanks were uncoated but entries in the HVPQ included the date of coating and date of the last coating inspection for the cargo tanks.', 'According to the attached HVPQ, the following items were found: \n1) 1.3.1.10 the owner took over from Dec 2019; \n2) 1.5.16/ 1.5.17/1.5.18/ 1.5.19 were blank, one exemption was kept at the time of inspection (9 nos. level gauges of cargo tanks); \n3) 2.1 parts of the annual survey were marked on 13 Nov 2023, including SCC/ IOPP/ IBWM; \n4) 2.2.1 Parts of publications were not the latest editions: ICS Guide to helicopter/ Ship Operations, Guidance Manual for Tanker structures; \n5) 7.1.3 Parts of bal']}, {'question_no': '9.30.1', 'repeat_count': 3, 'priority': 'LOW_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'Mooring line / wire / tail installation age', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following discrepancies were noted in the HVPQ: - Item 1.5.4.2: Lack of necessary information. - Item 1.5.5.1: Error in the due date recording, listed incorrectly as 03 April 2024 instead of 03 April 2022. - Item 1.5.11: Lack of necessary information. - Item 1.5.12: Lack of necessary information. - Item 1.5.16: An error discovered in the recording of the answer, which was incorrectly marked as NO instead of YES. - Item 1.5.17: Lack of necessary information. - Item 3.1.3: The number of rating', 'The following errors were noted in HVPQ uploaded on SIRE. - 3. 1. 1 - Minimum officers required as \nper MSMD were 6. - 9. 16. 10. 1 I 2 - Details not provided. - 9. 17. 1 - The cargo pump details of slop wing tanks were not recorded. - 9.30.1- Vessel was assigned IMO Type 2 I 3 chemical tanker notation. - 9.31. 10-IGS composition details not completed. - 10.1.4- The mooring winches were \nfitted with split drums. - 10.8. 1 - Manifold arrangement diagram not completed. - 10. 9. 1 - Last periodical', '(i) The HVPQ entries inside section Ref: 1.2.3.2/4 (CII-B by Class ABS)/ 1.5.14 (Condition of Class on LRIT conformance test)/ 1.5.5.2/\n9.30.17.2/ 9.30.1( IMO Type 2) / 10.9.1(Last Class annual examination) / 10.10.6/ 121.4 did not accurately reflect the information\nrelating to the ship at the time of inspection. The incorrect info and particulars were brought to the attention of the Master.\n(ii) The vessel designation as recorded as per IOPP cert ., and the name of vessel P&I club were not decl']}, {'question_no': '12.2.4', 'repeat_count': 3, 'priority': 'LOW_REPEAT', 'category': 'Ice / special operations / VOC', 'topic': 'Certificates / class survey dates / endorsements', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following discrepancies were noted in the HVPQ: - Item 1.5.4.2: Lack of necessary information. - Item 1.5.5.1: Error in the due date recording, listed incorrectly as 03 April 2024 instead of 03 April 2022. - Item 1.5.11: Lack of necessary information. - Item 1.5.12: Lack of necessary information. - Item 1.5.16: An error discovered in the recording of the answer, which was incorrectly marked as NO instead of YES. - Item 1.5.17: Lack of necessary information. - Item 3.1.3: The number of rating', 'The HVPQ submitted through the CVIQ was found to be not completed/wrongly filled/ not answered for the following sections as\nbelow: 1.5.5, 1.5.19, 1.8.1, 1.9.8, 2.2.1, 3.1.9, 4.1.1, 4.2.2, 10.1.3, 10.4.2, 10.6.3, 10.6.6, 10.6.7, 10.7.1, 10.8.1, 10.8.2-10.8.4, 11.1.8,\n11.1.9, 11.1.13, 11.2.2, 11.3.1, 11.3.4, 11.9.1, 12.1.7, 12.1.8, 12.1.14, 12.2.4, 12.4.1. For example: 1.9.8 Port State Control # 1 Date of\nlast Port State Control inspection and # 2 Port of last Port State Control inspection - were', 'Certain information in the HVPQ downloaded on 23 January 2026 was inaccurate or not updated in the following sections; \nQ. 1.4.3 (To be 02 Nov 2020) / \n2.1.5 (Issue dates of Safety Radio Certificate and Fitness Certificate) / \n3.1.3 / 3.1.9 / 4.1.1 (Model, software and firmware versions) / \n4.2.2 (Serial number, software and firmware versions) / \n6.1.1.4 / 6.1.14.2 (To be 5 bar) / \n7.1.3 (Date when tank was coated) / \n9.67.3 / 10.9.1 (Last annual test) / \n11.4.3 / 12.2.4.']}, {'question_no': '1.8.6', 'repeat_count': 3, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Tank coating / structural inspection dates and frequency', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following information was either not completed or was not accurately updated in the HVPQ dated 14-Sep-2025: (i) Date of last in water survey (1. 5. 5. 1 ): 16-Jan-2025; (ii) Assigned dead weight 4 1.e. 34,999 MT that was available was not included (1.8.6); (iii) The following ballast tank inspection dates were not updated (7.1.3): 1P on 08-Jul-2025, 2P on 08-Jul-2025, 3P on 08-Jul-2025, 5S on 7 7-Aug-2025; (iv) Accommodation ladder wire renewal date (10.10.2). 05-Jun-2025; (v) The renewal da', 'In the HVPQ dated 16 January 2025 uploaded in the system was observed: 8. Load Line Information : 1.8.6 Assigned dead weight 1 : 49999 MT, however on board was noted the correct one : 49795.7 MT. 1.8.7 What is the current in use assigned dwt: it was stated: 497796.00 MT. The correct is 49795.7 MT.', 'On the uploaded HVPQ the following items were not correct: 1.8.6 (LL #6 -33497 MT), 1.8.7 (Active DW was 39996 MT), 1.9.8.2 (Last PSC was at Dumaguete, Philippines), 6.1.14.2 (Bunker lines tested to 7.0 Bar), 9.3.7 (Total capacity of cargo tanks value was not correct), 10.1.4 (Vessel was provided with split drums).']}, {'question_no': '9.16.19', 'repeat_count': 3, 'priority': 'LOW_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'Venting / P-V valves / IGS', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['A few errors were noted in the HVPQ. The correct data was:\n1.5.4 Date of last drydock was 12 June 2024.\n1.5.6 Date of last special survey was 12 June 2024.\n9.16.19 The nitrogen system had one segregation.', 'A few errors were noted in the HVPQ. The correct answers were: 1.5.4.3 Next drydock was due on 05.02.2027. 1.5.5.2 Next IWS due on 05.02.2027. 1.5.6 Last special survey was an enhanced special survey. 9.16.19 The IG system had only 1 segregation.', 'It was observed that the HVPQ uploaded by the Operator on 09 Dec 2024 had significant errors at various locations including a)\n6.1.1- Various coaming heights , b) 9.5.3- Loading through pump cannot be bypassed, c) 9.7.3- All valves could be operated from\nthe CCR and e) 9.16.19- Vessel had seven IG segregations']}, {'question_no': '10.1.8', 'repeat_count': 3, 'priority': 'LOW_REPEAT', 'category': 'Mooring / lifting / SPM / ETA', 'topic': 'Mooring line / wire / tail installation age', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['Following were noted in HVPQ last updated 07 Feb 2025: - Entries missing in 1.5.14 & 10.1.8. Incorrect entries in 9.30.6, 10.6.3 & 10.9.1', 'Process - The below information was found missing from the HVPQ uploaded on the CVIQ (dated 06 Mar 2025), as follows: 10.7.1 Bow Mooring arrangement diagram 10.1.8 Manifold Arrangement Diagram 10.2.1 Diagram for the layout of Mooring Fairleads, Chocks and Bollards and Bitts. A revised uploaded HVPQ (dated 04 Apr 2025) containing the missing diagrams was presented.', 'The following items of the HVPQ dated 30 November 2024 were not updated : 1.3.1.5 (IMO of Owner); 1.3.1.10(Date when registered); 1.5.10(Date of last thickness measurements); 1.6.13-15, 1.6.17-1.6.19, 1.6.21-25, 1.6.28 (Distances and parallel body); 4.2.2(Communication equipment on board);9.16.18, 9.16.21 (Type of the deck seal and non-return valve);9.17.1 (Details of cargo pump); 10.1.8(Retirement policy); 10.2.2-3 (Details of bollards and fairleads); 10.4.2 ( Details of ETA); 10.8.1-4 (Manifol']}, {'question_no': '11.1.9', 'repeat_count': 3, 'priority': 'LOW_REPEAT', 'category': 'Engine room / machinery / generators', 'topic': 'Mooring line / wire / tail installation age', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ["The HVPQ data entered either in error or not populated for the following HVPQ numbers were noted:\n9.1.1 tank plan incomplete\n9.2.1 1 & 2 not populated\n9.15.1 heat exchanger internal, heating coils were fitted.\n10.1.6.4 All mooring ropes had indicator strands\n10.9.4 not populated\n11.1.9 not populated\n11.3.1 A/E fuel HFO only. (The HVPQ only allows one selection) the A/E's can use either HFO or MDO.\nRECTIFIED, where possible, during the inspection.", 'HVPQ submitted through the CVIQ was found to be not completed/ wrongly filled/ not answered for the following sections as below: 1.5.1,1.5.5,1.8.7, 4.1.1, 5.3.2, 7.1.6, 9.59.7, 10.2.1, 10.4.1, 10.7.1, 10.8.1, 10.10.7, 11.1.9, 11.2.2, 11.9.3, 12.4.1. For example, 1.5.1 Classification society #2 is Classification society is an IACS member, # does the ship have dual class ? 9.59.7 Compressors # Are they oil free ? There were no responses provided. 4.1.1 Navigational equipment fitted on board? # Glo', 'The HVPQ submitted through the CVIQ was found to be not completed/wrongly filled/ not answered for the following sections as\nbelow: 1.5.5, 1.5.19, 1.8.1, 1.9.8, 2.2.1, 3.1.9, 4.1.1, 4.2.2, 10.1.3, 10.4.2, 10.6.3, 10.6.6, 10.6.7, 10.7.1, 10.8.1, 10.8.2-10.8.4, 11.1.8,\n11.1.9, 11.1.13, 11.2.2, 11.3.1, 11.3.4, 11.9.1, 12.1.7, 12.1.8, 12.1.14, 12.2.4, 12.4.1. For example: 1.9.8 Port State Control # 1 Date of\nlast Port State Control inspection and # 2 Port of last Port State Control inspection - were']}, {'question_no': '9.5.3', 'repeat_count': 3, 'priority': 'LOW_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['There were minor errors having no impact on the inspection. 7.1.1 Tank Type 2G is a gas tanker The tank construction is SS cladding, which is not an option in the HVPQ drop-down menu. 9.1.1 Tank plan cross-section schematic is wrong. 9.5.3 Says that the pump cannot be by pass pump during loading; this is incorrect. 12.1.13 The vessel has closed chocks and bollards at the manifold, the distances are missing.', 'It was observed that the HVPQ uploaded by the Operator on 09 Dec 2024 had significant errors at various locations including a)\n6.1.1- Various coaming heights , b) 9.5.3- Loading through pump cannot be bypassed, c) 9.7.3- All valves could be operated from\nthe CCR and e) 9.16.19- Vessel had seven IG segregations', 'Below sections of the HVPQ were not correctly updated: 9.5.3 :Cargo pumps were not provided with the external deck mounted heat exchangers. 10.1.7:Mooring ropes no. 3 end for end date was not updated to Jan 2025. 10.9.,1 : Details pertaining to engine overhead crane SWL 3.5 tons, provision crane port side SWL 5 tons and provision crane starboard side SWT 2.5 tons were not updated.']}, {'question_no': '6.2.1', 'repeat_count': 3, 'priority': 'LOW_REPEAT', 'category': 'Pollution prevention', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['On the uploaded HVPQ, some entries were either missing or not accurately answered: - Items 6.1.8, 9.6.6, 11.8.3, 12.1.3 & 12.1.4 had blank responses. - The date entered in the item 1.1.4.4 was one day ahead of the actual date of change of flag. - Item 1.2.3 stated that a CII rating of A was obtained and verified by the Class. Actually, the vessel was delivered on 18 September 2025 and was not due to receive its initial CII rating. - The numbers of certified officers and ratings recorded in the i', "On HVPQ, several items were recorded with inaccurate information: \n- The HVPQ declared a No to the item 1.9.5, however the vessel had three reported incidents on record in the previous 12 months. \n- The HVPQ declared a Yes to the items 6.2.1, 6.2.2 & 6.2.3, but the vessel didn't hold a valid USCG VRP Approval Letter or COFR. \n- The HVPQ declared a Yes to the item 9.7.3, actually all manifold valves were manual. \n- The safe working load (SWL) of each provision crane and engine room crane was ente", 'On the latest published HVPQ dated 14 November 2025, a few entries were observed with incorrect data: - The amount of value was entered as USD 2 billion, which was not the same as presented in the uploaded P&I Club Certificate of Entry (Qu 1.1.13.3). – The edition number recorded for two publications (i.e., ICS Guide to Helicopter/Ship Operations & OCIMF/ICS Ship to Ship Transfer Guide (Petroleum)) was not updated to reflect the current edition of these publications onboard (Qu 2.2.1). – The HVP']}, {'question_no': '1.9.2', 'repeat_count': 3, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Incident declarations', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following items mentioned in the HVPQ were not correct: 1.9. I; 1.9.2; I I. 9. I.', 'Vessel had suffered contact (collision) incident with another vessel at Tanjung Pelepas on 01 July 2025. Vessel had further carried out repairs using shore fitters inside steering gear room, Fields 1.9.2 and 1.9.3 in HVPQ for unscheduled repairs and collision incident respectively were mentioned as “ No”. Field 1.9.1 which should have been mentioned as “ No” was reported as”Yes”.', "HVPQ 1.9.2 response was recorded as 'No', Vessel had conducted an unscheduled voyage repair from Yeosu to Busan for fresh water tank (S) bulkhead welding crack to aft peak tank which required attendance of shore repair personnel from 7 to 9 December 2024. Class Surveyor attendance on 9 December 2024 for verification of repairs was recorded."]}, {'question_no': '11.1.6', 'repeat_count': 3, 'priority': 'LOW_REPEAT', 'category': 'Engine room / machinery / generators', 'topic': 'Generators / emergency power', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ updated on 04 March 2025 was not accurately completed with respect to following items:\n1.5.9/1.5.10/2.2.1/5.3.2.10/8.2.3.3/9.6.6/9.8.14.3/9.15.1/9.30.6/10.4.2/10.9.1/11.1.6', 'HVPQ had some minor errors as follows: (1.1.7/3) Ship\'s email entered was ineffective as senders\' emails may be rejected & not received by vessel. (Only by adding "master." was ship receiving all emails. For information, using email given in HVPQ, Inspector was not able to contact ship for pre-boarding formalities); (3.2.1) Incorrectly entered YES when senior officers in actuality, did not return to same ship on rotational basis; (5.3.1.4) Foam supplied or tested date was entered as 11 November ', 'The following errors/omissions were evident in the HVPQ:\nIn Q11.1.6 it was indicated that the type of fuel used for main propulsion was HFO. The actual fuel used for main propulsion was\nVLSFO or MGO.\nIn Q 11.3.1 It was indicated that the vessel had one generator and that it utilized HFO. The vessel had three generators that utilized\nLSFO or MGO.']}, {'question_no': '9.7.3', 'repeat_count': 3, 'priority': 'LOW_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['Information in HVPQ was incorrect at 9.7.3 (all valves could not be controlled from the cargo control room), 10.1.4 (mooring winches were fitted with split drums and remote controls), 10.9.1 (date of last annual inspection of lifting appliances by class was 23 Sep 2024) and 11.3.1 (three diesel generators - each 600 kW rating were fitted).', "On HVPQ, several items were recorded with inaccurate information: \n- The HVPQ declared a No to the item 1.9.5, however the vessel had three reported incidents on record in the previous 12 months. \n- The HVPQ declared a Yes to the items 6.2.1, 6.2.2 & 6.2.3, but the vessel didn't hold a valid USCG VRP Approval Letter or COFR. \n- The HVPQ declared a Yes to the item 9.7.3, actually all manifold valves were manual. \n- The safe working load (SWL) of each provision crane and engine room crane was ente", 'It was observed that the HVPQ uploaded by the Operator on 09 Dec 2024 had significant errors at various locations including a)\n6.1.1- Various coaming heights , b) 9.5.3- Loading through pump cannot be bypassed, c) 9.7.3- All valves could be operated from\nthe CCR and e) 9.16.19- Vessel had seven IG segregations']}, {'question_no': '11.11.2', 'repeat_count': 3, 'priority': 'LOW_REPEAT', 'category': 'Engine room / machinery / generators', 'topic': 'Engine room / machinery / generators', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['As per HVPQ section 11.11.2, it was declared that the vessel was fitted with exhaust gas recirculation system however, such a system was not fitted.', 'Inspector Negative Comment As per HVPQ section 11.11.2, it was declared that the vessel was fitted with exhaust gas recirculation system however, such a system was not fitted.', 'The following sections were incorrect or missing in the HVPQ 1.5.11,1.6.25,7.1.1,7.1.3,10.1.3.2, 10.9.1, 11.3.1, 11.5.1,11.11.2.']}, {'question_no': '1.9.6', 'repeat_count': 3, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Incident declarations', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['According to attached HVPQ, following items were kept in blank or wrongly entered: 1.3.1.5/1.4.2/1.9.6/10.1.3.2', 'The uploaded HVPQ had the following incorrect information. 1.1.8, 1.9.5 and 1.9.6.', 'The information related to the following HVPQ numbers were incorrect or missing in the uploaded HVPQ; 1.1.8, 1.3.1.5, 1.3.1.7, 1.3.1.8, 1.5.11, 1.5.18, 1.5.19, 1.9.5, 1.9.6, 5.3.1.4, 7.1.3, 9.1.1, 10.1.3.2, 10.1.4, 10.7.1 and 10.9.1.']}, {'question_no': '9.6.2', 'repeat_count': 3, 'priority': 'LOW_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'Tank coating / structural inspection dates and frequency', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following entries were either incomplete or incorrect - 1.5.1, 1.5.12, 6.1.5, 9.6.2, 9.6.6, 9.8.17, 9.15.1, 9.16.11, 9.16.12, 10.9.1, 11.3.4, 11.4.3, 11.5.5.2, 9.10.9.2 to 4.', 'The HVPQ uploaded had the following discrepancies: 1.2.4 - EIV rating stated available in 2024, not entered. 1.5.14 - Open conditions of class not entered 7.1.1 - Cargo tank inspections frequency stated as annual, last inspection on 11 & 12 May 2024 except for cargo tanks 4S & 5W. 7.1.3 - Ballast tank inspections frequency stated as annual, apart from forepeak tank, all others were last inspected on 18, 19 & 25 May 2024. 9.6.2, 9.6.6 & 9.17.1- Cargo pump types stated "centrifugal", whereas deep ', 'Some missing or erroneous information was noted within the HVPQ under the following sections: 8.2.1, 9.3.3-4, 9.3.7, 9.6.2, 9.10.10, 9.11.4, 9.15.1, 9.17.1, 9.35.3, 10.1.4, 10.8.1, 12.1.8.']}, {'question_no': '7.1.6', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Structural assessment / tank coating', 'topic': 'Tank coating / structural inspection dates and frequency', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ published to the OCIMF website on 24 Sept 2024 was randomly reviewed on 25 Sept 2024. Some questions were not responded and some questions were wrongly responded. For eg - 1.1.1, 1.2.4, 1.5.4, 1.9.8, 2.2.1, 5.3.2, 6.1.12, 7.1.6, 9.59.5, 9.68.6, 10.1.3, 10.4.2, 10.10.6, 11.10.1, 11.10.2, 12.4.1', 'HVPQ submitted through the CVIQ was found to be not completed/ wrongly filled/ not answered for the following sections as below: 1.5.1,1.5.5,1.8.7, 4.1.1, 5.3.2, 7.1.6, 9.59.7, 10.2.1, 10.4.1, 10.7.1, 10.8.1, 10.10.7, 11.1.9, 11.2.2, 11.9.3, 12.4.1. For example, 1.5.1 Classification society #2 is Classification society is an IACS member, # does the ship have dual class ? 9.59.7 Compressors # Are they oil free ? There were no responses provided. 4.1.1 Navigational equipment fitted on board? # Glo']}, {'question_no': '1.4', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Mooring brake test date / brake holding capacity', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following error was recorded in the latest HVPQ dated 02 January 2025: HVPQ 10. 1.4 last brake holding capacity test dated 10 January 2024 instead of 07 January 2025.', 'HVPQ submitted through the CVIQ was found to be not completed, wrongly filled, not answered for the following sections as below: 1.1. 13, 1 .2.1, 1.2.4, 1.3.1, 1.3.2, 2. 1.4, 3.3.1, 4. 1. 1, 7. 1 .1, 7. 1.3, 9. 1 .1, 9.8.22, 9. 16.6, 9.17.1, 9.34.7, 10.1.3, 10. 1.4, 10.2.1, 10.4.2, 10.7.1, 10.8.1, 12. 1.14. For example: 1.1. 13 P and I Club# 3 Amount of P&I Cover; 9.16.6 What is the capacity of the IGS?-not responded; 1.3. 1 Registered Owner# 9 Number of years this ship has been owned by Registe']}, {'question_no': '1.1', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Venting / P-V valves / IGS', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['HVPQ submitted through the CVIQ was found to be not completed, wrongly filled, not answered for the following sections as below: 1.1. 13, 1 .2.1, 1.2.4, 1.3.1, 1.3.2, 2. 1.4, 3.3.1, 4. 1. 1, 7. 1 .1, 7. 1.3, 9. 1 .1, 9.8.22, 9. 16.6, 9.17.1, 9.34.7, 10.1.3, 10. 1.4, 10.2.1, 10.4.2, 10.7.1, 10.8.1, 12. 1.14. For example: 1.1. 13 P and I Club# 3 Amount of P&I Cover; 9.16.6 What is the capacity of the IGS?-not responded; 1.3. 1 Registered Owner# 9 Number of years this ship has been owned by Registe', 'It was noted in the pre-board review of the HVPQ uploaded that the annual survey date was incorrectly uploaded as 23 Dec. 2024, but the last annual survey was completed on 29 Sep. 2025. \nAlso, the bunker tanks were protected by the double hull construction required by MARPOL Annex I, Reg. 12A, and it was correctly identified on the IOPPC, Form B, 2A.1.1, but the HVPQ 6.1.19 ‘Are all oil fuel tanks protected by a double hull?" was noted as NO.']}, {'question_no': '9.34.7', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'Venting / P-V valves / IGS', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['HVPQ submitted through the CVIQ was found to be not completed, wrongly filled, not answered for the following sections as below: 1.1. 13, 1 .2.1, 1.2.4, 1.3.1, 1.3.2, 2. 1.4, 3.3.1, 4. 1. 1, 7. 1 .1, 7. 1.3, 9. 1 .1, 9.8.22, 9. 16.6, 9.17.1, 9.34.7, 10.1.3, 10. 1.4, 10.2.1, 10.4.2, 10.7.1, 10.8.1, 12. 1.14. For example: 1.1. 13 P and I Club# 3 Amount of P&I Cover; 9.16.6 What is the capacity of the IGS?-not responded; 1.3. 1 Registered Owner# 9 Number of years this ship has been owned by Registe', 'The following discrepancies were noted in the HVPQ document: - Item 1.3.1.6: Lack of the necessary information. - Item 1.5.10: Error in the recording of the date, listed incorrectly as 28 September 2022 instead of 19 September 2022.. - Item 1.5.12: Lack of the necessary information. - Item 3.1.1: Error in the recording of the quantity, listed incorrectly as 07 instead of 06. - Item 3.1.9: Error in the recording of the quantity, listed incorrectly as 08 instead of 07. - Item 9.10.4: Error in the ']}, {'question_no': '6.1.13', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Pollution prevention', 'topic': 'Pollution prevention', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following discrepancies were noted in the HVPQ: - Item 1.5.4.2: Lack of necessary information. - Item 1.5.5.1: Error in the due date recording, listed incorrectly as 03 April 2024 instead of 03 April 2022. - Item 1.5.11: Lack of necessary information. - Item 1.5.12: Lack of necessary information. - Item 1.5.16: An error discovered in the recording of the answer, which was incorrectly marked as NO instead of YES. - Item 1.5.17: Lack of necessary information. - Item 3.1.3: The number of rating', 'Reference HVPQ 6.1.13 specified, Cargo lines were subjected to 24 bars pressure at intervals not greater than 12 months, onboard records indicated the annual pressure test was done to 16 bars as the MAWP was also 16 bars.']}, {'question_no': '10.9.4', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Mooring / lifting / SPM / ETA', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following HVPQ Entries were not updated – The date of last winch test (10.1.4): 05-Aug-2025. There were no restrictions noted for the hose handling crane’s capability to maintain it’s design SWL when plumbing a point one metre outboard from the ship’s side over the full length of the manifold (10.9.1). the vessel was capable of carrying out operations at SBM /CBM but the arrangement at the vapor manifold area did not have arrangements for securing floating hoses. (10.9.4)', "The HVPQ data entered either in error or not populated for the following HVPQ numbers were noted:\n9.1.1 tank plan incomplete\n9.2.1 1 & 2 not populated\n9.15.1 heat exchanger internal, heating coils were fitted.\n10.1.6.4 All mooring ropes had indicator strands\n10.9.4 not populated\n11.1.9 not populated\n11.3.1 A/E fuel HFO only. (The HVPQ only allows one selection) the A/E's can use either HFO or MDO.\nRECTIFIED, where possible, during the inspection."]}, {'question_no': '1.9', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'PSC inspection / detention / deficiency declaration', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['On the uploaded HVPQ the following items were recorded with inaccurate information: 1.9.(5,6) (LTI on 04-Jun-2025), 1.9.8 (Last PSC inspection was at Nha Be, Vietnam on 24-Jun-2025), 10.9.1 (Last 5 yearly test was on 04-Jun-2025).', 'The following items mentioned in the HVPQ were not correct: 1.9. I; 1.9.2; I I. 9. I.']}, {'question_no': '10.10.2', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Mooring / lifting / SPM / ETA', 'topic': 'Tank coating / structural inspection dates and frequency', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following information was either not completed or was not accurately updated in the HVPQ dated 14-Sep-2025: (i) Date of last in water survey (1. 5. 5. 1 ): 16-Jan-2025; (ii) Assigned dead weight 4 1.e. 34,999 MT that was available was not included (1.8.6); (iii) The following ballast tank inspection dates were not updated (7.1.3): 1P on 08-Jul-2025, 2P on 08-Jul-2025, 3P on 08-Jul-2025, 5S on 7 7-Aug-2025; (iv) Accommodation ladder wire renewal date (10.10.2). 05-Jun-2025; (v) The renewal da', 'The following items of the HVPQ dated 30 November 2024 were not updated : 1.3.1.5 (IMO of Owner); 1.3.1.10(Date when registered); 1.5.10(Date of last thickness measurements); 1.6.13-15, 1.6.17-1.6.19, 1.6.21-25, 1.6.28 (Distances and parallel body); 4.2.2(Communication equipment on board);9.16.18, 9.16.21 (Type of the deck seal and non-return valve);9.17.1 (Details of cargo pump); 10.1.8(Retirement policy); 10.2.2-3 (Details of bollards and fairleads); 10.4.2 ( Details of ETA); 10.8.1-4 (Manifol']}, {'question_no': '9.30.6', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'Cargo systems / IGS / venting / pumps', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['Following were noted in HVPQ last updated 07 Feb 2025: - Entries missing in 1.5.14 & 10.1.8. Incorrect entries in 9.30.6, 10.6.3 & 10.9.1', 'The HVPQ updated on 04 March 2025 was not accurately completed with respect to following items:\n1.5.9/1.5.10/2.2.1/5.3.2.10/8.2.3.3/9.6.6/9.8.14.3/9.15.1/9.30.6/10.4.2/10.9.1/11.1.6']}, {'question_no': '9.2.1', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'Mooring line / wire / tail installation age', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ["The HVPQ data entered either in error or not populated for the following HVPQ numbers were noted:\n9.1.1 tank plan incomplete\n9.2.1 1 & 2 not populated\n9.15.1 heat exchanger internal, heating coils were fitted.\n10.1.6.4 All mooring ropes had indicator strands\n10.9.4 not populated\n11.1.9 not populated\n11.3.1 A/E fuel HFO only. (The HVPQ only allows one selection) the A/E's can use either HFO or MDO.\nRECTIFIED, where possible, during the inspection.", 'In HVPQ dated 14 Jan 2025, the following items were either missing or not provided correct information: 1.5.4.3 / 9.1.1 / 9.2.1 / 9.10.2 / 10.7.1 / 10.8.1']}, {'question_no': '10.1.6', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Mooring / lifting / SPM / ETA', 'topic': 'Mooring line / wire / tail installation age', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ["The HVPQ data entered either in error or not populated for the following HVPQ numbers were noted:\n9.1.1 tank plan incomplete\n9.2.1 1 & 2 not populated\n9.15.1 heat exchanger internal, heating coils were fitted.\n10.1.6.4 All mooring ropes had indicator strands\n10.9.4 not populated\n11.1.9 not populated\n11.3.1 A/E fuel HFO only. (The HVPQ only allows one selection) the A/E's can use either HFO or MDO.\nRECTIFIED, where possible, during the inspection.", 'Question no. 4.1.1, 10.1.3.2, 10.1.6, 10.2.1, 10.8, 10.10.6, 10.10.7 & 11.3 were incomplete or blank.']}, {'question_no': '9.32.2', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'Certificates / class survey dates / endorsements', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['(i) The HVPQ entries inside section Ref: 1.5.5.2/ 9.16.18/ 9.32.2(Dehumidifier not fitted)/ 11.3.1(KW incorrect) /11.7.1/ 11.7.4/ 12.1.4 did not accurately reflect the information relating to the ship at the time of inspection. The incorrect info & particulars were brought to the attention of Master.\n(ii) The vessel designation as recorded as per IOPP certificate was not declared inside the PIQ.', 'The HVPQ uploaded by the Operator on 12 Mar 2025 had errors including:\na) 6.1.1.4- Incorrect distance for coaming of height 260mm.\nb) 9.32.2.1- The vessel has a dehumidifier.\nc) 10.6.2- The vessel was fitted with only 1 bow stopper.']}, {'question_no': '10.6.2', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Mooring / lifting / SPM / ETA', 'topic': 'Mooring / lifting / SPM / ETA', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ uploaded by the Operator on 12 Mar 2025 had errors including:\na) 6.1.1.4- Incorrect distance for coaming of height 260mm.\nb) 9.32.2.1- The vessel has a dehumidifier.\nc) 10.6.2- The vessel was fitted with only 1 bow stopper.', 'The following deficiencies were listed in the ship\'s HVPQ (dated 30 October 2025): Item 1.5.5.1 IWS listed on 2 July 2024 (when SSH was also carried out as per item 1.5.6.1). Item 5.3.2.8 Fuel gas system fitted with fixed CO2 fixed fire fighting system (no Fuel Gas system fitted on this Crude oil/ Product carrier ship). Item 5.3.2.10 Hydraulic room fitted with fixed CO2 fixed fire fighting system (not correct). Item 9.8.13.2 Listed manufacturer of UTI tapes as "Hermetic" (not correct). Item 10.1']}, {'question_no': '10.4.1', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Mooring / lifting / SPM / ETA', 'topic': 'Certificates / class survey dates / endorsements', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['Review of the uploaded HVPQ dated 14 Apr 2025 indicated the following were not correctly updated. 1. HVPQ Q1.1.8 was not correctly updated to indicate the type of ship as an "Oil Tanker," as specified in the International Oil Pollution Prevention Certificate (IOPPC). 2. HVPQ Q10.4.1 was answered in negative and the question required to ignore the remainder of the section, but Item Q10.4.7 and Q10.5.3 was provided with information which generated a question in CVIQ for emergency towing equipment ', 'HVPQ submitted through the CVIQ was found to be not completed/ wrongly filled/ not answered for the following sections as below: 1.5.1,1.5.5,1.8.7, 4.1.1, 5.3.2, 7.1.6, 9.59.7, 10.2.1, 10.4.1, 10.7.1, 10.8.1, 10.10.7, 11.1.9, 11.2.2, 11.9.3, 12.4.1. For example, 1.5.1 Classification society #2 is Classification society is an IACS member, # does the ship have dual class ? 9.59.7 Compressors # Are they oil free ? There were no responses provided. 4.1.1 Navigational equipment fitted on board? # Glo']}, {'question_no': '1.4.2', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'General / certificates / PSC / ownership', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following entries were either incorrect or incomplete :- 1.2.4.3, 1.3.1.7 & 8, 1.4.2, 1.5.4.1, 1.5.6, 3.1.6.2, 3.2.1, 3.2.2, 6.1.8, 7.1.1, 9.15.1, 9.15.3.5, 9.16.10, 9.16.29.2, 10.1.7 and 12.1.9.', 'According to attached HVPQ, following items were kept in blank or wrongly entered: 1.3.1.5/1.4.2/1.9.6/10.1.3.2']}, {'question_no': '3.1.6', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Crew / training / operator assessments', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following entries were either incorrect or incomplete :- 1.2.4.3, 1.3.1.7 & 8, 1.4.2, 1.5.4.1, 1.5.6, 3.1.6.2, 3.2.1, 3.2.2, 6.1.8, 7.1.1, 9.15.1, 9.15.3.5, 9.16.10, 9.16.29.2, 10.1.7 and 12.1.9.', 'The HVPQ showed erroneous dates and data as follows:\n1.1.13.4: P&I cover include wreck removal no response\n1.5.5.2: incorrect date\n1.5.6.5: next special date no response\n1.9.8.1: PSC date incorrect\n3.1.6.1: should be yes\n3.1.6.2: should be no\n3.1.10: same as for officers no response needed\n4.1.1: type of navigation equipments not completed\n10.9.1: crane annual test date incorrect\n11.1.13.1: quick closing valves fitted\n11.1.13.2: no response\n11.4.3: no response\n11.10.1: no response\n12.1.1: should']}, {'question_no': '3.2.1', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Crew / training / operator assessments', 'topic': 'Crew / training / operator assessments', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following entries were either incorrect or incomplete :- 1.2.4.3, 1.3.1.7 & 8, 1.4.2, 1.5.4.1, 1.5.6, 3.1.6.2, 3.2.1, 3.2.2, 6.1.8, 7.1.1, 9.15.1, 9.15.3.5, 9.16.10, 9.16.29.2, 10.1.7 and 12.1.9.', 'HVPQ had some minor errors as follows: (1.1.7/3) Ship\'s email entered was ineffective as senders\' emails may be rejected & not received by vessel. (Only by adding "master." was ship receiving all emails. For information, using email given in HVPQ, Inspector was not able to contact ship for pre-boarding formalities); (3.2.1) Incorrectly entered YES when senior officers in actuality, did not return to same ship on rotational basis; (5.3.1.4) Foam supplied or tested date was entered as 11 November ']}, {'question_no': '9.15.3', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'Cargo systems / IGS / venting / pumps', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following entries were either incorrect or incomplete :- 1.2.4.3, 1.3.1.7 & 8, 1.4.2, 1.5.4.1, 1.5.6, 3.1.6.2, 3.2.1, 3.2.2, 6.1.8, 7.1.1, 9.15.1, 9.15.3.5, 9.16.10, 9.16.29.2, 10.1.7 and 12.1.9.', 'Some erroneous or missing information was noted within the HVPQ under the following sections: 5.3.1.4, 9.3.3, 9.3.3.4, 9.15.3.3-4, 9.15.4.4, 10.1.7, 10.8.1.']}, {'question_no': '1.5.1', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Certificates / class survey dates / endorsements', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['HVPQ submitted through the CVIQ was found to be not completed/ wrongly filled/ not answered for the following sections as below: 1.5.1,1.5.5,1.8.7, 4.1.1, 5.3.2, 7.1.6, 9.59.7, 10.2.1, 10.4.1, 10.7.1, 10.8.1, 10.10.7, 11.1.9, 11.2.2, 11.9.3, 12.4.1. For example, 1.5.1 Classification society #2 is Classification society is an IACS member, # does the ship have dual class ? 9.59.7 Compressors # Are they oil free ? There were no responses provided. 4.1.1 Navigational equipment fitted on board? # Glo', 'The following entries were either incomplete or incorrect - 1.5.1, 1.5.12, 6.1.5, 9.6.2, 9.6.6, 9.8.17, 9.15.1, 9.16.11, 9.16.12, 10.9.1, 11.3.4, 11.4.3, 11.5.5.2, 9.10.9.2 to 4.']}, {'question_no': '11.2.2', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Engine room / machinery / generators', 'topic': 'Engine room / machinery / generators', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['HVPQ submitted through the CVIQ was found to be not completed/ wrongly filled/ not answered for the following sections as below: 1.5.1,1.5.5,1.8.7, 4.1.1, 5.3.2, 7.1.6, 9.59.7, 10.2.1, 10.4.1, 10.7.1, 10.8.1, 10.10.7, 11.1.9, 11.2.2, 11.9.3, 12.4.1. For example, 1.5.1 Classification society #2 is Classification society is an IACS member, # does the ship have dual class ? 9.59.7 Compressors # Are they oil free ? There were no responses provided. 4.1.1 Navigational equipment fitted on board? # Glo', 'The HVPQ submitted through the CVIQ was found to be not completed/wrongly filled/ not answered for the following sections as\nbelow: 1.5.5, 1.5.19, 1.8.1, 1.9.8, 2.2.1, 3.1.9, 4.1.1, 4.2.2, 10.1.3, 10.4.2, 10.6.3, 10.6.6, 10.6.7, 10.7.1, 10.8.1, 10.8.2-10.8.4, 11.1.8,\n11.1.9, 11.1.13, 11.2.2, 11.3.1, 11.3.4, 11.9.1, 12.1.7, 12.1.8, 12.1.14, 12.2.4, 12.4.1. For example: 1.9.8 Port State Control # 1 Date of\nlast Port State Control inspection and # 2 Port of last Port State Control inspection - were']}, {'question_no': '1.5.3', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Certificates / class survey dates / endorsements', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['As per HVPQ point 1.4.7 Delivery date as recorder in Form A Q1.5.3 or Form B Q1.8.3 of the IOPPC 13 October 2022.\nAs per IOPPC Form B, point 1.8.3 Date of delivery: 07 October 2022.\nAs per Cargo Ship Safety Construction Certificate. Date of delivery: 13 October 2022.', 'The following discrepancies were found in the HVPQ document:\nItem 1.5.3.1: An error was discovered in the recording of the answer, which was incorrectly marked as NO instead of YES.\nItem 1.5.3.2: Lack of the necessary information.\nItem 1.5.3.3: Lack of the necessary information.\nItem 1.5.5.1: Lack of the necessary information.\nItem 1.5.5.2: Lack of the necessary information.\nItem 1.5.10: Error in recording the date, listed incorrectly 26 March 2024 instead 12 March 2024.\nItem 1.5.12: Lack of the']}, {'question_no': '8.2.3', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Cargo/ballast tank capacity', 'topic': 'Cargo/ballast tank capacity', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ updated on 04 March 2025 was not accurately completed with respect to following items:\n1.5.9/1.5.10/2.2.1/5.3.2.10/8.2.3.3/9.6.6/9.8.14.3/9.15.1/9.30.6/10.4.2/10.9.1/11.1.6', 'The following discrepancies were found in the HVPQ document:\nItem 1.5.3.1: An error was discovered in the recording of the answer, which was incorrectly marked as NO instead of YES.\nItem 1.5.3.2: Lack of the necessary information.\nItem 1.5.3.3: Lack of the necessary information.\nItem 1.5.5.1: Lack of the necessary information.\nItem 1.5.5.2: Lack of the necessary information.\nItem 1.5.10: Error in recording the date, listed incorrectly 26 March 2024 instead 12 March 2024.\nItem 1.5.12: Lack of the']}, {'question_no': '9.8.14', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'Cargo systems / IGS / venting / pumps', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ updated on 04 March 2025 was not accurately completed with respect to following items:\n1.5.9/1.5.10/2.2.1/5.3.2.10/8.2.3.3/9.6.6/9.8.14.3/9.15.1/9.30.6/10.4.2/10.9.1/11.1.6', "Following inaccuracies noted in HVPQ: 4.1.1 Engine order logger was fitted but was recorded as 'No'. 9.8.5 Fixed tank gauging calibration information was incorrect. Calibration was done by the shipyard. 9.8.14 Vapour lock calibration information was incorrect - it was done by the shipyard."]}, {'question_no': '3.3.4', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Crew / training / operator assessments', 'topic': 'Crew training / simulator courses', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The PIQ dated 26 March 2025 was observed not accurately completed with respect to following items:\n3.3.4 ( no evidence that Junior Engineer attended shore based simulator course according IMO 2.07).', 'The HVPQ had incorrect entries for the following sections: 10.9.1 (last annual and 5 yearly tests), 2.2.1 (old edition for publication), 3.3.4, 4.2.2 (survival craft radios not mentioned) & 5.1.1 (ISO mentioned instead of ISM).']}, {'question_no': '5.3.6', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Safety / firefighting / lifeboats', 'topic': 'Safety / firefighting / lifeboats', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The item 5.3.6 reported fixed with the sprinkler system, instead the cargo and bunker sample locker was fitted with a sea water spray fixed system.', 'The item 5.3.6 reported fixed with the sprinkler system, instead the cargo and bunker sample locker was fitted with a sea water spray fixed system.']}, {'question_no': '9.8.8', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'Tank coating / structural inspection dates and frequency', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['HVPQ had some minor errors as follows: (1.1.7/3) Ship\'s email entered was ineffective as senders\' emails may be rejected & not received by vessel. (Only by adding "master." was ship receiving all emails. For information, using email given in HVPQ, Inspector was not able to contact ship for pre-boarding formalities); (3.2.1) Incorrectly entered YES when senior officers in actuality, did not return to same ship on rotational basis; (5.3.1.4) Foam supplied or tested date was entered as 11 November ', "The following operator's comments in the HVPQ were not updated or omitted to the question:\na) Not updated: 7.1.3 - Date of last ballast tank inspection.\nb) Omitted to question: 9.8.8 - Bunker tank sounding pipe & 9.8.10 - Bunker tank high level alarm system"]}, {'question_no': '3.5', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Crew / training / operator assessments', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['HVPQ had some minor errors as follows: (1.1.7/3) Ship\'s email entered was ineffective as senders\' emails may be rejected & not received by vessel. (Only by adding "master." was ship receiving all emails. For information, using email given in HVPQ, Inspector was not able to contact ship for pre-boarding formalities); (3.2.1) Incorrectly entered YES when senior officers in actuality, did not return to same ship on rotational basis; (5.3.1.4) Foam supplied or tested date was entered as 11 November ', 'Below sections of the HVPQ were not correctly updated: 9.5.3 :Cargo pumps were not provided with the external deck mounted heat exchangers. 10.1.7:Mooring ropes no. 3 end for end date was not updated to Jan 2025. 10.9.,1 : Details pertaining to engine overhead crane SWL 3.5 tons, provision crane port side SWL 5 tons and provision crane starboard side SWT 2.5 tons were not updated.']}, {'question_no': '10.6.6', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Mooring / lifting / SPM / ETA', 'topic': 'Mooring / lifting / SPM / ETA', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ submitted through the CVIQ was found to be not completed/wrongly filled/ not answered for the following sections as\nbelow: 1.5.5, 1.5.19, 1.8.1, 1.9.8, 2.2.1, 3.1.9, 4.1.1, 4.2.2, 10.1.3, 10.4.2, 10.6.3, 10.6.6, 10.6.7, 10.7.1, 10.8.1, 10.8.2-10.8.4, 11.1.8,\n11.1.9, 11.1.13, 11.2.2, 11.3.1, 11.3.4, 11.9.1, 12.1.7, 12.1.8, 12.1.14, 12.2.4, 12.4.1. For example: 1.9.8 Port State Control # 1 Date of\nlast Port State Control inspection and # 2 Port of last Port State Control inspection - were', 'The following items were inaccurately declared: 1-9.16.5 Are fixed O2 alarms fitted in inert gas generating or storage spaces? : Yes, actually not fitted; 2-10.6.6 What is the distance between the bow fairlead and stopper/bracket? : 3200.00 Mtrs, actual 3.20 Mtrs.']}, {'question_no': '6.2.2', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Pollution prevention', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['Noted following HVPQ item were wrongly marked: Item 1.5.18 - Yes, but the vessel currently did not have any Class dispensation. Item 6.1.7&8 - Yes, but Cargo sea chest was not fitted onboard. Therefore invalid question 6.2.2 generated.', "On HVPQ, several items were recorded with inaccurate information: \n- The HVPQ declared a No to the item 1.9.5, however the vessel had three reported incidents on record in the previous 12 months. \n- The HVPQ declared a Yes to the items 6.2.1, 6.2.2 & 6.2.3, but the vessel didn't hold a valid USCG VRP Approval Letter or COFR. \n- The HVPQ declared a Yes to the item 9.7.3, actually all manifold valves were manual. \n- The safe working load (SWL) of each provision crane and engine room crane was ente"]}, {'question_no': '10.8', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Mooring / lifting / SPM / ETA', 'topic': 'Mooring line / wire / tail installation age', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['Question no. 4.1.1, 10.1.3.2, 10.1.6, 10.2.1, 10.8, 10.10.6, 10.10.7 & 11.3 were incomplete or blank.', 'The following errors were noted in HVPQ uploaded on SIRE. - 3. 1. 1 - Minimum officers required as \nper MSMD were 6. - 9. 16. 10. 1 I 2 - Details not provided. - 9. 17. 1 - The cargo pump details of slop wing tanks were not recorded. - 9.30.1- Vessel was assigned IMO Type 2 I 3 chemical tanker notation. - 9.31. 10-IGS composition details not completed. - 10.1.4- The mooring winches were \nfitted with split drums. - 10.8. 1 - Manifold arrangement diagram not completed. - 10. 9. 1 - Last periodical']}, {'question_no': '11.3', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Engine room / machinery / generators', 'topic': 'Mooring line / wire / tail installation age', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['Question no. 4.1.1, 10.1.3.2, 10.1.6, 10.2.1, 10.8, 10.10.6, 10.10.7 & 11.3 were incomplete or blank.', 'The following errors were noted in HVPQ uploaded on SIRE. - 3. 1. 1 - Minimum officers required as \nper MSMD were 6. - 9. 16. 10. 1 I 2 - Details not provided. - 9. 17. 1 - The cargo pump details of slop wing tanks were not recorded. - 9.30.1- Vessel was assigned IMO Type 2 I 3 chemical tanker notation. - 9.31. 10-IGS composition details not completed. - 10.1.4- The mooring winches were \nfitted with split drums. - 10.8. 1 - Manifold arrangement diagram not completed. - 10. 9. 1 - Last periodical']}, {'question_no': '1.6.13', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following items of the HVPQ dated 30 November 2024 were not updated : 1.3.1.5 (IMO of Owner); 1.3.1.10(Date when registered); 1.5.10(Date of last thickness measurements); 1.6.13-15, 1.6.17-1.6.19, 1.6.21-25, 1.6.28 (Distances and parallel body); 4.2.2(Communication equipment on board);9.16.18, 9.16.21 (Type of the deck seal and non-return valve);9.17.1 (Details of cargo pump); 10.1.8(Retirement policy); 10.2.2-3 (Details of bollards and fairleads); 10.4.2 ( Details of ETA); 10.8.1-4 (Manifol', 'Missing information in the following: 1.6.13/1.6.14/ 1.6.15/1.6.17/1.6.18/ 1.6.19/ 1.6.21/1.6.22/1.6.23/1.6.24. Question 11.1.7 no information provided reference another type of fuel used for propulsion. Question 10.9.1 Engine room crane on board is electrical type and reported like Hydraulic type.']}, {'question_no': '1.6.17', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following items of the HVPQ dated 30 November 2024 were not updated : 1.3.1.5 (IMO of Owner); 1.3.1.10(Date when registered); 1.5.10(Date of last thickness measurements); 1.6.13-15, 1.6.17-1.6.19, 1.6.21-25, 1.6.28 (Distances and parallel body); 4.2.2(Communication equipment on board);9.16.18, 9.16.21 (Type of the deck seal and non-return valve);9.17.1 (Details of cargo pump); 10.1.8(Retirement policy); 10.2.2-3 (Details of bollards and fairleads); 10.4.2 ( Details of ETA); 10.8.1-4 (Manifol', 'Missing information in the following: 1.6.13/1.6.14/ 1.6.15/1.6.17/1.6.18/ 1.6.19/ 1.6.21/1.6.22/1.6.23/1.6.24. Question 11.1.7 no information provided reference another type of fuel used for propulsion. Question 10.9.1 Engine room crane on board is electrical type and reported like Hydraulic type.']}, {'question_no': '1.6.19', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following items of the HVPQ dated 30 November 2024 were not updated : 1.3.1.5 (IMO of Owner); 1.3.1.10(Date when registered); 1.5.10(Date of last thickness measurements); 1.6.13-15, 1.6.17-1.6.19, 1.6.21-25, 1.6.28 (Distances and parallel body); 4.2.2(Communication equipment on board);9.16.18, 9.16.21 (Type of the deck seal and non-return valve);9.17.1 (Details of cargo pump); 10.1.8(Retirement policy); 10.2.2-3 (Details of bollards and fairleads); 10.4.2 ( Details of ETA); 10.8.1-4 (Manifol', 'Missing information in the following: 1.6.13/1.6.14/ 1.6.15/1.6.17/1.6.18/ 1.6.19/ 1.6.21/1.6.22/1.6.23/1.6.24. Question 11.1.7 no information provided reference another type of fuel used for propulsion. Question 10.9.1 Engine room crane on board is electrical type and reported like Hydraulic type.']}, {'question_no': '1.6.21', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The following items of the HVPQ dated 30 November 2024 were not updated : 1.3.1.5 (IMO of Owner); 1.3.1.10(Date when registered); 1.5.10(Date of last thickness measurements); 1.6.13-15, 1.6.17-1.6.19, 1.6.21-25, 1.6.28 (Distances and parallel body); 4.2.2(Communication equipment on board);9.16.18, 9.16.21 (Type of the deck seal and non-return valve);9.17.1 (Details of cargo pump); 10.1.8(Retirement policy); 10.2.2-3 (Details of bollards and fairleads); 10.4.2 ( Details of ETA); 10.8.1-4 (Manifol', 'Missing information in the following: 1.6.13/1.6.14/ 1.6.15/1.6.17/1.6.18/ 1.6.19/ 1.6.21/1.6.22/1.6.23/1.6.24. Question 11.1.7 no information provided reference another type of fuel used for propulsion. Question 10.9.1 Engine room crane on board is electrical type and reported like Hydraulic type.']}, {'question_no': '9.15.4', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'Cargo systems / IGS / venting / pumps', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['On the uploaded HVPQ dated 4 March 2025, some entries were either missing or not correctly answered: - Items 9.15.4, 10.1.3.2, 10.2.1, 10.5.4, 10.7.1, 10.8.1 & 11.10 had blank responses. - In item 7.1.1, the frequency of inspection for cargo tanks was recorded as 30 months instead of every 12 months.', 'Some erroneous or missing information was noted within the HVPQ under the following sections: 5.3.1.4, 9.3.3, 9.3.3.4, 9.15.3.3-4, 9.15.4.4, 10.1.7, 10.8.1.']}, {'question_no': '1.9.1', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'General / certificates / PSC / ownership', 'topic': 'Tank coating / structural inspection dates and frequency', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['Vessel had suffered contact (collision) incident with another vessel at Tanjung Pelepas on 01 July 2025. Vessel had further carried out repairs using shore fitters inside steering gear room, Fields 1.9.2 and 1.9.3 in HVPQ for unscheduled repairs and collision incident respectively were mentioned as “ No”. Field 1.9.1 which should have been mentioned as “ No” was reported as”Yes”.', 'Question 1.9.1 of the online HVPQ did not include unscheduled repairs carried out in June 2024.\nQuestion 7.1.1 in the HVPQ had cargo tank inspection dates from 2023 and not the latest tank inspections carried out on October\n24.\nQuestion 7.1.3 did not contain the latest tank inspection reports carried out over various dates towards the end of 2024.']}, {'question_no': '6.60', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Pollution prevention', 'topic': 'Mooring line / wire / tail installation age', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['PROCESS - The following was noted not updated accurately as per the HVPQ dated 26-Nov-2024: a) 2.2.1 - The edition numbers for the publications were not provided as required. b) 6.1.14 - The bunker pipeline annual pressure test was incorrectly stated as 6.60 bars instead of 4.40 bars. c) 7.1.3 - The FPT, 1W and 3W ballast tanks annual inspection dates were not updated and indicated as overdue instead. d) 10.1.3.2 - The diagram for the mooring winch layout was not provided. e) 10.7.1 - The bow mo', 'The HVPQ was last updated by the operator on 27th March 2025 and the followings were noted - Items 6.1.14: The bunker pipeline annual pressure test was incorrectly stated as 6.60 bars, instead of 4.40 bars. - Item 10.1.3.1: The diagram for the mooring winch layout was not provided. Item 12.1.8 - details for the bollards, chock, rollers and fairleads used for the STS operation indicated that the winches / capstan driven by the electric power, instead of hydraulic power.']}, {'question_no': '4.40', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Navigation and communication', 'topic': 'Mooring line / wire / tail installation age', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['PROCESS - The following was noted not updated accurately as per the HVPQ dated 26-Nov-2024: a) 2.2.1 - The edition numbers for the publications were not provided as required. b) 6.1.14 - The bunker pipeline annual pressure test was incorrectly stated as 6.60 bars instead of 4.40 bars. c) 7.1.3 - The FPT, 1W and 3W ballast tanks annual inspection dates were not updated and indicated as overdue instead. d) 10.1.3.2 - The diagram for the mooring winch layout was not provided. e) 10.7.1 - The bow mo', 'The HVPQ was last updated by the operator on 27th March 2025 and the followings were noted - Items 6.1.14: The bunker pipeline annual pressure test was incorrectly stated as 6.60 bars, instead of 4.40 bars. - Item 10.1.3.1: The diagram for the mooring winch layout was not provided. Item 12.1.8 - details for the bollards, chock, rollers and fairleads used for the STS operation indicated that the winches / capstan driven by the electric power, instead of hydraulic power.']}, {'question_no': '5.1.1', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Safety / firefighting / lifeboats', 'topic': 'Safety / firefighting / lifeboats', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ["Inaccurate information found in HVPQ para 10.9.1 regarding last cargo gear annual survey (24 January 2025), para 5.1.1.4\nregarding approval of the ship's quality management (17 July 2022) and para 1.5.19 regarding flag dispensation which is not\nexisting", 'The HVPQ had incorrect entries for the following sections: 10.9.1 (last annual and 5 yearly tests), 2.2.1 (old edition for publication), 3.3.4, 4.2.2 (survival craft radios not mentioned) & 5.1.1 (ISO mentioned instead of ISM).']}, {'question_no': '9.3.7', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'PSC inspection / detention / deficiency declaration', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['On the uploaded HVPQ the following items were not correct: 1.8.6 (LL #6 -33497 MT), 1.8.7 (Active DW was 39996 MT), 1.9.8.2 (Last PSC was at Dumaguete, Philippines), 6.1.14.2 (Bunker lines tested to 7.0 Bar), 9.3.7 (Total capacity of cargo tanks value was not correct), 10.1.4 (Vessel was provided with split drums).', 'Some missing or erroneous information was noted within the HVPQ under the following sections: 8.2.1, 9.3.3-4, 9.3.7, 9.6.2, 9.10.10, 9.11.4, 9.15.1, 9.17.1, 9.35.3, 10.1.4, 10.8.1, 12.1.8.']}, {'question_no': '9.16.1', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'Mooring brake test date / brake holding capacity', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['The HVPQ uploaded on 23 January 2026 indicated wrong or not updated information on; 9.16.1 Inert gas system. 10.9.1 Annual examination of the lifting equipment.', 'Erroneous entries on HVPQ 9.16.1 (IGS); 10.1.4 (Brake rendering load); 10.4.2 (Emergency towing equipment SWL).']}, {'question_no': '9.3.3', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'Cargo systems / IGS / venting / pumps', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['Some missing or erroneous information was noted within the HVPQ under the following sections: 8.2.1, 9.3.3-4, 9.3.7, 9.6.2, 9.10.10, 9.11.4, 9.15.1, 9.17.1, 9.35.3, 10.1.4, 10.8.1, 12.1.8.', 'Some erroneous or missing information was noted within the HVPQ under the following sections: 5.3.1.4, 9.3.3, 9.3.3.4, 9.15.3.3-4, 9.15.4.4, 10.1.7, 10.8.1.']}, {'question_no': '9.35.3', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'Mooring line / wire / tail installation age', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['Some missing or erroneous information was noted within the HVPQ under the following sections: 8.2.1, 9.3.3-4, 9.3.7, 9.6.2, 9.10.10, 9.11.4, 9.15.1, 9.17.1, 9.35.3, 10.1.4, 10.8.1, 12.1.8.', 'The uploaded HVPQ had the following discrepancies: 1. 5.4 and 1. 5. 5 last IWS was on 29th January 2024 and the last dry dock on 6th February 2024 just a week after 1.8.1 summer, winter and tropical details for free board, draft, dead weight and displacement were all stated as same 9.16.5 fixed 02 alarms in IG spaces stated yes 9.35.3 capacity of tank cleaning machine not stated .']}, {'question_no': '9.8.5', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Cargo systems / IGS / venting / pumps', 'topic': 'Cargo systems / IGS / venting / pumps', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['HVPQ question 1.3.3.4/1.3.3.5/1.3.3.6/1.3.3.7 Comercial operator information missing. Question 1.5.18 Dispensation of flag state in effect answered YES. 1.5.19 no information regarding the flag state dispensation Question 7.1.3 Ballast tanks coated answer NO instead of YES. Question 9.8.3 answer YES no local readouts sighted. Question 9.8.5 does not provide the name of the inspection company related to question 9.8.4. Question 10.2.1 No diagram layout for mooring fairleads, chocks and bitts disp', "Following inaccuracies noted in HVPQ: 4.1.1 Engine order logger was fitted but was recorded as 'No'. 9.8.5 Fixed tank gauging calibration information was incorrect. Calibration was done by the shipyard. 9.8.14 Vapour lock calibration information was incorrect - it was done by the shipyard."]}, {'question_no': '11.1.7', 'repeat_count': 2, 'priority': 'LOW_REPEAT', 'category': 'Engine room / machinery / generators', 'topic': 'Lifting appliances annual and five-year tests', 'machine_check_intent': 'Use as repeat-observation priority signal only. Do not create a defect unless actual extracted HVPQ/PIQ/Q88/Class data is missing, stale, contradictory, or illogical.', 'compare_scope': ['HVPQ', 'PIQ', 'Q88', 'CLASS'], 'evidence_examples': ['HVPQ question 1.3.3.4/1.3.3.5/1.3.3.6/1.3.3.7 Comercial operator information missing. Question 1.5.18 Dispensation of flag state in effect answered YES. 1.5.19 no information regarding the flag state dispensation Question 7.1.3 Ballast tanks coated answer NO instead of YES. Question 9.8.3 answer YES no local readouts sighted. Question 9.8.5 does not provide the name of the inspection company related to question 9.8.4. Question 10.2.1 No diagram layout for mooring fairleads, chocks and bitts disp', 'Missing information in the following: 1.6.13/1.6.14/ 1.6.15/1.6.17/1.6.18/ 1.6.19/ 1.6.21/1.6.22/1.6.23/1.6.24. Question 11.1.7 no information provided reference another type of fuel used for propulsion. Question 10.9.1 Engine room crane on board is electrical type and reported like Hydraulic type.']}], 'validation_rules': [{'rule_id': 'HVPQ-BLANK-001', 'source_scope': ['HVPQ'], 'question_refs': ['ALL'], 'category': 'Completeness', 'rule_type': 'blank_check', 'severity': 'MEDIUM', 'statement': 'No applicable HVPQ question should be blank.', 'machine_logic': 'Flag blank/NA unless parent response is No/not applicable or section is explicitly exempt.', 'evidence_required': 'Question text, parent answer, extracted answer.', 'action_if_fail': 'Update missing HVPQ response or mark as N/A with valid parent logic.', 'skip_when': ['Section 13 combination carriers', 'Section 9.6 LNG bunkers', 'chemical-only sections for pure oil ship']}, {'rule_id': 'HVPQ-CERT-001', 'source_scope': ['HVPQ', 'CLASS'], 'question_refs': ['2.1.5', '1.5.11', '1.5.12'], 'category': 'Certificates', 'rule_type': 'date_validity', 'severity': 'HIGH', 'statement': 'No certificate should be expired.', 'machine_logic': 'For each certificate expiry date, fail if expiry < reference date.', 'evidence_required': 'Certificate name, issue date, annual/intermediate endorsement, expiry.', 'action_if_fail': 'Renew certificate or correct HVPQ/Class entry.', 'skip_when': []}, {'rule_id': 'HVPQ-CERT-002', 'source_scope': ['HVPQ', 'CLASS'], 'question_refs': ['2.1.5', '1.5.11'], 'category': 'Certificates', 'rule_type': 'endorsement_interval', 'severity': 'HIGH', 'statement': 'Certificates issued more than 1 year ago require annual endorsement within 12 months and alignment with 1.5.11.', 'machine_logic': 'If issue_date + 12 months < reference_date, latest annual endorsement must be within 12 months and match HVPQ 1.5.11 where applicable.', 'evidence_required': 'Issue date, latest annual endorsement, HVPQ 1.5.11.', 'action_if_fail': 'Correct endorsement date or update HVPQ annual survey response.', 'skip_when': []}, {'rule_id': 'HVPQ-CERT-003', 'source_scope': ['HVPQ', 'CLASS'], 'question_refs': ['2.1.5', '1.5.12'], 'category': 'Certificates', 'rule_type': 'endorsement_interval', 'severity': 'HIGH', 'statement': 'Certificates issued more than 2.5 years ago require intermediate endorsement within 30 months and alignment with 1.5.12.', 'machine_logic': 'If issue_date + 30 months < reference_date, latest intermediate endorsement must be within 30 months and match HVPQ 1.5.12 where applicable.', 'evidence_required': 'Issue date, intermediate endorsement, HVPQ 1.5.12.', 'action_if_fail': 'Correct endorsement date or update HVPQ intermediate survey response.', 'skip_when': []}, {'rule_id': 'HVPQ-INS-001', 'source_scope': ['HVPQ'], 'question_refs': ['1.1.13.4'], 'category': 'Insurance', 'rule_type': 'expected_boolean', 'severity': 'HIGH', 'statement': 'P&I wreck removal cover must be Yes.', 'machine_logic': 'Normalize response; pass only if Yes.', 'evidence_required': 'HVPQ 1.1.13.4 answer and P&I evidence.', 'action_if_fail': 'Correct declaration or obtain P&I evidence.', 'skip_when': []}, {'rule_id': 'HVPQ-CAP-001', 'source_scope': ['HVPQ'], 'question_refs': ['1.4.7', '1.5.19'], 'category': 'Class / CAP', 'rule_type': 'conditional_required', 'severity': 'HIGH', 'statement': 'If vessel age is above 15 years, CAP rating is applicable and should be 1.', 'machine_logic': 'Compute age from delivery/build date; if >15 years, CAP response must exist and rating must equal 1.', 'evidence_required': 'Delivery/build date, CAP applicability, CAP rating.', 'action_if_fail': 'Update CAP declaration/rating or provide justification.', 'skip_when': []}, {'rule_id': 'HVPQ-CLASS-001', 'source_scope': ['HVPQ'], 'question_refs': ['1.5.1.2'], 'category': 'Class', 'rule_type': 'expected_boolean', 'severity': 'HIGH', 'statement': 'Class society IACS member response must be Yes.', 'machine_logic': 'Normalize response; pass only if Yes.', 'evidence_required': 'HVPQ 1.5.1.2 answer.', 'action_if_fail': 'Correct IACS response or class details.', 'skip_when': []}, {'rule_id': 'HVPQ-DD-001', 'source_scope': ['HVPQ', 'CLASS'], 'question_refs': ['1.5.4.1', '1.5.6.1'], 'category': 'Dry dock / survey', 'rule_type': 'date_alignment', 'severity': 'HIGH', 'statement': 'Last dry dock should not be older than 5 years and must match last special/dry dock response where applicable.', 'machine_logic': 'Fail if last dry dock date + 5 years < reference date; compare HVPQ 1.5.4.1 with 1.5.6.1/Class status date.', 'evidence_required': 'Last dry dock date, special survey date, Class status.', 'action_if_fail': 'Correct HVPQ/Class data or verify survey evidence.', 'skip_when': []}, {'rule_id': 'HVPQ-INC-001', 'source_scope': ['HVPQ'], 'question_refs': ['1.9.1', '1.9.2'], 'category': 'Incidents', 'rule_type': 'conditional_consistency', 'severity': 'MEDIUM', 'statement': 'If 1.9.1 is No, 1.9.2 cannot also be No/blank where follow-up details are required.', 'machine_logic': 'Check parent/child incident logic; blank/no contradiction must be flagged for manual review.', 'evidence_required': 'HVPQ 1.9.1 and 1.9.2.', 'action_if_fail': 'Clarify incident declaration logic.', 'skip_when': []}, {'rule_id': 'HVPQ-CREW-001', 'source_scope': ['HVPQ'], 'question_refs': ['3.2.1'], 'category': 'Crew / operator assessment', 'rule_type': 'expected_boolean', 'severity': 'MEDIUM', 'statement': 'Response for 3.2.1 should be No.', 'machine_logic': 'Normalize response; pass only if No.', 'evidence_required': 'HVPQ 3.2.1 answer.', 'action_if_fail': 'Correct answer or provide justification.', 'skip_when': []}, {'rule_id': 'HVPQ-TRAIN-001', 'source_scope': ['HVPQ', 'PIQ'], 'question_refs': ['3.3.4', 'PIQ 3.3.1', 'PIQ 3.3.3', 'PIQ 3.3.4'], 'category': 'Training', 'rule_type': 'keyword_presence', 'severity': 'MEDIUM', 'statement': 'HVPQ 3.3.4 courses should include ERM simulator, cargo simulator, and ship handling type courses and align with PIQ training entries.', 'machine_logic': 'Extract course text; require concepts engine-room resource management/simulator, cargo simulator, ship handling; compare with PIQ courses.', 'evidence_required': 'Course names in HVPQ and PIQ.', 'action_if_fail': 'Update course list or clarify equivalence.', 'skip_when': []}, {'rule_id': 'HVPQ-FIRE-001', 'source_scope': ['HVPQ'], 'question_refs': ['5.3.1.4'], 'category': 'Safety / firefighting', 'rule_type': 'date_validity', 'severity': 'MEDIUM', 'statement': '5.3.1.4 date should not be older than 12 months.', 'machine_logic': 'Fail if date + 12 months < reference date.', 'evidence_required': 'HVPQ 5.3.1.4 date.', 'action_if_fail': 'Update date/evidence.', 'skip_when': []}, {'rule_id': 'HVPQ-TANK-001', 'source_scope': ['HVPQ', 'PIQ'], 'question_refs': ['7.1.1', 'PIQ 2.3.3001'], 'category': 'Tank coating / cargo tanks', 'rule_type': 'date_frequency', 'severity': 'HIGH', 'statement': 'Cargo/slop tank coating inspection dates by ship staff must be within stated frequency and align with PIQ 2.3.3001.', 'machine_logic': 'Read table last inspection date column, not original coating date. Oldest last inspection + frequency months must be >= reference date. Compare PIQ oldest/frequency.', 'evidence_required': 'Frequency, all last inspection dates, oldest inspection date.', 'action_if_fail': 'Update actual coating inspection dates/frequency.', 'skip_when': []}, {'rule_id': 'HVPQ-TANK-002', 'source_scope': ['HVPQ', 'PIQ'], 'question_refs': ['7.1.3', 'PIQ 2.3.3002'], 'category': 'Tank coating / ballast tanks', 'rule_type': 'date_frequency', 'severity': 'HIGH', 'statement': 'Ballast tank coating inspections by competent person must be within stated frequency and align with PIQ 2.3.3002.', 'machine_logic': 'Read table last inspection date column, not coating application date. Oldest last inspection + frequency months must be >= reference date.', 'evidence_required': 'Frequency, all last inspection dates, oldest inspection date.', 'action_if_fail': 'Update actual coating inspection dates/frequency.', 'skip_when': []}, {'rule_id': 'HVPQ-TANK-003', 'source_scope': ['HVPQ'], 'question_refs': ['7.1.4.5'], 'category': 'Tank coating', 'rule_type': 'numeric_unit', 'severity': 'MEDIUM', 'statement': '7.1.4.5 response should be greater than zero and expressed in percent.', 'machine_logic': 'Parse numeric value and unit; fail if value <=0 or no % unit/percentage context.', 'evidence_required': 'HVPQ 7.1.4.5.', 'action_if_fail': 'Correct percentage value.', 'skip_when': []}, {'rule_id': 'HVPQ-PUMP-001', 'source_scope': ['HVPQ', 'Q88'], 'question_refs': ['1.6.1', '9.6.2', '9.17.1'], 'category': 'Cargo pumps', 'rule_type': 'conditional_value', 'severity': 'MEDIUM', 'statement': 'If cargo tank capacity/parameter in 1.6.1 is greater than 200, pump type should be Centrifugal; otherwise Deepwell.', 'machine_logic': 'Parse 1.6.1 numeric and pump type. If >200 require Centrifugal; else require Deepwell. Cross-check Q88 pump type.', 'evidence_required': 'HVPQ 1.6.1, HVPQ 9.6.2/9.17.1, Q88 pump table.', 'action_if_fail': 'Correct pump type or verify mapping.', 'skip_when': []}, {'rule_id': 'HVPQ-MOOR-001', 'source_scope': ['HVPQ', 'Q88'], 'question_refs': ['10.1.4'], 'category': 'Mooring', 'rule_type': 'date_frequency', 'severity': 'HIGH', 'statement': 'Date of last brake holding capacity test must be within brake test frequency.', 'machine_logic': 'Extract brake test date and frequency; fail if last test + frequency < reference date.', 'evidence_required': 'HVPQ/Q88 brake test date and frequency.', 'action_if_fail': 'Update brake test date or conduct test.', 'skip_when': []}, {'rule_id': 'HVPQ-MOOR-002', 'source_scope': ['HVPQ', 'Q88'], 'question_refs': ['10.1.7'], 'category': 'Mooring', 'rule_type': 'age_limit', 'severity': 'HIGH', 'statement': 'Wire, tails and ropes must be within allowed age: wire 10 years, tails 18 months, ropes 5 years.', 'machine_logic': 'For each mooring line row, parse type and installed date; fail if age exceeds type limit.', 'evidence_required': 'Line type and installed date for each wire/tail/rope.', 'action_if_fail': 'Replace overdue line or correct date/type.', 'skip_when': []}, {'rule_id': 'HVPQ-LIFT-001', 'source_scope': ['HVPQ', 'Q88'], 'question_refs': ['10.9.1'], 'category': 'Lifting appliances', 'rule_type': 'date_frequency', 'severity': 'HIGH', 'statement': 'Last annual and five-year lifting appliance tests must not be expired.', 'machine_logic': 'Annual test <=12 months; five-year test <=60 months from reference date.', 'evidence_required': 'Annual test date, 5-year test date.', 'action_if_fail': 'Update test details or arrange test.', 'skip_when': []}, {'rule_id': 'PIQ-TYPE-001', 'source_scope': ['PIQ', 'HVPQ'], 'question_refs': ['PIQ 1.1.1', 'HVPQ 2.1.4', 'HVPQ 1.1.8'], 'category': 'Vessel type', 'rule_type': 'cross_document_match', 'severity': 'MEDIUM', 'statement': 'PIQ vessel type must match HVPQ vessel type.', 'machine_logic': 'Normalize vessel type categories and compare.', 'evidence_required': 'PIQ 1.1.1, HVPQ vessel type.', 'action_if_fail': 'Correct one document or justify terminology difference.', 'skip_when': []}, {'rule_id': 'PIQ-NAV-001', 'source_scope': ['PIQ'], 'question_refs': ['PIQ 3.2.1', 'PIQ 3.2.2'], 'category': 'Navigation assessment', 'rule_type': 'mutual_exclusive_boolean', 'severity': 'MEDIUM', 'statement': 'Static and dynamic navigational assessments cannot both be Yes or both be No; responses must be opposite.', 'machine_logic': 'Normalize both answers; fail if equal or blank.', 'evidence_required': 'PIQ 3.2.1 and 3.2.2.', 'action_if_fail': 'Correct assessment type declaration.', 'skip_when': []}, {'rule_id': 'PIQ-NAV-002', 'source_scope': ['PIQ'], 'question_refs': ['PIQ 3.2.1', 'PIQ 3.2.2'], 'category': 'Navigation assessment', 'rule_type': 'date_validity', 'severity': 'HIGH', 'statement': 'Last navigational assessment date must not be older than 12 months.', 'machine_logic': 'Use whichever assessment is marked Yes; latest date + 12 months must be >= reference date.', 'evidence_required': 'Assessment type, date.', 'action_if_fail': 'Update assessment date/evidence.', 'skip_when': []}, {'rule_id': 'PIQ-AUDIT-001', 'source_scope': ['PIQ'], 'question_refs': ['PIQ 3.5', 'PIQ 3.6', 'PIQ 3.7'], 'category': 'Audits', 'rule_type': 'conditional_date_validity', 'severity': 'MEDIUM', 'statement': 'If Chapter 3.5, 3.6 or 3.7 is marked Yes, last audit date must not be older than 12 months.', 'machine_logic': 'For each Yes audit, parse date; fail if date +12 months < reference date.', 'evidence_required': 'Audit answer and date.', 'action_if_fail': 'Update audit evidence.', 'skip_when': []}, {'rule_id': 'PIQ-MAST-001', 'source_scope': ['PIQ'], 'question_refs': ['PIQ 3.4.2001', 'PIQ 3.4.2002'], 'category': 'Master review / office assessment', 'rule_type': 'date_validity', 'severity': 'MEDIUM', 'statement': 'PIQ 3.4.2001 date must be within 12 months and 3.4.2002 within 3 months.', 'machine_logic': 'Fail if 3.4.2001 +12 months or 3.4.2002 +3 months is before reference date.', 'evidence_required': 'PIQ dates.', 'action_if_fail': 'Update/revalidate dates.', 'skip_when': []}, {'rule_id': 'PIQ-FIRE-001', 'source_scope': ['PIQ', 'HVPQ'], 'question_refs': ['PIQ 5.2.4', 'HVPQ 5.3.2.4'], 'category': 'Safety / firefighting', 'rule_type': 'cross_document_match', 'severity': 'MEDIUM', 'statement': 'PIQ 5.2.4 must match HVPQ 5.3.2 sub question 4.', 'machine_logic': 'Normalize and compare answers.', 'evidence_required': 'PIQ and HVPQ relevant answers.', 'action_if_fail': 'Correct mismatch.', 'skip_when': []}, {'rule_id': 'PIQ-INC-001', 'source_scope': ['PIQ', 'HVPQ', 'CLASS'], 'question_refs': ['PIQ 5.7.1001-1029', 'HVPQ 1.9.1-1.9.7', 'PIQ 2.1.1'], 'category': 'Incidents / damage survey', 'rule_type': 'cross_document_alignment', 'severity': 'HIGH', 'statement': 'PIQ incident declarations must align with HVPQ incident responses, and damage/other/occasional Class visit purpose may imply damage/repair verification.', 'machine_logic': 'Compare PIQ incident yes/no matrix with HVPQ 1.9.1-1.9.7. If Class visit purpose includes damage/other/occasional, require incident/repair explanation.', 'evidence_required': 'PIQ incident matrix, HVPQ incidents, Class/PIQ survey purpose.', 'action_if_fail': 'Correct declarations or add explanation.', 'skip_when': []}, {'rule_id': 'PIQ-PSC-001', 'source_scope': ['PIQ', 'HVPQ'], 'question_refs': ['PIQ 2.8', 'HVPQ 1.9.8', 'HVPQ 1.9.9'], 'category': 'PSC', 'rule_type': 'cross_document_alignment', 'severity': 'HIGH', 'statement': 'PIQ PSC information must align with HVPQ PSC responses.', 'machine_logic': 'Compare last PSC date/port/detention/deficiency data.', 'evidence_required': 'PIQ Chapter 2.8 and HVPQ 1.9.8/1.9.9.', 'action_if_fail': 'Update PSC records.', 'skip_when': []}, {'rule_id': 'PIQ-VEC-001', 'source_scope': ['PIQ', 'HVPQ'], 'question_refs': ['PIQ 8.3', 'HVPQ 9.9', 'HVPQ 9.10'], 'category': 'VEC / venting', 'rule_type': 'cross_document_alignment', 'severity': 'MEDIUM', 'statement': 'PIQ Section 8.3 responses must align with HVPQ vapor emission control and venting sections.', 'machine_logic': 'Map VEC/venting fitted/operational/arrangement answers and compare.', 'evidence_required': 'PIQ 8.3, HVPQ 9.9/9.10.', 'action_if_fail': 'Correct inconsistent VEC/venting declarations.', 'skip_when': []}, {'rule_id': 'PIQ-GEN-001', 'source_scope': ['PIQ', 'HVPQ'], 'question_refs': ['PIQ 10.2.1', 'HVPQ 11.3.3'], 'category': 'Machinery / generators', 'rule_type': 'cross_document_match', 'severity': 'MEDIUM', 'statement': 'PIQ 10.2.1 must match HVPQ 11.3.3.', 'machine_logic': 'Normalize and compare answer/value.', 'evidence_required': 'PIQ 10.2.1 and HVPQ 11.3.3.', 'action_if_fail': 'Correct mismatch.', 'skip_when': []}, {'rule_id': 'PIQ-ENG-001', 'source_scope': ['PIQ', 'HVPQ'], 'question_refs': ['PIQ 10.2.3', 'HVPQ 11.9.1'], 'category': 'Machinery / alarms / safety', 'rule_type': 'cross_document_match', 'severity': 'MEDIUM', 'statement': 'PIQ 10.2.3 must match HVPQ 11.9.1.', 'machine_logic': 'Normalize and compare answer/value.', 'evidence_required': 'PIQ 10.2.3 and HVPQ 11.9.1.', 'action_if_fail': 'Correct mismatch.', 'skip_when': []}, {'rule_id': 'PIQ-SUPT-001', 'source_scope': ['PIQ'], 'question_refs': ['PIQ 2.2.1001'], 'category': 'Management oversight', 'rule_type': 'strict_interval', 'severity': 'HIGH', 'statement': 'Technical Superintendent visit must be within 7 months and successive visit gaps must not exceed 7.0 months.', 'machine_logic': 'Sort visit date ranges; latest visit end/date must be <=7.0 months old and every successive gap <=7.0 months. Strict no tolerance.', 'evidence_required': 'PIQ 2.2.1001 visit table dates.', 'action_if_fail': 'Arrange/update inspection or explain non-compliance.', 'skip_when': []}, {'rule_id': 'PIQ-SUPT-002', 'source_scope': ['PIQ'], 'question_refs': ['PIQ 2.2.1002'], 'category': 'Management oversight', 'rule_type': 'strict_interval', 'severity': 'HIGH', 'statement': 'Marine Superintendent visit must be within 12 months and successive visit gaps must not exceed 12.0 months.', 'machine_logic': 'Sort visit date ranges; latest visit end/date must be <=12.0 months old and every successive gap <=12.0 months. Strict no tolerance.', 'evidence_required': 'PIQ 2.2.1002 visit table dates.', 'action_if_fail': 'Arrange/update inspection or explain non-compliance.', 'skip_when': []}, {'rule_id': 'PIQ-TANK-001', 'source_scope': ['PIQ', 'HVPQ'], 'question_refs': ['PIQ 2.3.3001', 'HVPQ 7.1.1'], 'category': 'Structural assessment', 'rule_type': 'date_frequency_alignment', 'severity': 'HIGH', 'statement': 'PIQ cargo/slop tank inspection frequency and oldest date must align with HVPQ 7.1.1.', 'machine_logic': 'Compare PIQ required frequency and oldest inspection date against HVPQ tank coating last inspection dates and frequency.', 'evidence_required': 'PIQ 2.3.3001 and HVPQ 7.1.1 table.', 'action_if_fail': 'Correct stale/mismatched tank inspection data.', 'skip_when': []}, {'rule_id': 'PIQ-TANK-002', 'source_scope': ['PIQ', 'HVPQ'], 'question_refs': ['PIQ 2.3.3002', 'HVPQ 7.1.3'], 'category': 'Structural assessment', 'rule_type': 'date_frequency_alignment', 'severity': 'HIGH', 'statement': 'PIQ ballast tank inspection frequency and oldest date must align with HVPQ 7.1.3.', 'machine_logic': 'Compare PIQ required frequency and oldest inspection date against HVPQ tank coating last inspection dates and frequency.', 'evidence_required': 'PIQ 2.3.3002 and HVPQ 7.1.3 table.', 'action_if_fail': 'Correct stale/mismatched ballast tank inspection data.', 'skip_when': []}]}


def embedded_observation_df() -> pd.DataFrame:
    """Return observation history in the same row format expected by older helper functions.

    The app uses this only to prioritise/strengthen reasons for real extracted defects,
    not to create hundreds of generic manual findings.
    """
    rows = []
    for item in EMBEDDED_KNOWLEDGE_BASE.get("observation_library", []):
        q = clean_text(item.get("question_no", ""))
        ex = " | ".join([clean_text(x) for x in item.get("evidence_examples", [])])
        rows.append({
            "question_no": q,
            "repeat_count": item.get("repeat_count", 0),
            "priority": item.get("priority", ""),
            "category": item.get("category", ""),
            "topic": item.get("topic", ""),
            "joined": f"HVPQ {q} | repeat_count={item.get('repeat_count', 0)} | {item.get('topic', '')} | {ex}",
        })
    return pd.DataFrame(rows)


def embedded_validation_rules_df() -> pd.DataFrame:
    """Machine-readable validation rule register for display/export/debug."""
    rows = []
    for r in EMBEDDED_KNOWLEDGE_BASE.get("validation_rules", []):
        rows.append({
            "Rule ID": r.get("rule_id", ""),
            "Source scope": ", ".join(r.get("source_scope", [])),
            "Question refs": ", ".join(r.get("question_refs", [])),
            "Category": r.get("category", ""),
            "Rule type": r.get("rule_type", ""),
            "Severity": r.get("severity", ""),
            "Statement": r.get("statement", ""),
            "Machine logic": r.get("machine_logic", ""),
            "Skip when": "; ".join(r.get("skip_when", [])),
            "Action if fail": r.get("action_if_fail", ""),
        })
    return pd.DataFrame(rows)


def embedded_observation_rules_df() -> pd.DataFrame:
    rows = []
    for item in EMBEDDED_KNOWLEDGE_BASE.get("observation_library", []):
        rows.append({
            "Question No.": item.get("question_no", ""),
            "Repeat count": item.get("repeat_count", 0),
            "Priority": item.get("priority", ""),
            "Category": item.get("category", ""),
            "Topic": item.get("topic", ""),
            "Machine check intent": item.get("machine_check_intent", ""),
            "Example basis": " | ".join(item.get("evidence_examples", [])[:2]),
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(["Repeat count", "Question No."], ascending=[False, True])
    return df


RULE_QIDS = ["1.1.13.4", "1.5.1.2", "1.5.4.1", "1.5.6.1", "1.5.11", "1.5.12", "1.9.1", "1.9.2", "3.2.1", "3.3.4", "5.3.1.4", "7.1.1", "7.1.3", "7.1.4.5", "9.6.2", "10.1.4", "10.1.7", "10.9.1", "PIQ 1.1.1", "PIQ 3.2.1", "PIQ 3.2.2", "PIQ 3.2.5", "PIQ 3.2.6", "PIQ 3.2.7", "PIQ 2.2.1001", "PIQ 2.2.1002", "PIQ 2.3.3001", "PIQ 2.3.3002", "PIQ 2.8.2", "PIQ 5.7.1001-1029"]

def _normalise_qid_token(token: str) -> str:
    """Normalise HVPQ-like question numbers and reject dates/loose Q88 line numbers.

    The observation Excel often contains dates such as 05.02.2027 or Q88-style
    two-part numbers such as 1.11. These must not become HVPQ manual checks.
    For this checker we only treat three-part numeric references as HVPQ question
    references, e.g. 2.1.5, 7.1.3, 10.1.4, 2.2.1001.
    """
    t = str(token).strip().strip(".,;:()[]{}")
    parts = t.split(".")
    if len(parts) != 3:
        return ""
    if not all(p.isdigit() for p in parts):
        return ""
    a, b, c = [int(p) for p in parts]
    # HVPQ chapters are normally 1-13. This removes random values but keeps 10.1.4 etc.
    if not (1 <= a <= 13):
        return ""
    # Reject obvious dates in dd.mm.yyyy / mm.dd.yyyy form.
    if 1900 <= c <= 2100 and 1 <= a <= 31 and 1 <= b <= 12:
        return ""
    # Reject improbable month/year-looking values even with leading zero.
    if len(parts[2]) == 4 and c >= 1900:
        return ""
    return f"{a}.{b}.{c}"

def extract_qids_from_obs(obs_df: pd.DataFrame) -> set:
    qids=set()
    if obs_df is None or obs_df.empty:
        return qids
    try:
        joined = obs_df.fillna("").astype(str).apply(lambda r: " ".join(r.tolist()), axis=1)
    except Exception:
        return qids
    for s in joined:
        # Require three components. This intentionally ignores Q88-style 1.11 / 1.15
        # because the observation library is used only to prioritise HVPQ correction checks.
        for m in re.findall(r"\b\d{1,2}\.\d{1,2}\.\d{1,4}\b", str(s)):
            q=_normalise_qid_token(m)
            if q:
                qids.add(q)
    return qids

def obs_reason_for(qno: str, obs_qids: set) -> str:
    if not qno or not obs_qids:
        return ""
    qno_s=str(qno)
    tokens=set()
    for m in re.findall(r"\b\d{1,2}\.\d{1,2}\.\d{1,4}\b", qno_s):
        q=_normalise_qid_token(m)
        if q:
            tokens.add(q)
    if tokens.intersection(obs_qids):
        return "Similar issues have appeared in the uploaded observation history, so this HVPQ item is prioritised for correction/review."
    for q in obs_qids:
        for t in tokens:
            if t.startswith(q + ".") or q.startswith(t + "."):
                return "Similar issues have appeared in the uploaded observation history, so this HVPQ item is prioritised for correction/review."
    return ""

def concise_check_name(text_in: str) -> str:
    s=clean_text(text_in)
    s=re.sub(r"^(HVPQ|PIQ|Q88)\s+(missing|missing/not extracted)\s+", "", s, flags=re.I)
    s=re.sub(r"\b(completeness|value was extracted for review)\b", "", s, flags=re.I)
    s=re.sub(r"\s+", " ", s).strip(" -:")
    return s

def add_v15_validation_findings(findings: List[Finding], fields: List[FieldRecord], ref_date: date):
    """Add deterministic/manual validation rows from the user's comparison rules. Avoid false confidence: if not extractable, add manual confirmation."""
    hvpq_raw = "\n".join([f.raw for f in fields if f.source=="HVPQ"])

    pni_sec = first_field(fields,"HVPQ","section.1.1.13") or ""
    if pni_sec and re.search(r"wreck\s+removal[^\n]{0,80}\bno\b", pni_sec, re.I):
        add_finding(findings, area="Insurance", check="P&I wreck removal cover declaration", status="MISMATCH", risk="HIGH", hvpq_value=pni_sec[:350], reason="HVPQ 1.1.13.4 should confirm wreck removal cover. Extracted text suggests it may not be Yes.", action="Correct HVPQ or provide evidence that P&I cover includes wreck removal.")
    elif not pni_sec:
        add_finding(findings, area="Insurance", check="P&I wreck removal cover declaration", status="MANUAL CHECK", risk="MEDIUM", hvpq_value="Not reliably extracted", reason="HVPQ 1.1.13.4 could not be reliably checked. Rule requires the response to be Yes.", action="Vessel/office to confirm HVPQ 1.1.13.4 is Yes and supported by P&I evidence.")

    iacs_sec = section_text_by_qid(hvpq_raw, "1.5.1") or first_field(fields,"HVPQ","classification.class_society")
    if iacs_sec and re.search(r"IACS[^\n]{0,100}\bNo\b", iacs_sec, re.I):
        add_finding(findings, area="Class", check="IACS membership declaration", status="MISMATCH", risk="HIGH", hvpq_value=iacs_sec[:350], reason="HVPQ 1.5.1.2 should be Yes for IACS member class society.", action="Correct HVPQ 1.5.1.2 or verify class society details.")
    elif not iacs_sec:
        add_finding(findings, area="Class", check="IACS membership declaration", status="MANUAL CHECK", risk="MEDIUM", hvpq_value="Not reliably extracted", reason="HVPQ 1.5.1.2 could not be reliably checked. Rule requires IACS member response to be Yes.", action="Vessel/office to confirm HVPQ 1.5.1.2.")

    last_dd=parse_date_any(first_field(fields,"HVPQ","surveys.last_drydock"))
    last_sp=parse_date_any(first_field(fields,"HVPQ","surveys.last_special"))
    if last_dd:
        if last_dd + relativedelta(years=5) < ref_date:
            add_finding(findings, area="Class / Survey", check="Last dry dock age", status="MISMATCH", risk="HIGH", hvpq_value=last_dd.isoformat(), reason="HVPQ 1.5.4.1 last dry dock appears older than 5 years.", action="Verify dry dock/Class Status and update HVPQ if stale.")
        if last_sp and abs((last_dd-last_sp).days)>14:
            add_finding(findings, area="Class / Survey", check="Last dry dock vs last special survey date", status="MANUAL CHECK", risk="MEDIUM", hvpq_value=f"Dry dock {last_dd.isoformat()} / special {last_sp.isoformat()}", reason="Rule expects HVPQ 1.5.4.1 to align with 1.5.6.1 where renewal/special survey was completed at drydock; dates differ beyond tolerance.", action="Confirm whether dates should match based on latest Class Status and update stale HVPQ entry if required.")
    else:
        add_finding(findings, area="Class / Survey", check="Last dry dock date", status="MANUAL CHECK", risk="MEDIUM", hvpq_value="Not reliably extracted", reason="HVPQ 1.5.4.1 could not be reliably checked.", action="Verify last dry dock date against Class Status.")

    foam_dt=None
    foam_sec=first_field(fields,"HVPQ","section.5.3.1") or ""
    dates=all_dates_in_text(foam_sec)
    if dates: foam_dt=max(dates)
    if foam_dt:
        if foam_dt + relativedelta(months=12) < ref_date:
            add_finding(findings, area="Firefighting", check="Foam test / supply date", status="MISMATCH", risk="HIGH", hvpq_value=foam_dt.isoformat(), reason="HVPQ 5.3.1.4 foam supply/test analysis date appears older than 1 year.", action="Verify latest foam test/supply certificate and update HVPQ.")
    else:
        add_finding(findings, area="Firefighting", check="Foam test / supply date", status="MANUAL CHECK", risk="MEDIUM", hvpq_value="Not reliably extracted", reason="HVPQ 5.3.1.4 could not be reliably checked from extracted text.", action="Vessel to confirm latest foam supply/test analysis date is within 1 year.")

    # Tank coating inspection dates: table-aware check. Uses the last inspection date column, not original coating date.
    for fid, q, label in [("tank.cargo_hvpq.last_inspection_dates", "7.1.1", "Cargo/slop tank coating inspection dates"), ("tank.ballast_hvpq.last_inspection_dates", "7.1.3", "Ballast tank coating inspection dates")]:
        oldest = parse_date_any(first_field(fields, "HVPQ", fid + ".oldest"))
        count = first_field(fields, "HVPQ", fid + ".count")
        if oldest:
            due = oldest + relativedelta(months=12)
            if due < ref_date:
                add_finding(findings, area="Tank inspection", check=label, status="MISMATCH", risk="HIGH", hvpq_value=f"Oldest inspection date {oldest.isoformat()} from {count or '?'} parsed entries", reason=f"HVPQ {q} last coating inspection date appears outside the 12-month annual frequency. The check used the inspection-date column, not the original coating date.", action="Verify every tank entry and update HVPQ if any coating inspection is overdue/stale.")
        else:
            if first_field(fields, "HVPQ", f"section.{q}"):
                add_finding(findings, area="Tank inspection", check=label, status="MANUAL CHECK", risk="MEDIUM", hvpq_value="Table found, last-inspection column not reliably parsed", reason=f"HVPQ {q} table was located, but the last coating inspection dates could not be reliably separated from original coating dates. Manual check is preferred.", action="Vessel to verify the last coating inspection date column against the stated frequency.")
            else:
                add_finding(findings, area="Tank inspection", check=label, status="MANUAL CHECK", risk="MEDIUM", hvpq_value="Section not reliably extracted", reason=f"HVPQ {q} was not reliably located for tank coating inspection validation.", action="Vessel to verify all coating inspection dates against stated frequency.")

    anode_sec=section_text_by_qid(hvpq_raw, "7.1.4")
    if anode_sec and not re.search(r"\b([1-9][0-9]?(?:\.\d+)?)\s*%", anode_sec):
        add_finding(findings, area="Tank inspection", check="Ballast tank anode wastage percentage", status="MANUAL CHECK", risk="MEDIUM", hvpq_value=anode_sec[:350], reason="HVPQ 7.1.4.5 should contain wastage greater than 0 with % unit where anodes are applicable; extraction did not show a clear percentage.", action="Vessel to verify anode entries and correct HVPQ if blank/zero/wrong unit.")

    pump_sec=section_text_by_qid(hvpq_raw, "9.6.2")
    if not pump_sec:
        add_finding(findings, area="Cargo systems", check="Cargo pump type vs vessel length", status="MANUAL CHECK", risk="MEDIUM", hvpq_value="HVPQ 9.6.2 not reliably extracted", reason="Comparison rule for pump type could not be reliably checked.", action="Verify HVPQ 9.6.2 manually against vessel length and cargo pump arrangement.")

    lift_sec=section_text_by_qid(hvpq_raw, "10.9.1")
    if lift_sec:
        dates=all_dates_in_text(lift_sec)
        if dates:
            newest=max(dates)
            if newest + relativedelta(months=12) < ref_date:
                add_finding(findings, area="Lifting gear", check="Annual lifting gear test date", status="MISMATCH", risk="HIGH", hvpq_value=newest.isoformat(), reason="Latest visible lifting gear test date appears older than annual requirement; table may need detailed checking.", action="Vessel to verify annual and 5-year lifting gear/crane tests and update HVPQ 10.9.1.")
        else:
            add_finding(findings, area="Lifting gear", check="Annual and 5-year lifting gear test dates", status="MANUAL CHECK", risk="MEDIUM", hvpq_value="Not reliably extracted", reason="HVPQ 10.9.1 could not be reliably checked.", action="Vessel to confirm annual and 5-year lifting gear test dates are current.")
    else:
        add_finding(findings, area="Lifting gear", check="Annual and 5-year lifting gear test dates", status="MANUAL CHECK", risk="MEDIUM", hvpq_value="Section not reliably extracted", reason="HVPQ 10.9.1 was not reliably located in extraction.", action="Vessel to verify lifting gear/crane entries.")

    static=normalize_bool(first_field(fields,"PIQ","piq.static_nav_assessment"))
    dyn=normalize_bool(first_field(fields,"PIQ","piq.dynamic_nav_assessment_shore"))
    if static and dyn and static==dyn:
        add_finding(findings, area="PIQ Navigational Assessment", check="Static and dynamic assessment responses", status="MISMATCH", risk="HIGH", piq_value=f"Static {static}; dynamic {dyn}", reason="PIQ 3.2.1 and 3.2.2 should not both be Yes or both be No as per provided rule.", action="Verify PIQ navigation assessment responses and dates.")
    elif not static or not dyn:
        add_finding(findings, area="PIQ Navigational Assessment", check="Static and dynamic assessment responses", status="MANUAL CHECK", risk="MEDIUM", piq_value=f"Static {static or 'not extracted'}; dynamic {dyn or 'not extracted'}", reason="PIQ 3.2.1/3.2.2 could not be fully extracted for rule check.", action="Verify static/dynamic navigational assessment responses and last assessment date.")

    for fid,label,q in [("piq.cargo_audit","Cargo audit","PIQ 3.2.5"),("piq.engineering_audit","Engineering audit","PIQ 3.2.6"),("piq.mooring_anchoring_audit","Mooring/anchoring audit","PIQ 3.2.7")]:
        v=normalize_bool(first_field(fields,"PIQ",fid))
        raw=next((f.raw for f in fields if f.source=="PIQ" and f.field_id==fid),"")
        dates=all_dates_in_text(raw)
        if v=="yes" and dates:
            latest=max(dates)
            if latest + relativedelta(months=12) < ref_date:
                add_finding(findings, area="PIQ Audit", check=f"{label} date", status="MISMATCH", risk="HIGH", piq_value=latest.isoformat(), reason=f"{q} is marked Yes but latest visible date appears older than 12 months.", action="Verify latest audit date and update PIQ.")
        elif v=="yes" and not dates:
            add_finding(findings, area="PIQ Audit", check=f"{label} date", status="MANUAL CHECK", risk="MEDIUM", piq_value="Yes, date not reliably extracted", reason=f"{q} is marked Yes but date could not be reliably checked.", action="Verify audit date is within 12 months.")

    return dedupe_findings(findings)


def build_hvpq_checks_v15(fields: List[FieldRecord], findings: List[Finding], ref_date: date, obs_qids: set) -> pd.DataFrame:
    rows=[]; seen=set()
    def add_row(priority,qno,area,check,status,hvpq_val,ref_source,ref_val,interp,action):
        check=concise_check_name(check)
        obs_reason=obs_reason_for(qno, obs_qids)
        if obs_reason and obs_reason not in interp:
            interp=(interp + " " + obs_reason).strip()
        key=(priority,qno,area,check,status,hvpq_val,ref_val)
        if key in seen: return
        seen.add(key)
        rows.append({"Priority":priority,"Question / Section":qno,"Area":area,"Check":check,"Status":status,"HVPQ value":hvpq_val,"Reference source":ref_source,"Reference value":ref_val,"Finding / interpretation":interp,"Action requested":action})
    for f in findings:
        if f.risk.upper() not in ["CRITICAL","HIGH","MEDIUM"]: continue
        if f.hvpq_value or f.class_value or f.area in ["Certificates","Class / Survey","Classification","Blank / Missing","Environment","Ownership / Operation","Ownership","Insurance","Vessel Type","Class","Tank inspection","Firefighting","Cargo systems","Lifting gear"]:
            add_row(f.risk, qno_for_finding(f), f.area, f.check, f.status, f.hvpq_value, "Class Status" if f.class_value else ("Q88" if f.q88_value else "PIQ/manual"), f.class_value or f.q88_value or f.piq_value, f.reason, f.action)
    for r in cert_validity_rows(fields, ref_date):
        add_row(r["Priority"], r["Question / Section"], r["Area"], r["Check"], r["Status"], r["Document value"], "Class Status/latest certificate", r["Reference value"], r["Finding / interpretation"], r["Action requested"])
    for row in [_hvpq_ops_row(fields, ref_date, "Brake testing", "mooring.brake_test_date", 12, "10.1.4", "Mooring", "Latest brake test date found in HVPQ/Q88 text"), _hvpq_ops_row(fields, ref_date, "Mooring ropes age / visible date", "mooring.ropes.latest_visible_date", 60, "10.1.7", "Mooring", "Latest rope/tail visible date found; verify every rope individually")]:
        if row["Priority"] != "Manual":
            add_row(row["Priority"], row["Question / Section"], row["Area"], row["Check"], row["Status"], row["HVPQ value"], row["Reference source"], row["Reference value"], row["Finding / interpretation"], row["Action requested"])
    # Positive tank coating checks: only show if extraction is reliable and no overdue issue was found.
    for fid, qno, label in [("tank.cargo_hvpq.last_inspection_dates", "7.1.1", "Cargo/slop tank coating inspection dates"), ("tank.ballast_hvpq.last_inspection_dates", "7.1.3", "Ballast tank coating inspection dates")]:
        oldest = parse_date_any(first_field(fields, "HVPQ", fid + ".oldest"))
        count = first_field(fields, "HVPQ", fid + ".count")
        if oldest and oldest + relativedelta(months=12) >= ref_date:
            add_row("OK", qno, "Tank inspection", label, "In order", f"Oldest last-inspection date {oldest.isoformat()} from {count or '?'} parsed entries", "HVPQ table", "Annual frequency", "The HVPQ tank coating table was read using the last-inspection-date column, and the oldest parsed inspection date is within 12 months.", "No action required unless vessel knows any individual tank inspection is missing from HVPQ.")

    # Observation-library question numbers are deliberately not converted into standalone
    # manual checks. They are used only to strengthen the reason for an actual HVPQ
    # issue/blank/mismatch already detected above. This avoids discouraging users with
    # hundreds of generic repeat-observation rows when the extracted HVPQ entry appears
    # present and not doubtful.
    cols=["Priority","Question / Section","Area","Check","Status","HVPQ value","Reference source","Reference value","Finding / interpretation","Action requested"]
    df=pd.DataFrame(rows)
    if df.empty: return pd.DataFrame(columns=cols)
    df=df[cols].drop_duplicates()
    df["_rank"]=df["Priority"].map(lambda x: {"Critical":0,"CRITICAL":0,"High":1,"HIGH":1,"Medium":2,"MEDIUM":2,"Manual":3,"OK":4}.get(str(x),5))
    return df.sort_values(["_rank","Area","Question / Section"]).drop(columns="_rank")


def build_piq_checks_v15(fields: List[FieldRecord], findings: List[Finding], ref_date: date) -> pd.DataFrame:
    rows=[]
    base=build_piq_checks(fields, findings, ref_date)
    if not base.empty:
        for _,r in base.iterrows():
            d=r.to_dict()
            d["Area"]=str(d.get("Area","")).replace("PIQ completeness","PIQ declaration")
            d["Check"]=concise_check_name(str(d.get("Check","")))
            interp=str(d.get("Finding / interpretation",""))
            interp=re.sub(r"PIQ value was extracted for review\.?", "The PIQ entry was extracted and included in the review.", interp, flags=re.I)
            d["Finding / interpretation"]=interp
            rows.append(d)
    if not any(str(r.get("Question / Section","")).startswith("PIQ 5.7") for r in rows):
        inc_vals=[f.value for f in fields if f.source=="PIQ" and f.field_id.startswith("incidents.")]
        status="In order / positive confirmation" if inc_vals and all(normalize_bool(v)=="no" for v in inc_vals if v) else "Manual confirmation"
        rows.append({"Priority":"Manual","Question / Section":"PIQ 5.7.1001-1029","Area":"Incident declarations","Check":"Incident declaration alignment","Status":status,"Document value":"; ".join(inc_vals[:8]) or "Not reliably extracted","Reference value":"HVPQ 1.9.1-1.9.7","Finding / interpretation":"Incident-related PIQ answers should align with HVPQ 1.9.1-1.9.7 and any Class survey purpose indicating damage/repair/occasional survey.","Action requested":"Vessel/office to positively confirm there were no reportable machinery, injury, mooring, pollution, security, navigation or operational incidents omitted from PIQ/HVPQ."})
    df=pd.DataFrame(rows)
    cols=["Priority","Question / Section","Area","Check","Status","Document value","Reference value","Finding / interpretation","Action requested"]
    for c in cols:
        if c not in df.columns: df[c]=""
    df=df[cols].drop_duplicates()
    df["_rank"]=df["Priority"].map(lambda x: {"Critical":0,"CRITICAL":0,"High":1,"HIGH":1,"Medium":2,"MEDIUM":2,"Manual":3,"OK":4}.get(str(x),5))
    return df.sort_values(["_rank","Area","Question / Section"]).drop(columns="_rank")


def build_q88_checks_v15(fields: List[FieldRecord], findings: List[Finding]) -> pd.DataFrame:
    df=build_q88_checks(fields, findings)
    if df.empty: return df
    df=df.copy()
    df["Area"]=df["Area"].astype(str).str.replace("Q88 completeness","Q88 value-add", regex=False)
    df["Check"]=df["Check"].map(concise_check_name)
    df["Finding / interpretation"]=df["Finding / interpretation"].astype(str).str.replace("Q88 value was extracted for cross-check.", "The Q88 entry was extracted and cross-checked as value-add information.", regex=False)
    return df


def build_summary_paragraphs_v15(hvpq_df: pd.DataFrame, q88_df: pd.DataFrame, piq_df: pd.DataFrame) -> Dict[str,str]:
    def list_items(df, priorities, limit=10):
        if df is None or df.empty or "Priority" not in df.columns: return []
        mask=df["Priority"].astype(str).str.upper().isin([p.upper() for p in priorities])
        vals=[]
        for _,r in df[mask].head(limit).iterrows():
            chk=clean_text(r.get("Check", r.get("Area","")))
            q=clean_text(r.get("Question / Section",""))
            vals.append(f"{q} {chk}".strip())
        return vals
    ok=list_items(hvpq_df,["OK"],8)+list_items(piq_df,["OK"],8)+list_items(q88_df,["OK"],4)
    bad=list_items(hvpq_df,["CRITICAL","HIGH","MEDIUM"],8)+list_items(piq_df,["CRITICAL","HIGH","MEDIUM"],8)+list_items(q88_df,["CRITICAL","HIGH","MEDIUM"],5)
    manual=list_items(hvpq_df,["MANUAL"],8)+list_items(piq_df,["MANUAL"],8)+list_items(q88_df,["MANUAL"],5)
    ok_text="; ".join(ok) if ok else "No positive conclusion is shown for areas where extraction was not strong enough; those items are carried as manual confirmations instead."
    bad_text="; ".join(bad) if bad else "No critical/high mismatch was identified from the mapped checks."
    manual_text="; ".join(manual) if manual else "No major manual confirmation gap was generated."
    return {
        "checked": "The review checked HVPQ as the main correction document, used Class Status only as the reference for certificate/survey dates and Conditions/Memoranda/dispensations, checked Q88 as value-add information, and reviewed PIQ operational declarations, superintendent intervals, tank inspection cycles, PSC, MOC and incident declarations.",
        "ok": "From the extracted data, the following checks appear in order: " + ok_text,
        "bad": "The following items need correction or review: " + bad_text,
        "manual": "The following items could not be confirmed reliably and are included for manual confirmation: " + manual_text,
    }


def make_excel_v15(hvpq_df: pd.DataFrame, q88_df: pd.DataFrame, piq_df: pd.DataFrame) -> bytes:
    bio=io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        sheets=[("HVPQ Checks", hvpq_df), ("Q88 Value Add", q88_df), ("PIQ Checks", piq_df)]
        for name, df in sheets:
            df.to_excel(writer, index=False, sheet_name=name)
        for ws in writer.book.worksheets:
            ws.freeze_panes="A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font=Font(bold=True, color="FFFFFF")
                cell.fill=PatternFill("solid", fgColor="1F4E78")
                cell.alignment=Alignment(wrap_text=True, vertical="center")
            for row in ws.iter_rows(min_row=2):
                pr = str(row[0].value).upper() if row and row[0].value is not None else ""
                fill = None
                if "CRITICAL" in pr or "HIGH" in pr:
                    fill=PatternFill("solid", fgColor="FCE4D6")
                elif "MEDIUM" in pr or "MANUAL" in pr:
                    fill=PatternFill("solid", fgColor="EAF2F8")
                elif "OK" in pr:
                    fill=PatternFill("solid", fgColor="E2F0D9")
                for cell in row:
                    cell.alignment=Alignment(wrap_text=True, vertical="top")
                    if fill: cell.fill=fill
            widths={"A":13,"B":18,"C":24,"D":30,"E":20,"F":44,"G":28,"H":44,"I":60,"J":58}
            for idx, col in enumerate(ws.columns, start=1):
                letter=get_column_letter(idx)
                max_len=max((len(str(c.value)) if c.value is not None else 0) for c in col)
                ws.column_dimensions[letter].width=min(max(widths.get(letter, 14), min(max_len+2, 55)), 65)
            for r in range(2, min(ws.max_row, 500)+1):
                ws.row_dimensions[r].height=62
    return bio.getvalue()

# ----------------------------- Streamlit app -----------------------------

def main():
    st.set_page_config(page_title=APP_TITLE, layout="wide")
    st.title(APP_TITLE)
    st.caption(APP_SUBTITLE)

    with st.sidebar:
        st.header("Upload documents")
        hvpq_file = st.file_uploader("HVPQ PDF — main document to correct", type=["pdf"], key="hvpq")
        hvpq_xml = st.file_uploader("HVPQ XML (optional)", type=["xml"], key="xml")
        piq_file = st.file_uploader("PIQ PDF", type=["pdf"], key="piq")
        q88_file = st.file_uploader("Q88 PDF — value-add cross-check", type=["pdf"], key="q88")
        class_file = st.file_uploader("Class Status PDF — certificate/survey authority", type=["pdf"], key="class")
        st.caption("Built-in rule base active: observation history, incident checks and comparison rules are embedded in the app.")
        st.divider()
        ref_date_input = st.date_input("Reference / review date", value=date.today())
        show_low = st.checkbox("Show low-risk findings", value=False)
        use_llm = st.checkbox("Use local LLM extraction assist (Ollama)", value=False)
        ollama_url = st.text_input("Ollama URL", value="http://localhost:11434", disabled=not use_llm)
        ollama_model = st.text_input("Ollama model", value="qwen2.5:14b", disabled=not use_llm)
        run_btn = st.button("Run checks", type="primary")

    if not run_btn:
        st.info("Upload HVPQ, PIQ, Q88 and Class Status where available, then click **Run checks**. Observation history and comparison rules are built in; no separate rule/observation upload is required.")
        return

    settings = {"show_low": show_low}
    all_fields: List[FieldRecord] = []
    page_cache = {}

    with st.spinner("Extracting structured data and building source-specific registers..."):
        if hvpq_file:
            pages = extract_pdf_pages(hvpq_file); page_cache["HVPQ"] = pages
            all_fields += extract_hvpq(pages)
            if use_llm: all_fields += llm_assist_extract("HVPQ", pages, ollama_url, ollama_model)
        if piq_file:
            pages = extract_pdf_pages(piq_file); page_cache["PIQ"] = pages
            all_fields += extract_piq(pages)
            if use_llm: all_fields += llm_assist_extract("PIQ", pages, ollama_url, ollama_model)
        if q88_file:
            pages = extract_pdf_pages(q88_file); page_cache["Q88"] = pages
            all_fields += extract_q88(pages)
            if use_llm: all_fields += llm_assist_extract("Q88", pages, ollama_url, ollama_model)
        if class_file:
            pages = extract_pdf_pages(class_file); page_cache["CLASS"] = pages
            all_fields += extract_class_status(pages)
            if use_llm: all_fields += llm_assist_extract("CLASS", pages, ollama_url, ollama_model)
        if hvpq_xml:
            all_fields += extract_xml(hvpq_xml)
        obs_df = embedded_observation_df()
        all_fields = dedupe_fields(all_fields)

    findings = run_rules(all_fields, ref_date_input, settings, obs_df)
    findings = add_v15_validation_findings(findings, all_fields, ref_date_input)
    if not show_low:
        findings = [f for f in findings if f.risk.upper() != "LOW"]
    hvpq_text = raw_text_for_source(page_cache, "HVPQ")
    obs_qid_df = hvpq_qid_status_df(obs_df, hvpq_text)
    obs_qids = extract_qids_from_obs(obs_df)
    hvpq_checks_df = build_hvpq_checks_v15(all_fields, findings, ref_date_input, obs_qids)
    q88_checks_df = build_q88_checks_v15(all_fields, findings)
    piq_checks_df = build_piq_checks_v15(all_fields, findings, ref_date_input)
    summary = build_summary_paragraphs_v15(hvpq_checks_df, q88_checks_df, piq_checks_df)

    # Summary metrics
    def count_bad(df):
        if df is None or df.empty or "Priority" not in df.columns: return 0
        return int(df["Priority"].astype(str).str.contains("Critical|High|Medium|Manual", case=False, na=False).sum())
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("HVPQ checks", len(hvpq_checks_df))
    c2.metric("Q88 checks", len(q88_checks_df))
    c3.metric("PIQ checks", len(piq_checks_df))
    manual_count = int((hvpq_checks_df.get('Priority', pd.Series(dtype=str)).astype(str).str.upper()=='MANUAL').sum() + (q88_checks_df.get('Priority', pd.Series(dtype=str)).astype(str).str.upper()=='MANUAL').sum() + (piq_checks_df.get('Priority', pd.Series(dtype=str)).astype(str).str.upper()=='MANUAL').sum())
    c4.metric("Manual checks", manual_count)

    st.subheader("Office review summary")
    st.markdown(summary["checked"])
    st.success(summary["ok"])
    if count_bad(hvpq_checks_df)+count_bad(q88_checks_df)+count_bad(piq_checks_df):
        st.warning(summary["bad"])
    else:
        st.success(summary["bad"])
    st.info(summary["manual"])

    xlsx = make_excel_v15(hvpq_checks_df, q88_checks_df, piq_checks_df)
    st.download_button("Download Excel check register", xlsx, file_name="hvpq_piq_q88_check_register_v17.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    tabs = st.tabs(["HVPQ Checks", "Q88 Value Add", "PIQ Checks", "Built-in Rule Base", "Advanced Extraction"])
    with tabs[0]:
        st.subheader("HVPQ checks — main correction register")
        st.caption("HVPQ is the document to correct. Class Status is used as authority for certificate/survey dates; Q88 is only value-add.")
        st.dataframe(style_priority_dataframe(hvpq_checks_df), use_container_width=True, height=680)
    with tabs[1]:
        st.subheader("Q88 checks — value-add consistency")
        st.caption("Q88 is not authoritative by itself. Use this to identify Q88/HVPQ blanks or mismatches requiring review.")
        st.dataframe(style_priority_dataframe(q88_checks_df), use_container_width=True, height=680)
    with tabs[2]:
        st.subheader("PIQ checks — operational declarations and intervals")
        st.caption("Includes PIQ superintendent intervals, tank inspection cycles, MOC/retrofit, PSC, incidents and key PIQ rule checks.")
        st.dataframe(style_priority_dataframe(piq_checks_df), use_container_width=True, height=680)
    with tabs[3]:
        st.subheader("Built-in machine-readable rule base")
        st.caption("Rules and repeated-observation priorities are bundled inside the app. Observation rows are sorted by repeat count and used only as priority signals, not standalone defects.")
        kb1, kb2 = st.tabs(["Validation rules", "Observation priorities"])
        with kb1:
            st.dataframe(embedded_validation_rules_df(), use_container_width=True, height=650)
        with kb2:
            st.dataframe(embedded_observation_rules_df(), use_container_width=True, height=650)
    with tabs[4]:
        st.subheader("Advanced extraction review")
        st.caption("For troubleshooting extraction only. This is not intended as the main user workflow.")
        st.dataframe(df_from_fields(all_fields), use_container_width=True, height=600)
        if st.checkbox("Show raw detailed findings"):
            st.dataframe(df_from_findings(findings), use_container_width=True, height=600)


if __name__ == "__main__":
    main()
